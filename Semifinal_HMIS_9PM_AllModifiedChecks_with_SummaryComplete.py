import sys, re
import numpy as np
import pandas as pd
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QFileDialog, QMainWindow, QMessageBox, QWidgetAction
from PyQt5.QtCore import Qt
import matplotlib.pyplot as plt



########################################### PandasModel to display it on tableView #####################

class PandasModel(QtCore.QAbstractTableModel):
    def __init__(self, df=pd.DataFrame(), parent=None):
        QtCore.QAbstractTableModel.__init__(self, parent=parent)
        self._df = df.copy()

    def toDataFrame(self):
        return self._df.copy()

    def headerData(self, section, orientation, role=QtCore.Qt.DisplayRole):
        if role != QtCore.Qt.DisplayRole:
            return QtCore.QVariant()

        if orientation == QtCore.Qt.Horizontal:
            try:
                return self._df.columns.tolist()[section]
            except (IndexError, ):
                return QtCore.QVariant()
        elif orientation == QtCore.Qt.Vertical:
            try:
                # return self.df.index.tolist()
                return self._df.index.tolist()[section]
            except (IndexError, ):
                return QtCore.QVariant()

    def data(self, index, role=QtCore.Qt.DisplayRole):
        if role != QtCore.Qt.DisplayRole:
            return QtCore.QVariant()

        if not index.isValid():
            return QtCore.QVariant()

        return QtCore.QVariant(str(self._df.iloc[index.row(), index.column()]))

    def setData(self, index, value, role):
        row = self._df.index[index.row()]
        col = self._df.columns[index.column()]
        if hasattr(value, 'toPyObject'):
            # PyQt4 gets a QVariant
            value = value.toPyObject()
        else:
            # PySide gets an unicode
            dtype = self._df[col].dtype
            if dtype != object:
                value = None if value == '' else dtype.type(value)
        self._df.set_value(row, col, value)
        return True

    def rowCount(self, parent=QtCore.QModelIndex()):
        return len(self._df.index)

    def columnCount(self, parent=QtCore.QModelIndex()):
        return len(self._df.columns)

    def sort(self, column, order):
        colname = self._df.columns.tolist()[column]
        self.layoutAboutToBeChanged.emit()
        self._df.sort_values(colname, ascending=order ==
                             QtCore.Qt.AscendingOrder, inplace=True)
        self._df.reset_index(inplace=True, drop=True)
        self.layoutChanged.emit()


class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(1347, 859)
        self.pushButton = QtWidgets.QPushButton(Dialog)
        self.pushButton.setGeometry(QtCore.QRect(30, 20, 251, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.upload)
        self.pushButton_2 = QtWidgets.QPushButton(Dialog)
        self.pushButton_2.setGeometry(QtCore.QRect(30, 160, 251, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)

        # 2
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.clicked.connect(self.VerifyFType)
        self.comboBox = QtWidgets.QComboBox(Dialog)
        
        self.comboBox.setGeometry(QtCore.QRect(30, 100, 251, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.comboBox.setFont(font)
        self.comboBox.setObjectName("comboBox")
    
        self.label = QtWidgets.QLabel(Dialog)
        self.label.setGeometry(QtCore.QRect(30, 60, 251, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(Dialog)
        self.label_2.setGeometry(QtCore.QRect(30, 210, 241, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(Dialog)
        self.label_3.setGeometry(QtCore.QRect(30, 320, 251, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.pushButton_3 = QtWidgets.QPushButton(Dialog)
        self.pushButton_3.clicked.connect(self.export)
        self.pushButton_3.setGeometry(QtCore.QRect(30, 740, 251, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setObjectName("pushButton_3")
        self.tableView = QtWidgets.QTableView(Dialog)
        self.tableView.setGeometry(QtCore.QRect(320, 20, 1011, 821))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.tableView.setFont(font)
        self.tableView.setObjectName("tableView")
        self.label_4 = QtWidgets.QLabel(Dialog)
        self.label_4.setGeometry(QtCore.QRect(30, 430, 251, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(Dialog)
        self.label_5.setGeometry(QtCore.QRect(30, 530, 251, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(Dialog)
        self.label_6.setGeometry(QtCore.QRect(30, 630, 251, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")

        # 4
        self.pushButton_4 = QtWidgets.QPushButton(Dialog)
        self.pushButton_4.installEventFilter(Dialog)
        self.pushButton_4.clicked.connect(self.onSelectState)

        self.pushButton_4.setGeometry(QtCore.QRect(30, 260, 251, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setObjectName("pushButton_4")

        # 5
        self.pushButton_5 = QtWidgets.QPushButton(Dialog)
        self.pushButton_5.installEventFilter(Dialog)
        self.pushButton_5.clicked.connect(self.onSelectDistrict)

        self.pushButton_5.setGeometry(QtCore.QRect(30, 360, 251, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton_5.setFont(font)
        self.pushButton_5.setObjectName("pushButton_5")

        # 6
        self.pushButton_6 = QtWidgets.QPushButton(Dialog)
        self.pushButton_6.installEventFilter(Dialog)
        self.pushButton_6.clicked.connect(self.onSelectFacilityName)

        self.pushButton_6.setGeometry(QtCore.QRect(30, 470, 251, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton_6.setFont(font)
        self.pushButton_6.setObjectName("pushButton_6")

        # 7
        self.pushButton_7 = QtWidgets.QPushButton(Dialog)
        self.pushButton_7.installEventFilter(Dialog)
        self.pushButton_7.clicked.connect(self.onSelectMonth)

        self.pushButton_7.setGeometry(QtCore.QRect(30, 570, 251, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.pushButton_7.setFont(font)
        self.pushButton_7.setObjectName("pushButton_7")

        # 8
        self.pushButton_8 = QtWidgets.QPushButton(Dialog)
        self.pushButton_8.installEventFilter(Dialog)
        self.pushButton_8.clicked.connect(self.onSelectYear)

        self.pushButton_8.setGeometry(QtCore.QRect(30, 670, 251, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(False)
        font.setWeight(50)
        self.pushButton_8.setFont(font)
        self.pushButton_8.setObjectName("pushButton_8")

        # 9
        self.pushButton_9 = QtWidgets.QPushButton(Dialog)
        self.pushButton_9.clicked.connect(self.reset)
        self.pushButton_9.setGeometry(QtCore.QRect(30, 800, 251, 41))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_9.setFont(font)
        self.pushButton_9.setObjectName("pushButton_9")

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Dialog"))
        self.pushButton.setText(_translate("Dialog", "UPLOAD "))
        self.pushButton_2.setText(_translate("Dialog", "VALIDATE "))
        self.label.setText(_translate("Dialog", "Facility Type"))
        self.label_2.setText(_translate("Dialog", "Filter State"))
        self.label_3.setText(_translate("Dialog", "Filter District"))
        self.pushButton_3.setText(_translate("Dialog", "Export CSV"))
        self.label_4.setText(_translate("Dialog", "Filter Facility Name"))
        self.label_5.setText(_translate("Dialog", "Filter Month"))
        self.label_6.setText(_translate("Dialog", "Filter Year"))
        self.pushButton_4.setText(_translate("Dialog", "- All Selected -"))
        self.pushButton_5.setText(_translate("Dialog", "- All Selected -"))
        self.pushButton_6.setText(_translate("Dialog", "- All Selected -"))
        self.pushButton_7.setText(_translate("Dialog", "- All Selected -"))
        self.pushButton_8.setText(_translate("Dialog", "- All Selected -"))
        self.pushButton_9.setText(_translate("Dialog", "Reset"))

    def upload(self):
        global df_, res_dict

        # Validation for uploaded valid excel file
        try:
            # Upload file by opening filedialog
            fileName, _ = QFileDialog.getOpenFileName(
                Dialog, "Open Excel", (QtCore.QDir.homePath()), "Excel (*.xls *.xlsx)")

            # Read uploaded excel file
            df_ = pd.read_excel(fileName)

            # Converted again to csv file
            df_.to_csv("FileName.csv")

            # Read converted csv file
            df_ = pd.read_csv("FileName.csv")
        except:
            msg = QMessageBox()
            msg.setWindowTitle("Uploaded File Error Message")
            msg.setText(
                "The file which you have uploaded is not in the valid format of excel, Please upload valid excel file")
            msg.exec()

            try:
                # Upload file by opening filedialog
                fileName, _ = QFileDialog.getOpenFileName(
                    Dialog, "Open Excel", (QtCore.QDir.homePath()), "Excel (*.xls *.xlsx)")

                # Read uploaded excel file
                df_ = pd.read_excel(fileName)

                # Converted again to csv file
                df_.to_csv("FileName.csv")

                # Read converted csv file
                df_ = pd.read_csv("FileName.csv")
            except:
                msg = QMessageBox()
                msg.setWindowTitle("Uploaded File Error Message")
                msg.setText("Please upload valid excel file, Try again!")
                msg.exec()
                self.close()

        # Dropping last two rows
        df_.drop(df_.index[[-1, -2]], inplace=True)

        # Extracting string from 1st cell of dataframe
        str_to_extr_MonthYear = str(df_.iloc[0])

        # grab the first row for the header
        new_header = df_.iloc[1]

        # #take the data less the header row
        df_ = df_[1:]

        # set the header row as the df header
        df_.columns = new_header

        # Extracting Month , Year from string
        results = re.findall(
            r"[abceglnoprtuvyADFJMNOS|]{3}[\s-]\d{2,4}", str_to_extr_MonthYear)
        print(results)

        # Splitting Month and Year
        MYList = results[0].split('-')
        print(MYList)

        # Partial list of headers
        lst1 = df_.columns[:16].values

        # Picking row items after 18th row to merge with lst1
        lst2 = df_.iloc[1, 16:].to_numpy()

        # Merging both lists
        lst3 = np.concatenate((lst1, lst2))

        # Assign lst3 as new column header
        df_.columns = lst3

        # Taking DataFrame from second row
        df_ = df_[2:]

        # Insering Month and Year to the orignal dataframe
        df_.insert(1, 'Month', MYList[0])
        df_.insert(2, 'Year', MYList[1])

        # Removing A column named as # coming from orignal data
        df_ = df_.loc[:, df_.columns != '#']

        # Reindexing dataframe
        df_ = df_.reset_index(drop=True)

        df_ = df_.iloc[:, 1:]

        # Temporary column to verify modified checks
        temp_columns = ['col_' + str(index)
                        for index in range(1, len(df_.columns)+1)]

        # Merging and converting temp_columns to orignal header to dictionary
        res_dict = {temp_columns[i]: df_.columns[i]
                    for i in range(len(temp_columns))}

        # Picking the temporary column names and renaming column headers with it
        df_.columns = [i for i in res_dict.keys()]

        # Orignal Header
        df_OrgHeaders = [i for i in res_dict.values()]

        # convert the set to the list and fill inside comboBox to select facility type
        list_set = df_['col_12'].tolist()
        unique_list = set(list_set)
        print(unique_list)

        self.comboBox.addItems(['-select-'])
        self.comboBox.addItems(["{0}".format(col) for col in unique_list])

        # Displaying uploaded dataframe
        self.tableView.setModel(PandasModel(df_))

        # Disabling upload Button
        self.pushButton.setDisabled(True)
        return df_

    
    # Upload file button functionality
    def loadFile(self, df_):
        return df_


    # Filtering Facility Type
    def VerifyFType(self):
        global df, FType

        df = self.loadFile(df_)

        FType = self.comboBox.currentText()
        print(FType)

        if (FType == 'Primary Health Centre'):
            print('Facility Type - ',FType)

            # Signaling PHC_Validate function i.e function where validation checks are present
            self.PHC_Validate()

        elif (FType == 'Health Sub Centre' ):
            print('Facility Type - ', FType)
            
            # Signaling HSC_Validate function i.e function where validation checks are present
            self.HSC_Validate()

        elif (FType == 'District Hospital'):
            print('Facility Type - ',FType)
            self.DH_Validate()

        elif (FType == 'Sub District Hospital'):
            print('Facility Type - ',FType)
            self.SDH_Validate()

        elif (FType == 'Community Health Centre'):
            print('Facility Type - ',FType)
            self.CHC_Validate()

        else:
            raise Exception('Facility Type Name is not matching')


    # District Hospital Validation Rules Function
    # ===============================================
    def DH_Validate(self):
        global df

        df = self.loadFile(df_)
        filterString = self.comboBox.currentText()
        df = df_.loc[df_['col_12'] == filterString]

        print('Entered DH_Validate')

        # Modified Checks of DH

        # 1.1.1 (22) <=1.1 (23)
        def res1(df):
            if pd.isnull(df['col_23']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_23']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_23']):
                    return 'Probable Reporting Error(1.1.1 is blank)'
                elif pd.isnull(float(df['col_22'])):
                    return 'Inconsistent'
            elif float(df['col_23']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 15.3.1.b (244)<=15.3.1.a (243)
        def res2(df):
            if pd.isnull(df['col_244']) and pd.isnull(df['col_243']):
                return 'Blank'
            elif pd.isnull(df['col_244']) or pd.isnull(df['col_243']):
                if pd.isnull(df['col_244']):
                    return 'Probable Reporting Error(15.3.1.b is blank)'
                elif pd.isnull(float(df['col_243'])):
                    return 'Inconsistent'
            elif float(df['col_244']) > float(df['col_243']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.2.4(27) <= 1.1(22)
        def res3(df):
            if pd.isnull(df['col_27']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_27']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_27']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_27'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df
        
        #1.2.5(28) <= 1.1(22)
        def res4(df):
            if pd.isnull(df['col_28']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_28']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_28']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
             # If value exists for all the elements
            else:

                lhs_value = float(df['col_28'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df   

        #1.2.7(30) <= 1.1(22)
        def res5(df):
            if pd.isnull(df['col_30']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_30']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_30']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
             # If value exists for all the elements
            else:

                lhs_value = float(df['col_30'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df    

        #1.3.1.a(33) <= 1.3.1(32)
        def res6(df):
            if pd.isnull(df['col_33']) and pd.isnull(df['col_32']):
                return 'Blank'
            elif pd.isnull(df['col_33']) or pd.isnull(df['col_32']):
                if pd.isnull(df['col_33']):
                    return 'Probable Reporting Error(1.3.1.a is blank)'
                elif pd.isnull(float(df['col_32'])):
                    return 'Inconsistent'
            elif float(df['col_33']) > float(df['col_32']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.3.2(34) <= 2.1(47)
        def res7(df):
            if pd.isnull(df['col_34']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_34']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_34']):
                    return 'Probable Reporting Error(1.3.2 is blank)'
                elif pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
            elif float(df['col_34']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #1.4.4(38) <= 1.4.3(37)
        def res8(df):
            if pd.isnull(df['col_38']) and pd.isnull(df['col_37']):
                return 'Blank'
            elif pd.isnull(df['col_38']) or pd.isnull(df['col_37']):
                if pd.isnull(df['col_38']) and not pd.isnull(float(df['col_37'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
             # If value exists for all the elements
            else:

                lhs_value = float(df['col_38'])
                rhs_value = float(df['col_37'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df  

        #1.5.1(39) <= 1.1(22)
        def res9(df):
            if pd.isnull(df['col_39']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_39']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_39']):
                    return 'Probable Reporting Error(1.5.1 is blank)'
                elif pd.isnull(float(df['col_22'])):
                    return 'Inconsistent'
            elif float(df['col_39']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.5.2(40) <= 1.5.1(39)
        def res10(df):
            if pd.isnull(df['col_40']) and pd.isnull(df['col_39']):
                return 'Blank'
            elif pd.isnull(df['col_40']) or pd.isnull(df['col_39']):
                if pd.isnull(df['col_40']):
                    return 'Probable Reporting Error(1.5.2 is blank)'
                elif pd.isnull(float(df['col_39'])):
                    return 'Inconsistent'
            elif float(df['col_40']) > float(df['col_39']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        
         #1.5.3(41) <= 1.5.2(40)
        def res11(df):
            if pd.isnull(df['col_41']) and pd.isnull(df['col_40']):
                return 'Blank'
            elif pd.isnull(df['col_41']) or pd.isnull(df['col_40']):
                if pd.isnull(df['col_41']):
                    return 'Probable Reporting Error(1.5.3 is blank)'
                elif pd.isnull(float(df['col_40'])):
                    return 'Inconsistent'
            elif float(df['col_41']) > float(df['col_40']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #1.6.1.a(42) <= 1.1(22)
        def res12(df):
            if pd.isnull(df['col_42']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_42']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_42']):
                    return 'Probable Reporting Error(1.6.1 is blank)'
                elif pd.isnull(float(df['col_22'])):
                    return 'Inconsistent'
            elif float(df['col_42']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.6.1.b(43) <= 1.6.1.a(42)
        def res13(df):
            if pd.isnull(df['col_43']) and pd.isnull(df['col_42']):
                return 'Blank'
            elif pd.isnull(df['col_43']) or pd.isnull(df['col_42']):
                if pd.isnull(df['col_43']):
                    return 'Probable Reporting Error(1.6.1.b is blank)'
                elif pd.isnull(float(df['col_42'])):
                    return 'Inconsistent'
            elif float(df['col_43']) > float(df['col_42']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.6.1.c(44) <= 1.6.1.b(43)
        def res14(df):
            if pd.isnull(df['col_44']) and pd.isnull(df['col_43']):
                return 'Blank'
            elif pd.isnull(df['col_44']) or pd.isnull(df['col_43']):
                if pd.isnull(df['col_44']):
                    return 'Probable Reporting Error(1.6.1.c is blank)'
                elif pd.isnull(float(df['col_43'])):
                    return 'Inconsistent'
            elif float(df['col_44']) > float(df['col_43']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.6.1.e(46) <= 1.6.1.d(45)
        def res15(df):
            if pd.isnull(df['col_46']) and pd.isnull(df['col_45']):
                return 'Blank'
            elif pd.isnull(df['col_46']) or pd.isnull(df['col_45']):
                if pd.isnull(df['col_46']):
                    return 'Probable Reporting Error(1.6.1.e is blank)'
                elif pd.isnull(float(df['col_45'])):
                    return 'Inconsistent'
            elif float(df['col_46']) > float(df['col_45']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #2.1.1(48) <= 2.1(47)
        def res16(df):
            if pd.isnull(df['col_48']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_48']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_48']):
                    return 'Probable Reporting Error(2.1.1 is blank)'
                elif pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
            elif float(df['col_48']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #3.1(49) <= 2.1(47)
        def res17(df):
            if pd.isnull(df['col_49']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_49']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_49']):
                    return 'Probable Reporting Error(3.1 is blank)'
                elif pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
            elif float(df['col_49']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #3.1.1(50) <= 3.1(49)
        def res18(df):
            if pd.isnull(df['col_50']) and pd.isnull(df['col_50']):
                return 'Blank'
            elif pd.isnull(df['col_51']) or pd.isnull(df['col_50']):
                if pd.isnull(df['col_51']):
                    return 'Probable Reporting Error(3.1.1 is blank)'
                elif pd.isnull(float(df['col_50'])):
                    return 'Inconsistent'
            elif float(df['col_51']) > float(df['col_50']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.1.1.a(51) + 4.1.1.b(52) + 4.1.3(54)<= 2.1(47)
        def res19(df):
            if pd.isnull(df['col_51']) and pd.isnull(df['col_52']) and pd.isnull(df['col_54']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_51']) or pd.isnull(df['col_52']) or pd.isnull(df['col_54']) or pd.isnull(df['col_47']):
                if pd.isnull((float(df['col_51'])) + (float(df['col_52'])) + (float(df['col_54']))) and not pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
                elif not pd.isnull((float(df['col_51'])) + (float(df['col_52'])) + (float(df['col_54']))) and pd.isnull(float(df['col_47'])):
                    return 'Probable Reporting Error'
            elif float(df['col_47']) < float(df['col_51']) + float(df['col_52']) + float(df['col_54']) :
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.1.2(53) <= 4.1.1.a(51) + 4.1.1.b(52)
        def res20(df):
            if pd.isnull(df['col_53']) and pd.isnull(df['col_51']) and pd.isnull(df['col_52']):
                return 'Blank'
            elif pd.isnull(df['col_53']) or pd.isnull(df['col_51']) or pd.isnull(df['col_52']):
                if pd.isnull(df['col_53']):
                    return 'Probable Reporting Error(4.1.2 is blank)'
                elif pd.isnull((float(df['col_51']) + float(df['col_52']))):
                    return 'Inconsistent'
            elif float(df['col_53']) > float(df['col_51']) + float(df['col_52']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #4.3.2.a(58) <= 4.3.1.a(56) + 4.3.1.b(57) + 4.2(55)
        def res21(df):
            if pd.isnull(df['col_58']) and pd.isnull(df['col_56']) and pd.isnull(df['col_57']) and pd.isnull(df['col_55']):
                return 'Blank'
            elif pd.isnull(df['col_58']) or pd.isnull(df['col_56']) or pd.isnull(df['col_57']) or pd.isnull(df['col_55']):
                if pd.isnull(df['col_58']):
                    return 'Probable Reporting Error(4.3.2.a is blank)'
                elif pd.isnull((float(df['col_56']) + float(df['col_57']) + float(df['col_55']))):
                    return 'Inconsistent'
            elif float(df['col_58']) > float(df['col_56']) + float(df['col_57']) + float(df['col_55']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.3.2.b(59) <= 4.3.2.a(58)
        def res22(df):
            if pd.isnull(df['col_59']) and pd.isnull(df['col_58']):
                return 'Blank'
            elif pd.isnull(df['col_59']) or pd.isnull(df['col_58']):
                if pd.isnull(df['col_59']):
                    return 'Probable Reporting Error(4.3.2.b is blank)'
                elif pd.isnull((float(df['col_58']))):
                    return 'Inconsistent'
            elif float(df['col_59']) > float(df['col_58']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.3.3(60) <= 4.3.1.a(56) + 4.3.1.b(57) + 4.2(55)
        def res23(df):
            if pd.isnull(df['col_60']) and pd.isnull(df['col_56']) and pd.isnull(df['col_57']) and pd.isnull(df['col_55']):
                return 'Blank'
            elif pd.isnull(df['col_60']) or pd.isnull(df['col_56']) or pd.isnull(df['col_57']) or pd.isnull(df['col_55']):
                if pd.isnull(df['col_60']):
                    return 'Probable Reporting Error(4.3.3 is blank)'
                elif pd.isnull((float(df['col_56'])) + (float(df['col_57'])) + (float(df['col_55']))):
                    return 'Inconsistent'
            elif float(df['col_60']) > float(df['col_56']) + float(df['col_57']) + float(df['col_55']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
         #4.4.1(61) <= 4.1.1.a(51) + 4.1.1.b(52)
        def res24(df):
            if pd.isnull(df['col_61']) and pd.isnull(df['col_51']) and pd.isnull(df['col_52']):
                return 'Blank'
            elif pd.isnull(df['col_61']) or pd.isnull(df['col_51']) or pd.isnull(df['col_52']):
                if pd.isnull(df['col_61']):
                    return 'Probable Reporting Error(4.4.1 is blank)'
                elif pd.isnull((float(df['col_51'])) + (float(df['col_52']))):
                    return 'Inconsistent'
            elif float(df['col_61']) > float(df['col_51']) + float(df['col_52']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.4.2(62) <= 4.4.1(61)
        def res25(df):
            if pd.isnull(df['col_62']) and pd.isnull(df['col_61']):
                return 'Blank'
            elif pd.isnull(df['col_62']) or pd.isnull(df['col_61']):
                if pd.isnull(df['col_62']):
                    return 'Probable Reporting Error(4.4.2 is blank)'
                elif pd.isnull((float(df['col_61']))):
                    return 'Inconsistent'
            elif float(df['col_62']) > float(df['col_61']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.4.3(63) <= 4.1.1.a(51) + 4.1.1.b(52)
        def res26(df):
            if pd.isnull(df['col_63']) and pd.isnull(df['col_51']) and pd.isnull(df['col_52']):
                return 'Blank'
            elif pd.isnull(df['col_63']) or pd.isnull(df['col_51']) or pd.isnull(df['col_52']):
                if pd.isnull(df['col_63']):
                    return 'Probable Reporting Error(4.4.3 is blank)'
                elif pd.isnull((float(df['col_51'])) + (float(df['col_52']))):
                    return 'Inconsistent'
            elif float(df['col_63']) > float(df['col_51']) + float(df['col_52']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #5.2(65) <= 5.1(64)
        def res27(df):
            if pd.isnull(df['col_65']) and pd.isnull(df['col_64']):
                return 'Blank'
            elif pd.isnull(df['col_65']) or pd.isnull(df['col_64']):
                if pd.isnull(df['col_65']):
                    return 'Probable Reporting Error(5.2 is blank)'
                elif pd.isnull((float(df['col_64']))):
                    return 'Inconsistent'
            elif float(df['col_65']) > float(df['col_64']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #6.3(68) <= 2.1(47)
        def res28(df):
            if pd.isnull(df['col_68']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_68']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_68']) and not pd.isnull(float(df['col_47'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
             # If value exists for all the elements
            else:

                lhs_value = float(df['col_68'])
                rhs_value = float(df['col_47'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df 
        
        #6.4(69) <= 2.1(47)
        def res29(df):
            if pd.isnull(df['col_69']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_69']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_69']) and not pd.isnull(float(df['col_47'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
             # If value exists for all the elements
            else:

                lhs_value = float(df['col_69'])
                rhs_value = float(df['col_47'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df 
        
        #7.2.1(72) <= 7.1.1(70)
        def res30(df):
            if pd.isnull(df['col_72']) and pd.isnull(df['col_70']):
                return 'Blank'
            elif pd.isnull(df['col_72']) or pd.isnull(df['col_70']):
                if pd.isnull(df['col_72']):
                    return 'Probable Reporting Error(7.2.1 is blank)'
                elif pd.isnull((float(df['col_70']))):
                    return 'Inconsistent'
            elif float(df['col_72']) > float(df['col_70']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #7.2.2(73) <= 7.1.2(71)
        def res31(df):
            if pd.isnull(df['col_73']) and pd.isnull(df['col_71']):
                return 'Blank'
            elif pd.isnull(df['col_73']) or pd.isnull(df['col_71']):
                if pd.isnull(df['col_73']):
                    return 'Probable Reporting Error(7.2.2 is blank)'
                elif pd.isnull((float(df['col_71']))):
                    return 'Inconsistent'
            elif float(df['col_73']) > float(df['col_71']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.2.3(77) <= 2.1(47)
        def res32(df):
            if pd.isnull(df['col_77']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_77']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_77']):
                    return 'Probable Reporting Error(8.2.3 is blank)'
                elif pd.isnull((float(df['col_47']))):
                    return 'Inconsistent'
            elif float(df['col_77']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
         #8.4(80) <= 2.1(47)
        def res33(df):
            if pd.isnull(df['col_80']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_80']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_80']):
                    return 'Probable Reporting Error(8.4 is blank)'
                elif pd.isnull((float(df['col_47']))):
                    return 'Inconsistent'
            elif float(df['col_80']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.7(83) <= 8.3(79) + 8.4(80) + 8.5(81)
        def res34(df):
            if pd.isnull(df['col_83']) and pd.isnull(df['col_79']) and pd.isnull(df['col_80']) and pd.isnull(df['col_81']):
                return 'Blank'
            elif pd.isnull(df['col_83']) or pd.isnull(df['col_79']) or pd.isnull(df['col_80']) or pd.isnull(df['col_81']):
                if pd.isnull(df['col_83']):
                    return 'Probable Reporting Error(8.7 is blank)'
                elif pd.isnull((float(df['col_79'])) + (float(df['col_80'])) + (float(df['col_81']))):
                    return 'Inconsistent'
            elif float(df['col_83']) > float(df['col_79']) + float(df['col_80']) + float(df['col_81']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.17.1(93) <= 8.1.1(74)
        def res35(df):
            if pd.isnull(df['col_93']) and pd.isnull(df['col_74']):
                return 'Blank'
            elif pd.isnull(df['col_93']) or pd.isnull(df['col_74']):
                if pd.isnull(df['col_93']):
                    return 'Probable Reporting Error(8.17.1 is blank)'
                elif pd.isnull((float(df['col_74']))):
                    return 'Inconsistent'
            elif float(df['col_93']) > float(df['col_74']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.17.2(94) <= 8.2.1(75) + 8.2.2(76) + 8.2.3(77) + 8.2.4(78)
        def res36(df):
            if pd.isnull(df['col_94']) and pd.isnull(df['col_75']) and pd.isnull(df['col_76']) and pd.isnull(df['col_77']) and pd.isnull(df['col_78']):
                return 'Blank'
            elif pd.isnull(df['col_94']) or pd.isnull(df['col_75']) or pd.isnull(df['col_76']) or pd.isnull(df['col_77']) or pd.isnull(df['col_78']):
                if pd.isnull(df['col_94']):
                    return 'Probable Reporting Error(8.17.2 is blank)'
                elif pd.isnull((float(df['col_75'])) + (float(df['col_76'])) + (float(df['col_77'])) + (float(df['col_78']))):
                    return 'Inconsistent'
            elif float(df['col_94']) > float(df['col_75']) + float(df['col_76']) + float(df['col_77']) + float(df['col_78']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.1(99) <= 4.1.1.a(51) + 4.1.1.b(52)
        def res37(df):
            if pd.isnull(df['col_99']) and pd.isnull(df['col_51']) and pd.isnull(df['col_52']):
                return 'Blank'
            elif pd.isnull(df['col_99']) or pd.isnull(df['col_51']) or pd.isnull(df['col_52']):
                if pd.isnull(df['col_99']):
                    return 'Probable Reporting Error(9.1.1 is blank)'
                elif pd.isnull((float(df['col_51'])) + (float(df['col_52']))):
                    return 'Inconsistent'
            elif float(df['col_99']) > float(df['col_51']) + float(df['col_52']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.2(100) <= 4.1.1.a(51) + 4.1.1.b(52)
        def res38(df):
            if pd.isnull(df['col_100']) and pd.isnull(df['col_51']) and pd.isnull(df['col_52']):
                return 'Blank'
            elif pd.isnull(df['col_100']) or pd.isnull(df['col_51']) or pd.isnull(df['col_52']):
                if pd.isnull(df['col_100']):
                    return 'Probable Reporting Error(9.1.2 is blank)'
                elif pd.isnull((float(df['col_51'])) + (float(df['col_52']))):
                    return 'Inconsistent'
            elif float(df['col_100']) > float(df['col_51']) + float(df['col_52']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.9(107) <= 4.1.1.a(51) + 4.1.1.b(52)
        def res39(df):
            if pd.isnull(df['col_107']) and pd.isnull(df['col_51']) and pd.isnull(df['col_52']):
                return 'Blank'
            elif pd.isnull(df['col_107']) or pd.isnull(df['col_51']) or pd.isnull(df['col_52']):
                if pd.isnull(df['col_107']):
                    return 'Probable Reporting Error(9.1.9 is blank)'
                elif pd.isnull((float(df['col_51'])) + (float(df['col_52']))):
                    return 'Inconsistent'
            elif float(df['col_107']) > float(df['col_51']) + float(df['col_52']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.13(111) <= 4.1.1.a(51) + 4.1.1.b(52)
        def res40(df):
            if pd.isnull(df['col_111']) and pd.isnull(df['col_51']) and pd.isnull(df['col_52']):
                return 'Blank'
            elif pd.isnull(df['col_111']) or pd.isnull(df['col_51']) or pd.isnull(df['col_52']):
                if pd.isnull(df['col_111']):
                    return 'Probable Reporting Error(9.1.13 is blank)'
                elif pd.isnull((float(df['col_51'])) + (float(df['col_52']))):
                    return 'Inconsistent'
            elif float(df['col_111']) > float(df['col_51']) + float(df['col_52']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.2.4.a(123) + 9.2.4.b(124) <= 9.2.1(120) + 9.2.2(121)
        def res41(df):
            if pd.isnull(df['col_123']) and pd.isnull(df['col_124']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']):
                return 'Blank'
            elif pd.isnull(df['col_123']) or pd.isnull(df['col_124']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']):
                if pd.isnull((float(df['col_123'])) + (float(df['col_124']))) and not pd.isnull(float(df['col_120']) + float(df['col_121'])):
                    return 'Probable Reporting Error'
                elif not pd.isnull((float(df['col_122'])) + (float(df['col_123']))) and pd.isnull(float(df['col_120']) + float(df['col_121'])):
                    return 'Inconsistent'
            elif (float(df['col_123']) + float(df['col_124'])) > (float(df['col_120']) + float(df['col_121'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #11.2.2(170) <= 11.2.1(169)
        def res42(df):
            if pd.isnull(df['col_170']) and pd.isnull(df['col_169']):
                return 'Blank'
            elif pd.isnull(df['col_170']) or pd.isnull(df['col_169']):
                if pd.isnull(df['col_170']):
                    return 'Probable Reporting Error(11.2.2 is blank)'
                elif pd.isnull((float(df['col_169']))):
                    return 'Inconsistent'
            elif float(df['col_170']) > float(df['col_169']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
         #11.4.2(175) <= 11.4.1(174)
        def res43(df):
            if pd.isnull(df['col_175']) and pd.isnull(df['col_174']):
                return 'Blank'
            elif pd.isnull(df['col_175']) or pd.isnull(df['col_174']):
                if pd.isnull(df['col_175']):
                    return 'Probable Reporting Error(11.2.2 is blank)'
                elif pd.isnull((float(df['col_174']))):
                    return 'Inconsistent'
            elif float(df['col_175']) > float(df['col_174']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
         #12.1.2.a(178) <= 12.1.1.a(176)
        def res44(df):
            if pd.isnull(df['col_178']) and pd.isnull(df['col_176']):
                return 'Blank'
            elif pd.isnull(df['col_178']) or pd.isnull(df['col_176']):
                if pd.isnull(df['col_178']):
                    return 'Probable Reporting Error(12.1.2.a is blank)'
                elif pd.isnull((float(df['col_176']))):
                    return 'Inconsistent'
            elif float(df['col_178']) > float(df['col_176']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #12.1.2.b(179) <= 12.1.1.b(177)
        def res45(df):
            if pd.isnull(df['col_179']) and pd.isnull(df['col_177']):
                return 'Blank'
            elif pd.isnull(df['col_179']) or pd.isnull(df['col_177']):
                if pd.isnull(df['col_179']):
                    return 'Probable Reporting Error(12.1.2.b is blank)'
                elif pd.isnull((float(df['col_177']))):
                    return 'Inconsistent'
            elif float(df['col_179']) > float(df['col_177']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #12.1.3.a(180) <= 12.1.1.a(176)
        def res46(df):
            if pd.isnull(df['col_180']) and pd.isnull(df['col_176']):
                return 'Blank'
            elif pd.isnull(df['col_180']) or pd.isnull(df['col_176']):
                if pd.isnull(df['col_180']):
                    return 'Probable Reporting Error(12.1.3.a is blank)'
                elif pd.isnull((float(df['col_176']))):
                    return 'Inconsistent'
            elif float(df['col_180']) > float(df['col_176']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #12.1.3.b(181) <= 12.1.1.b(177)
        def res47(df):
            if pd.isnull(df['col_181']) and pd.isnull(df['col_177']):
                return 'Blank'
            elif pd.isnull(df['col_181']) or pd.isnull(df['col_177']):
                if pd.isnull(df['col_181']):
                    return 'Probable Reporting Error(12.1.3.b is blank)'
                elif pd.isnull((float(df['col_177']))):
                    return 'Inconsistent'
            elif float(df['col_181']) > float(df['col_177']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.2.1(193) +14.2.2(194) >= 14.1.1(184) +14.1.2(184) +14.1.3(185) +14.1.4(186) +14.1.5(187) +14.1.6(188) +14.1.7(189) +14.1.8(190) +14.1.9(191)
        def res48(df):
            if pd.isnull(df['col_193']) and pd.isnull(df['col_194']) and pd.isnull(df['col_184']) and pd.isnull(df['col_185']) and pd.isnull(df['col_186']) and pd.isnull(df['col_187']) and pd.isnull(df['col_188']) and pd.isnull(df['col_189']) and pd.isnull(df['col_190']) and pd.isnull(df['col_191']) and pd.isnull(df['col_192']):
                return 'Blank'
            elif pd.isnull(df['col_193']) or pd.isnull(df['col_194']) or pd.isnull(df['col_184']) or pd.isnull(df['col_185']) or pd.isnull(df['col_186']) or pd.isnull(df['col_187']) or pd.isnull(df['col_188']) or pd.isnull(df['col_189']) or pd.isnull(df['col_190']) or pd.isnull(df['col_191']) or pd.isnull(df['col_192']):
                if pd.isnull((float(df['col_193'])) + (float(df['col_194']))):
                    return 'Inconsistent'
                elif pd.isnull(float(df['col_184']) + float(df['col_185']) + float(df['col_186'])+ float(df['col_187']) + float(df['col_188']) + float(df['col_189'])+ float(df['col_190']) + float(df['col_191']) + float(df['col_192'])):
                    return 'Probable Reporting Error'
            elif (float(df['col_193']) + float(df['col_194'])) < (float(df['col_184']) + float(df['col_185']) + float(df['col_186'])+ float(df['col_187']) + float(df['col_188']) + float(df['col_189'])+ float(df['col_190']) + float(df['col_191']) + float(df['col_192'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.3.3(199) <= 14.3.1.a(195) +14.3.1.b(196) +14.3.2.a(197) +14.3.2.b(198)
        def res49(df):
            if pd.isnull(df['col_199']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']):
                return 'Blank'
            elif pd.isnull(df['col_199']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']):
                if pd.isnull(df['col_199']):
                    return 'Probable Reporting Error(14.3.3 is blank)'
                elif pd.isnull((float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198']))):
                    return 'Inconsistent'
            elif float(df['col_199']) > float(df['col_195']) + float(df['col_196']) + float(df['col_197']) + float(df['col_198']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.4.1(200) <=14.3.1.a(195) +14.3.1.b(196) +14.3.2.a(197) +14.3.2.b(198)
        def res50(df):
            if pd.isnull(df['col_200']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']):
                return 'Blank'
            elif pd.isnull(df['col_200']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']):
                if pd.isnull(df['col_200']):
                    return 'Probable Reporting Error(14.4.1 is blank)'
                elif pd.isnull((float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198']))):
                    return 'Inconsistent'
            elif float(df['col_200']) > float(df['col_195']) + float(df['col_196']) + float(df['col_197']) + float(df['col_198']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        
        #14.4.2 (201)<=14.3.1.a (195)+14.3.1.b (196)+14.3.2.a (197)+14.3.2.b (198)
        def res51(df):
            if pd.isnull(df['col_201']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_201']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']):
                if pd.isnull(df['col_201']):
                    return 'Probable Reporting Error(14.4.2 is blank)'
                elif pd.isnull((float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198']))):
                    return 'Inconsistent'
            elif float(df['col_201']) > float(df['col_195']) + float(df['col_196']) + float(df['col_197']) + float(df['col_198']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #14.4.3 (202)<=14.3.1.a (195)+14.3.1.b (196)+14.3.2.a (197)+14.3.2.b (198)
        def res52(df):
            if pd.isnull(df['col_202']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']):
                return 'Blank'
            elif pd.isnull(df['col_202']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']):
                if pd.isnull(df['col_202']):
                    return 'Probable Reporting Error(14.4.3 is blank)'
                elif pd.isnull((float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198']))):
                    return 'Inconsistent'
            elif float(df['col_202']) > float(df['col_195']) + float(df['col_196']) + float(df['col_197']) + float(df['col_198']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.4.4 (203)<=14.3.1.a (195)+14.3.1.b (196)+14.3.2.a (197)+14.3.2.b (198)
        def res53(df):
            if pd.isnull(df['col_203']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']):
                return 'Blank'
            elif pd.isnull(df['col_203']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']):
                if pd.isnull(df['col_203']):
                    return 'Probable Reporting Error(14.4.4 is blank)'
                elif pd.isnull((float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198']))):
                    return 'Inconsistent'
            elif float(df['col_203']) > float(df['col_195']) + float(df['col_196']) + float(df['col_197']) + float(df['col_198']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

         #14.4.5 (204)<=14.3.1.a (195)+14.3.1.b (196)+14.3.2.a (197)+14.3.2.b (198)
        def res54(df):
            if pd.isnull(df['col_204']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']):
                return 'Blank'
            elif pd.isnull(df['col_204']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']):
                if pd.isnull(df['col_204']):
                    return 'Probable Reporting Error(14.4.5 is blank)'
                elif pd.isnull((float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198']))):
                    return 'Inconsistent'
            elif float(df['col_204']) > float(df['col_195']) + float(df['col_196']) + float(df['col_197']) + float(df['col_198']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
         #14.4.6 (205)<=14.3.1.a (195)+14.3.1.b (196)+14.3.2.a (197)+14.3.2.b (198)
        def res55(df):
            if pd.isnull(df['col_205']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']):
                return 'Blank'
            elif pd.isnull(df['col_205']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']):
                if pd.isnull(df['col_205']):
                    return 'Probable Reporting Error(14.4.6 is blank)'
                elif pd.isnull((float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198']))):
                    return 'Inconsistent'
            elif float(df['col_205']) > float(df['col_195']) + float(df['col_196']) + float(df['col_197']) + float(df['col_198']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df        
        
        #14.4.7 (206)<=14.3.1.a (195)+14.3.1.b (196)+14.3.2.a (197)+14.3.2.b (198)
        def res56(df):
            if pd.isnull(df['col_206']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']):
                return 'Blank'
            elif pd.isnull(df['col_206']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']):
                if pd.isnull(df['col_206']):
                    return 'Probable Reporting Error(14.4.7 is blank)'
                elif pd.isnull((float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198']))):
                    return 'Inconsistent'
            elif float(df['col_206']) > float(df['col_195']) + float(df['col_196']) + float(df['col_197']) + float(df['col_198']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df      
        
         #14.4.8 (207)<=14.3.1.a (195)+14.3.1.b (196)+14.3.2.a (197)+14.3.2.b (198)
        def res57(df):
            if pd.isnull(df['col_207']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']):
                return 'Blank'
            elif pd.isnull(df['col_207']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']):
                if pd.isnull(df['col_207']):
                    return 'Probable Reporting Error(14.4.8 is blank)'
                elif pd.isnull((float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198']))):
                    return 'Inconsistent'
            elif float(df['col_207']) > float(df['col_195']) + float(df['col_196']) + float(df['col_197']) + float(df['col_198']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df    

        #14.6.1(209) <= 14.5(208)
        def res58(df):
            if pd.isnull(df['col_209']) and pd.isnull(df['col_208']):
                return 'Blank'
            elif pd.isnull(df['col_209']) or pd.isnull(df['col_208']):
                if pd.isnull(df['col_209']):
                    return 'Probable Reporting Error(14.6.1 is blank)'
                elif pd.isnull(float(df['col_208'])):
                    return 'Inconsistent'
            elif float(df['col_209']) > float(df['col_208']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df        

        #14.6.2(210) <= 14.5(208)
        def res59(df):
            if pd.isnull(df['col_210']) and pd.isnull(df['col_208']):
                return 'Blank'
            elif pd.isnull(df['col_210']) or pd.isnull(df['col_208']):
                if pd.isnull(df['col_210']):
                    return 'Probable Reporting Error(14.6.2 is blank)'
                elif pd.isnull(float(df['col_208'])):
                    return 'Inconsistent'
            elif float(df['col_210']) > float(df['col_208']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #14.6.3(211) <= 14.5(208)
        def res60(df):
            if pd.isnull(df['col_211']) and pd.isnull(df['col_208']):
                return 'Blank'
            elif pd.isnull(df['col_211']) or pd.isnull(df['col_208']):
                if pd.isnull(df['col_211']):
                    return 'Probable Reporting Error(14.6.3 is blank)'
                elif pd.isnull(float(df['col_208'])):
                    return 'Inconsistent'
            elif float(df['col_211']) > float(df['col_208']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #14.6.4(212) <= 14.5(208)
        def res61(df):
            if pd.isnull(df['col_212']) and pd.isnull(df['col_208']):
                return 'Blank'
            elif pd.isnull(df['col_212']) or pd.isnull(df['col_208']):
                if pd.isnull(df['col_212']):
                    return 'Probable Reporting Error(14.6.4 is blank)'
                elif pd.isnull(float(df['col_208'])):
                    return 'Inconsistent'
            elif float(df['col_212']) > float(df['col_208']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df             

        #14.6.5(213) <= 14.5(208)
        def res62(df):
            if pd.isnull(df['col_213']) and pd.isnull(df['col_208']):
                return 'Blank'
            elif pd.isnull(df['col_213']) or pd.isnull(df['col_208']):
                if pd.isnull(df['col_213']):
                    return 'Probable Reporting Error(14.6.5 is blank)'
                elif pd.isnull(float(df['col_208'])):
                    return 'Inconsistent'
            elif float(df['col_213']) > float(df['col_208']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #14.6.6(213) <= 14.5(207)
        def res63(df):
            if pd.isnull(df['col_213']) and pd.isnull(df['col_207']):
                return 'Blank'
            elif pd.isnull(df['col_213']) or pd.isnull(df['col_207']):
                if pd.isnull(df['col_213']):
                    return 'Probable Reporting Error(14.6.6 is blank)'
                elif pd.isnull(float(df['col_207'])):
                    return 'Inconsistent'
            elif float(df['col_213']) > float(df['col_207']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #14.7(215) <= 14.5(208)
        def res64(df):
            if pd.isnull(df['col_215']) and pd.isnull(df['col_208']):
                return 'Blank'
            elif pd.isnull(df['col_215']) or pd.isnull(df['col_208']):
                if pd.isnull(df['col_215']):
                    return 'Probable Reporting Error(14.7 is blank)'
                elif pd.isnull(float(df['col_208'])):
                    return 'Inconsistent'
            elif float(df['col_215']) > float(df['col_208']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #14.14.2(234) <= 14.14.1(233)
        def res65(df):
            if pd.isnull(df['col_234']) and pd.isnull(df['col_233']):
                return 'Blank'
            elif pd.isnull(df['col_234']) or pd.isnull(df['col_233']):
                if pd.isnull(df['col_234']):
                    return 'Probable Reporting Error(14.14.2 is blank)'
                elif pd.isnull(float(df['col_233'])):
                    return 'Inconsistent'
            elif float(df['col_234']) > float(df['col_233']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

       #15.2.2(242) <= 15.2.1(241)
        def res66(df):
            if pd.isnull(df['col_242']) and pd.isnull(df['col_241']):
                return 'Blank'
            elif pd.isnull(df['col_242']) or pd.isnull(df['col_241']):
                if pd.isnull(df['col_242']):
                    return 'Probable Reporting Error(15.2.2 is blank)'
                elif pd.isnull(float(df['col_241'])):
                    return 'Inconsistent'
            elif float(df['col_242']) > float(df['col_241']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #15.3.2b(246) <= 15.3.2a(245)
        def res67(df):
            if pd.isnull(df['col_246']) and pd.isnull(df['col_245']):
                return 'Blank'
            elif pd.isnull(df['col_246']) or pd.isnull(df['col_245']):
                if pd.isnull(df['col_246']):
                    return 'Probable Reporting Error(15.3.2b is blank)'
                elif pd.isnull(float(df['col_245'])):
                    return 'Inconsistent'
            elif float(df['col_246']) > float(df['col_245']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #15.3.3b(248) <= 15.3.3a(247)
        def res68(df):
            if pd.isnull(df['col_248']) and pd.isnull(df['col_247']):
                return 'Blank'
            elif pd.isnull(df['col_248']) or pd.isnull(df['col_247']):
                if pd.isnull(df['col_248']):
                    return 'Probable Reporting Error(15.3.3b is blank)'
                elif pd.isnull(float(df['col_247'])):
                    return 'Inconsistent'
            elif float(df['col_248']) > float(df['col_247']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #15.3.3c(249) <= 15.3.3b(248)
        def res69(df):
            if pd.isnull(df['col_249']) and pd.isnull(df['col_248']):
                return 'Blank'
            elif pd.isnull(df['col_249']) or pd.isnull(df['col_248']):
                if pd.isnull(df['col_249']):
                    return 'Probable Reporting Error(15.3.3c is blank)'
                elif pd.isnull(float(df['col_248'])):
                    return 'Inconsistent'
            elif float(df['col_249']) > float(df['col_248']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #15.4.2(255) <= 15.4.1(254)
        def res70(df):
            if pd.isnull(df['col_255']) and pd.isnull(df['col_254']):
                return 'Blank'
            elif pd.isnull(df['col_255']) or pd.isnull(df['col_254']):
                if pd.isnull(df['col_255']):
                    return 'Probable Reporting Error(15.4.2 is blank)'
                elif pd.isnull(float(df['col_254'])):
                    return 'Inconsistent'
            elif float(df['col_255']) > float(df['col_254']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
    
       #9.6.1 (138) <=9.1.1 (99)+9.1.2 (100)+9.1.3 (101)+9.1.4 (102) +9.1.5 (103)+9.1.6 (104)+9.1.7 (105)+9.1.8 (106)+9.1.13 (111)+9.1.14 (112)+9.1.15 (113)+9.1.16 (114)+9.1.17 (115)+9.1.18 (116)+9.1.19 (117)+9.1.20 (118)+9.1.21 (119)+9.2.1 (120)+9.2.2 (121)+9.2.3 (122)+9.3.1 (125)+9.3.2 (126)+9.3.3 (127)+9.4.1 (128)+9.4.2 (129)+9.4.3 (130)+9.4.5 (132)+9.4.6 (133)+9.5.1 (134)+9.5.2 (135)+9.5.3 (136)+9.5.4 (137)
        def res71(df):
            if pd.isnull(df['col_138']) and pd.isnull(df['col_99']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_111']) and pd.isnull(df['col_112']) and pd.isnull(df['col_113']) and pd.isnull(df['col_114']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_125']) and pd.isnull(df['col_126']) and pd.isnull(df['col_127']) and pd.isnull(df['col_128']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_132']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_135']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']):
                return 'Blank'
            elif pd.isnull(df['col_138']) or pd.isnull(df['col_99']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_111']) or pd.isnull(df['col_112']) or pd.isnull(df['col_113']) or pd.isnull(df['col_114']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_125']) or pd.isnull(df['col_126']) or pd.isnull(df['col_127']) or pd.isnull(df['col_128']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_132']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_135']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']):
                if pd.isnull(df['col_138']):
                    return 'Probable Reporting Error(9.6.1 is blank)'
                elif pd.isnull((float(df['col_99'])) + (float(df['col_100'])) + (float(df['col_101'])) + (float(df['col_101'])) + (float(df['col_102'])) + (float(df['col_103'])) + (float(df['col_104'])) + (float(df['col_105'])) + (float(df['col_110'])) + (float(df['col_111'])) + (float(df['col_112'])) + (float(df['col_113'])) + (float(df['col_114'])) + (float(df['col_115'])) + (float(df['col_116'])) + (float(df['col_117'])) + (float(df['col_118'])) + (float(df['col_119'])) + (float(df['col_120'])) + (float(df['col_121'])) + (float(df['col_124'])) + (float(df['col_125'])) + (float(df['col_126'])) + (float(df['col_127'])) + (float(df['col_128'])) + (float(df['col_129'])) + (float(df['col_131'])) + (float(df['col_132'])) + (float(df['col_133'])) + (float(df['col_134'])) + (float(df['col_135'])) + (float(df['col_136']))+ (float(df['col_137']))):
                    return 'Inconsistent'
            elif float(df['col_138']) > float(df['col_99']) + (float(df['col_100']) + float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_110']) + float(df['col_111']) + float(df['col_112']) + float(df['col_113']) + float(df['col_114']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_124']) + float(df['col_125']) + float(df['col_126']) + float(df['col_127']) + float(df['col_128']) + float(df['col_129']) + float(df['col_131']) + float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + (float(df['col_137']))):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #9.6.2 (139) <=9.1.1 (99)+9.1.2 (100)+9.1.3 (101)+9.1.4 (102) +9.1.5 (103)+9.1.6 (104)+9.1.7 (105)+9.1.8 (106)+9.1.13 (111)+9.1.14 (112)+9.1.15 (113)+9.1.16 (114)+9.1.17 (115)+9.1.18 (116)+9.1.19 (117)+9.1.20 (118)+9.1.21 (119)+9.2.1 (120)+9.2.2 (121)+9.2.3 (122)+9.3.1 (125)+9.3.2 (126)+9.3.3 (127)+9.4.1 (128)+9.4.2 (129)+9.4.3 (130)+9.4.5 (132)+9.4.6 (133)+9.5.1 (134)+9.5.2 (135)+9.5.3 (136)+9.5.4 (137)
        def res72(df):
            if pd.isnull(df['col_139']) and pd.isnull(df['col_99']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_111']) and pd.isnull(df['col_112']) and pd.isnull(df['col_113']) and pd.isnull(df['col_114']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_125']) and pd.isnull(df['col_126']) and pd.isnull(df['col_127']) and pd.isnull(df['col_128']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_132']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_135']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']):
                return 'Blank'
            elif pd.isnull(df['col_139']) or pd.isnull(df['col_99']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_111']) or pd.isnull(df['col_112']) or pd.isnull(df['col_113']) or pd.isnull(df['col_114']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_125']) or pd.isnull(df['col_126']) or pd.isnull(df['col_127']) or pd.isnull(df['col_128']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_132']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_135']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']):
                if pd.isnull(df['col_139']):
                    return 'Probable Reporting Error(9.6.1 is blank)'
                elif pd.isnull((float(df['col_99'])) + (float(df['col_100'])) + (float(df['col_101'])) + (float(df['col_101'])) + (float(df['col_102'])) + (float(df['col_103'])) + (float(df['col_104'])) + (float(df['col_105'])) + (float(df['col_110'])) + (float(df['col_111'])) + (float(df['col_112'])) + (float(df['col_113'])) + (float(df['col_114'])) + (float(df['col_115'])) + (float(df['col_116'])) + (float(df['col_117'])) + (float(df['col_118'])) + (float(df['col_119'])) + (float(df['col_120'])) + (float(df['col_121'])) + (float(df['col_124'])) + (float(df['col_125'])) + (float(df['col_126'])) + (float(df['col_127'])) + (float(df['col_128'])) + (float(df['col_129'])) + (float(df['col_131'])) + (float(df['col_132'])) + (float(df['col_133'])) + (float(df['col_134'])) + (float(df['col_135'])) + (float(df['col_136']))+ (float(df['col_137']))):
                    return 'Inconsistent'
            elif float(df['col_139']) > float(df['col_99']) + float(df['col_100']) + float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_110']) + float(df['col_111']) + float(df['col_112']) + float(df['col_113']) + float(df['col_114']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_124']) + float(df['col_125']) + float(df['col_126']) + float(df['col_127']) + float(df['col_128']) + float(df['col_129']) + float(df['col_131']) + float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + float(df['col_137']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #9.6.3 (140) <=9.1.1 (99)+9.1.2 (100)+9.1.3 (101)+9.1.4 (102) +9.1.5 (103)+9.1.6 (104)+9.1.7 (105)+9.1.8 (106)+9.1.13 (111)+9.1.14 (112)+9.1.15 (113)+9.1.16 (114)+9.1.17 (115)+9.1.18 (116)+9.1.19 (117)+9.1.20 (118)+9.1.21 (119)+9.2.1 (120)+9.2.2 (121)+9.2.3 (122)+9.3.1 (125)+9.3.2 (126)+9.3.3 (127)+9.4.1 (128)+9.4.2 (129)+9.4.3 (130)+9.4.5 (132)+9.4.6 (133)+9.5.1 (134)+9.5.2 (135)+9.5.3 (136)+9.5.4 (137)
        def res73(df):
            if pd.isnull(df['col_140']) and pd.isnull(df['col_99']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_111']) and pd.isnull(df['col_112']) and pd.isnull(df['col_113']) and pd.isnull(df['col_114']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_125']) and pd.isnull(df['col_126']) and pd.isnull(df['col_127']) and pd.isnull(df['col_128']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_132']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_135']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']):
                    return 'Blank'
            elif pd.isnull(df['col_140']) or pd.isnull(df['col_99']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_111']) or pd.isnull(df['col_112']) or pd.isnull(df['col_113']) or pd.isnull(df['col_114']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_125']) or pd.isnull(df['col_126']) or pd.isnull(df['col_127']) or pd.isnull(df['col_128']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_132']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_135']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']):
                if pd.isnull(df['col_140']):
                    return 'Probable Reporting Error(9.6.1 is blank)'
                elif pd.isnull((float(df['col_99'])) + (float(df['col_100'])) + (float(df['col_101'])) + (float(df['col_101'])) + (float(df['col_102'])) + (float(df['col_103'])) + (float(df['col_104'])) + (float(df['col_105'])) + (float(df['col_110'])) + (float(df['col_111'])) + (float(df['col_112'])) + (float(df['col_113'])) + (float(df['col_114'])) + (float(df['col_115'])) + (float(df['col_116'])) + (float(df['col_117'])) + (float(df['col_118'])) + (float(df['col_119'])) + (float(df['col_120'])) + (float(df['col_121'])) + (float(df['col_124'])) + (float(df['col_125'])) + (float(df['col_126'])) + (float(df['col_127'])) + (float(df['col_128'])) + (float(df['col_129'])) + (float(df['col_131'])) + (float(df['col_132'])) + (float(df['col_133'])) + (float(df['col_134'])) + (float(df['col_135'])) + (float(df['col_136']))+ (float(df['col_137']))):
                    return 'Inconsistent'
            elif float(df['col_140']) > float(df['col_99']) + (float(df['col_100']) + float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_110']) + float(df['col_111']) + float(df['col_112']) + float(df['col_113']) + float(df['col_114']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_124']) + float(df['col_125']) + float(df['col_126']) + float(df['col_127']) + float(df['col_128']) + float(df['col_129']) + float(df['col_131']) + float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + float(df['col_137'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

      #9.7.2(142) <= 9.7.1(141)
        def res74(df):
            if pd.isnull(df['col_142']) and pd.isnull(df['col_141']):
                return 'Blank'
            elif pd.isnull(df['col_142']) or pd.isnull(df['col_141']):
                if pd.isnull(df['col_142']):
                    return 'Probable Reporting Error(9.7.2 is blank)'
                elif pd.isnull(float(df['col_141'])):
                    return 'Inconsistent'
            elif float(df['col_142']) > float(df['col_141']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #9.7.3(143) <= 9.7.2(142)
        def res75(df):
            if pd.isnull(df['col_143']) and pd.isnull(df['col_142']):
                return 'Blank'
            elif pd.isnull(df['col_143']) or pd.isnull(df['col_142']):
                if pd.isnull(df['col_143']):
                    return 'Probable Reporting Error(9.7.3 is blank)'
                elif pd.isnull(float(df['col_142'])):
                    return 'Inconsistent'
            elif float(df['col_143']) > float(df['col_142']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.1b (164) <= 11.1.1a (163)
        def res76(df):
            if pd.isnull(df['col_164']) and pd.isnull(df['col_163']):
                return 'Blank'
            elif pd.isnull(df['col_164']) or pd.isnull(df['col_163']):
                if pd.isnull(df['col_164']):
                    return 'Probable Reporting Error(11.1.1b is blank)'
                elif pd.isnull(float(df['col_163'])):
                    return 'Inconsistent'
            elif float(df['col_164']) > float(df['col_163']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.1c (165) <= 11.1.1a (163)
        def res77(df):
            if pd.isnull(df['col_165']) and pd.isnull(df['col_163']):
                return 'Blank'
            elif pd.isnull(df['col_165']) or pd.isnull(df['col_163']):
                if pd.isnull(df['col_165']):
                    return 'Probable Reporting Error(11.1.1c is blank)'
                elif pd.isnull(float(df['col_163'])):
                    return 'Inconsistent'
            elif float(df['col_165']) > float(df['col_163']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.2b (167) <= 11.1.2a (166)
        def res78(df):
            if pd.isnull(df['col_167']) and pd.isnull(df['col_166']):
                return 'Blank'
            elif pd.isnull(df['col_167']) or pd.isnull(df['col_166']):
                if pd.isnull(df['col_167']):
                    return 'Probable Reporting Error(11.1.2b is blank)'
                elif pd.isnull(float(df['col_166'])):
                    return 'Inconsistent'
            elif float(df['col_167']) > float(df['col_166']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.2c (168) <= 11.1.2a (166)
        def res79(df):
            if pd.isnull(df['col_168']) and pd.isnull(df['col_166']):
                return 'Blank'
            elif pd.isnull(df['col_168']) or pd.isnull(df['col_166']):
                if pd.isnull(df['col_168']):
                    return 'Probable Reporting Error(11.1.2c is blank)'
                elif pd.isnull(float(df['col_166'])):
                    return 'Inconsistent'
            elif float(df['col_168']) > float(df['col_166']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #14.18 (238) <= 14.8.6 (221)
        def res80(df):
            if pd.isnull(df['col_238']) and pd.isnull(df['col_221']):
                return 'Blank'
            elif pd.isnull(df['col_238']) or pd.isnull(df['col_221']):
                if pd.isnull(df['col_238']):
                    return 'Inconsistent'
                elif pd.isnull(float(df['col_221'])):
                    return 'Probable Reporting Error(14.8.6 is blank)'
            elif float(df['col_238']) > float(df['col_221']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #14.9.1 (223)<=14.3.1.a (195)+14.3.1.b (196)
        def res81(df):
            if pd.isnull(df['col_223']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']):
                return 'Blank'
            elif pd.isnull(df['col_223']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']):
                if pd.isnull(df['col_223']):
                    return 'Probable Reporting Error(14.9.1 is blank)'
                elif pd.isnull((float(df['col_195'])) + (float(df['col_196']))):
                    return 'Inconsistent'
            elif float(df['col_223']) > float(df['col_195']) + float(df['col_196']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df      

         #14.9.2 (224)<=14.3.2.a (197)+14.3.2.b (198)
        def res82(df):
            if pd.isnull(df['col_224']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']):
                return 'Blank'
            elif pd.isnull(df['col_224']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']):
                if pd.isnull(df['col_224']):
                    return 'Probable Reporting Error(14.9.2 is blank)'
                elif pd.isnull((float(df['col_197'])) + (float(df['col_198']))):
                    return 'Inconsistent'
            elif float(df['col_224']) > float(df['col_197']) + float(df['col_198']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df     

         #14.13 (232)<=14.12.1 (227)+14.12.2 (228)+14.12.3 (229)+14.12.4 (230)+14.12.5 (231)
        def res83(df):
            if pd.isnull(df['col_232']) and pd.isnull(df['col_227']) and pd.isnull(df['col_228']) and pd.isnull(df['col_229']) and pd.isnull(df['col_230']) and pd.isnull(df['col_231']):
                return 'Blank'
            elif pd.isnull(df['col_232']) or pd.isnull(df['col_227']) or pd.isnull(df['col_228']) or pd.isnull(df['col_229']) or pd.isnull(df['col_230']) or pd.isnull(df['col_231']):
                if pd.isnull(df['col_232']):
                    return 'Probable Reporting Error(14.13 is blank)'
                elif pd.isnull((float(df['col_227'])) + (float(df['col_228'])) + (float(df['col_229'])) + (float(df['col_230'])) + (float(df['col_231']))):
                    return 'Inconsistent'
            elif float(df['col_232']) > float(df['col_228']) + float(df['col_229']) + float(df['col_230']) + float(df['col_231']) + float(df['col_232']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df      

        #14.8.2(217) <= 14.8.1(216)
        def res84(df):
            if pd.isnull(df['col_217']) and pd.isnull(df['col_216']):
                return 'Blank'
            elif pd.isnull(df['col_217']) or pd.isnull(df['col_216']):
                if pd.isnull(df['col_217']):
                    return 'Probable Reporting Error(14.8.2 is blank)'
                elif pd.isnull(float(df['col_216'])):
                    return 'Inconsistent'
            elif float(df['col_217']) > float(df['col_216']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #14.8.3(218) <= 14.8.1(216)
        def res85(df):
            if pd.isnull(df['col_218']) and pd.isnull(df['col_216']):
                return 'Blank'
            elif pd.isnull(df['col_218']) or pd.isnull(df['col_216']):
                if pd.isnull(df['col_218']):
                    return 'Probable Reporting Error(14.8.3 is blank)'
                elif pd.isnull(float(df['col_216'])):
                    return 'Inconsistent'
            elif float(df['col_218']) > float(df['col_216']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
         #15.3.4b(251) <= 15.3.4a(250)
        def res86(df):
            if pd.isnull(df['col_251']) and pd.isnull(df['col_250']):
                return 'Blank'
            elif pd.isnull(df['col_251']) or pd.isnull(df['col_250']):
                if pd.isnull(df['col_251']):
                    return 'Probable Reporting Error(15.3.4b is blank)'
                elif pd.isnull(float(df['col_250'])):
                    return 'Inconsistent'
            elif float(df['col_251']) > float(df['col_250']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df       
        
         #15.3.4d(253) <= 15.3.4c(252)
        def res87(df):
            if pd.isnull(df['col_253']) and pd.isnull(df['col_252']):
                return 'Blank'
            elif pd.isnull(df['col_253']) or pd.isnull(df['col_252']):
                if pd.isnull(df['col_253']):
                    return 'Probable Reporting Error(15.3.4d is blank)'
                elif pd.isnull(float(df['col_252'])):
                    return 'Inconsistent'
            elif float(df['col_253']) > float(df['col_252']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  


        # Renamining column names=======================DH
        df['1.1.1<=1.1'] = df.apply(res1, axis=1)
        df['15.3.1.b<=15.3.1.a'] = df.apply(res2, axis=1)
        df['1.2.4<=1.1'] = df.apply(res3, axis=1)
        df['1.2.5<=1.1'] = df.apply(res4, axis=1)
        df['1.2.7<=1.1'] = df.apply(res5, axis=1)
        df['1.3.1.a<=1.3.1'] = df.apply(res6, axis=1)
        df['1.3.2<=2.1'] = df.apply(res7, axis=1)
        df['1.4.4>=1.4.3'] = df.apply(res8, axis=1)
        df['1.5.1<=1.1'] = df.apply(res9, axis=1)
        df['1.5.2<=1.5.1'] = df.apply(res10, axis=1)
        df['1.5.3<=1.5.2'] = df.apply(res11, axis=1)
        df['1.6.1.a<=1.1'] = df.apply(res12, axis=1)
        df['1.6.1.b<=1.6.1.a'] = df.apply(res13, axis=1)
        df['1.6.1.c<=1.6.1.b'] = df.apply(res14, axis=1)
        df['1.6.1.e<=1.6.1.d'] = df.apply(res15, axis=1)
        df['2.1.1<=2.1'] = df.apply(res16, axis=1)
        df['3.1<=2.1'] = df.apply(res17, axis=1)
        df['3.1.1<=3.1'] = df.apply(res18, axis=1)
        df['4.1.1.a+4.1.1.b+4.1.3>=2.1'] = df.apply(res19, axis=1)
        df['4.1.2<=4.1.1.a+4.1.1.b'] = df.apply(res20, axis=1)
        df['4.3.2.a<=4.3.1.a+4.3.1.b+4.2'] = df.apply(res21, axis=1)
        df['4.3.2.b<=4.3.2.a'] = df.apply(res22, axis=1)
        df['4.3.3<=4.3.1.a+4.3.1.b+4.2'] = df.apply(res23, axis=1)
        df['4.4.1<=4.1.1.a+4.1.1.b'] = df.apply(res24, axis=1)
        df['4.4.2<=4.4.1'] = df.apply(res25, axis=1)
        df['4.4.3<=4.1.1.a+4.1.1.b'] = df.apply(res26, axis=1)
        df['5.2<=5.1'] = df.apply(res27, axis=1)
        df['6.3<=2.1'] = df.apply(res28, axis=1)
        df['6.4<=2.1'] = df.apply(res29, axis=1)
        df['7.2.1<=7.1.1'] = df.apply(res30, axis=1)
        df['7.2.2<=7.1.2'] = df.apply(res31, axis=1)
        df['8.2.3<=2.1'] = df.apply(res32, axis=1)
        df['8.4<=2.1'] = df.apply(res33, axis=1)
        df['8.7<=8.3+8.4+8.5'] = df.apply(res34, axis=1)
        df['8.17.1<=8.1.1'] = df.apply(res35, axis=1)
        df['8.17.2<=8.2.1+8.2.2+8.2.3+8.2.4'] = df.apply(res36, axis=1)
        df['9.1.1<=4.1.1.a+4.1.1.b'] = df.apply(res37, axis=1)
        df['9.1.2<=4.1.1.a+4.1.1.b'] = df.apply(res38, axis=1)
        df['9.1.9<=4.1.1.a+4.1.1.b'] = df.apply(res39, axis=1)
        df['9.1.13<=4.1.1.a+4.1.1.b'] = df.apply(res40, axis=1)
        df['9.2.4.a+9.2.4.b<=9.2.1+ 9.2.2'] = df.apply(res41, axis=1)
        df['11.2.2<=11.2.1'] = df.apply(res42, axis=1)
        df['11.4.2<=11.4.1'] = df.apply(res43, axis=1)
        df['12.1.2.a<=12.1.1.a'] = df.apply(res44, axis=1)
        df['12.1.2.b<=12.1.1.b'] = df.apply(res45, axis=1)
        df['12.1.3.a<=12.1.1.a'] = df.apply(res46, axis=1)
        df['12.1.3.b<=12.1.1.b'] = df.apply(res47, axis=1)
        df['14.2.1+14.2.2>=14.1.1+14.1.2+14.1.3+14.1.4+14.1.5+14.1.6+14.1.7+14.1.8+14.1.9'] = df.apply(res48, axis=1)
        df['14.3.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res49, axis=1)
        df['14.4.1<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res50, axis=1)
        df['14.4.2<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res51, axis=1)
        df['14.4.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res52, axis=1)
        df['14.4.4<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res53, axis=1)
        df['14.4.5<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res54, axis=1)
        df['14.4.6<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res55, axis=1)
        df['14.4.7<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res56, axis=1)
        df['14.4.8<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res57, axis=1)
        df['14.6.1<=14.5'] = df.apply(res58, axis=1)
        df['14.6.2<=14.5'] = df.apply(res59, axis=1)
        df['14.6.3<=14.5'] = df.apply(res60, axis=1)
        df['14.6.4<=14.5'] = df.apply(res61, axis=1)
        df['14.6.5<=14.5'] = df.apply(res62, axis=1)
        df['14.6.6<=14.5'] = df.apply(res63, axis=1)
        df['14.7<=14.5'] = df.apply(res64, axis=1)
        df['14.14.2<=14.14.1'] = df.apply(res65, axis=1)
        df['15.2.2<=15.2.1'] = df.apply(res66, axis=1)
        df['15.3.2.b<=15.3.2.a'] = df.apply(res67, axis=1)
        df['15.3.3.b<=15.3.3.a'] = df.apply(res68, axis=1)
        df['15.3.3.c<=15.3.3.b'] = df.apply(res69, axis=1)
        df['15.4.2<=15.4.1'] = df.apply(res70, axis=1)
        df['9.6.1<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'] = df.apply(res71, axis=1)
        df['9.6.2<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'] = df.apply(res72, axis=1)
        df['9.6.3<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'] = df.apply(res73, axis=1)
        df['9.7.2<=9.7.1'] = df.apply(res74, axis=1)
        df['9.7.3<=9.7.2'] = df.apply(res75, axis=1)
        df['11.1.1.b<=11.1.1.a'] = df.apply(res76, axis=1)
        df['11.1.1.c<=11.1.1.a'] = df.apply(res77, axis=1)
        df['11.1.2.b<=11.1.2.a'] = df.apply(res78, axis=1)
        df['11.1.2.c<=11.1.2.a'] = df.apply(res79, axis=1)
        df['14.18>=14.8.6'] = df.apply(res80, axis=1)
        df['14.9.1<=14.3.1.a+14.3.1.b'] = df.apply(res81, axis=1)
        df['14.9.2<=14.3.2.a+14.3.2.b'] = df.apply(res82, axis=1)
        df['14.13<=14.12.1+14.12.2+14.12.3+14.12.4+14.12.5'] = df.apply(res83, axis=1)
        df['14.8.2<=14.8.1'] = df.apply(res84, axis=1)
        df['14.8.3<=14.8.1'] = df.apply(res85, axis=1)
        df['15.3.4.b<=15.3.4.a'] = df.apply(res86, axis=1)
        df['15.3.4.d<=15.3.4.c'] = df.apply(res87, axis=1)



        # Merging all the renamed columns
        #==================================================

        df = pd.concat([df['1.1.1<=1.1'],
                        df ['15.3.1.b<=15.3.1.a'],
                        df ['1.2.4<=1.1'],
                        df ['1.2.5<=1.1'],
                        df ['1.2.7<=1.1'],
                        df ['1.3.1.a<=1.3.1'],
                            df ['1.3.2<=2.1'],
                            df ['1.4.4>=1.4.3'],
                            df ['1.5.1<=1.1'],
                            df ['1.5.2<=1.5.1'],
                            df ['1.5.3<=1.5.2'],
                                df ['1.6.1.a<=1.1'],
                                df ['1.6.1.b<=1.6.1.a'],
                                df ['1.6.1.c<=1.6.1.b'],
                                df ['1.6.1.e<=1.6.1.d'],
                                df ['2.1.1<=2.1'],
                                    df ['3.1<=2.1'],
                                    df ['3.1.1<=3.1'],
                                    df ['4.1.1.a+4.1.1.b+4.1.3>=2.1'],
                                    df ['4.1.2<=4.1.1.a+4.1.1.b'],
                                    df ['4.3.2.a<=4.3.1.a+4.3.1.b+4.2'],
                                    df ['4.3.2.b<=4.3.2.a'],
                                        df ['4.3.3<=4.3.1.a+4.3.1.b+4.2'],
                                        df ['4.4.1<=4.1.1.a+4.1.1.b'],
                                        df ['4.4.2<=4.4.1'],
                                        df ['4.4.3<=4.1.1.a+4.1.1.b'],
                                        df ['5.2<=5.1'],
                                        df ['6.3<=2.1'],
                                        df ['6.4<=2.1'],
                                        df ['7.2.1<=7.1.1'],
                                        df ['7.2.2<=7.1.2'],
                                            df ['8.2.3<=2.1'],
                                            df ['8.4<=2.1'],
                                            df ['8.7<=8.3+8.4+8.5'],
                                            df ['8.17.1<=8.1.1'],
                                            df ['8.17.2<=8.2.1+8.2.2+8.2.3+8.2.4'],
                                            df ['9.1.1<=4.1.1.a+4.1.1.b'],
                                            df ['9.1.2<=4.1.1.a+4.1.1.b'],
                                            df ['9.1.9<=4.1.1.a+4.1.1.b'],
                                            df ['9.1.13<=4.1.1.a+4.1.1.b'],
                                                df ['9.2.4.a+9.2.4.b<=9.2.1+ 9.2.2'],
                                                df ['11.2.2<=11.2.1'],
                                                df ['11.4.2<=11.4.1'],
                                                df ['12.1.2.a<=12.1.1.a'],
                                                df ['12.1.2.b<=12.1.1.b'],
                                                df ['12.1.3.a<=12.1.1.a'],
                                                df ['12.1.3.b<=12.1.1.b'],
                                                df ['14.2.1+14.2.2>=14.1.1+14.1.2+14.1.3+14.1.4+14.1.5+14.1.6+14.1.7+14.1.8+14.1.9'],
                                                df ['14.3.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                                    df ['14.4.1<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                                    df ['14.4.2<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                                    df ['14.4.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                                    df ['14.4.4<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                                    df ['14.4.5<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                                    df ['14.4.6<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                                    df ['14.4.7<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                                    df ['14.4.8<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                                    df ['14.6.1<=14.5'],
                                                    df ['14.6.2<=14.5'],
                                                    df ['14.6.3<=14.5'],
                                                        df ['14.6.4<=14.5'],
                                                        df ['14.6.5<=14.5'],
                                                        df ['14.6.6<=14.5'],
                                                        df ['14.7<=14.5'],
                                                        df ['14.14.2<=14.14.1'],
                                                        df ['15.2.2<=15.2.1'],
                                                        df ['15.3.2.b<=15.3.2.a'],
                                                        df ['15.3.3.b<=15.3.3.a'],
                                                        df ['15.3.3.c<=15.3.3.b'],
                                                        df ['15.4.2<=15.4.1'],
                                                            df ['9.6.1<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'],
                                                            df ['9.6.2<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'],
                                                            df ['9.6.3<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'],
                                                            df ['9.7.2<=9.7.1'],
                                                            df ['9.7.3<=9.7.2'],
                                                            df ['11.1.1.b<=11.1.1.a'],
                                                            df ['11.1.1.c<=11.1.1.a'],
                                                            df ['11.1.2.b<=11.1.2.a'],
                                                            df ['11.1.2.c<=11.1.2.a'],
                                                            df ['14.18>=14.8.6'],
                                                            df ['14.9.1<=14.3.1.a+14.3.1.b'],
                                                            df ['14.9.2<=14.3.2.a+14.3.2.b'],
                                                                df ['14.13<=14.12.1+14.12.2+14.12.3+14.12.4+14.12.5'],
                                                                df ['14.8.2<=14.8.1'],
                                                                df ['14.8.3<=14.8.1'],
                                                                df ['15.3.4.b<=15.3.4.a'],
                                                                df ['15.3.4.d<=15.3.4.c']], axis=1)     


        # Mergining current result of modified checks with original dataframe and displaying it on screen
        frames = [df_, df]
        print(frames)
        df = pd.concat(frames, axis=1, sort=False)
        #df = df.dropna(axis=0, subset=['col_2'])
        self.tableView.setModel(PandasModel(df))

        msg = QMessageBox()
        msg.setWindowTitle(" Validation Completion Message")
        msg.setText("District Hospital Validation Complete")
        msg.setIcon(QMessageBox.Information)
        msg.exec()

        return df


    #############################################################################################
    # Public Health Center Validation Rules Function
    def PHC_Validate(self):
        global df

        df = self.loadFile(df_)

        filterString = self.comboBox.currentText()
        
        df = df_.loc[df_['col_12'] == filterString]
        print(df)

        print('Entered PHC_Validate')

        # Modified Checks of PHC
        
        # 9.1.1(103) <= 4.1.1.a(56) + 4.1.1.b(57)
        def res1(df):
            if pd.isnull(df['col_103']) and pd.isnull(df['col_56']) and pd.isnull(df['col_57']):
                return 'Blank'
            elif pd.isnull(df['col_103']) or pd.isnull(df['col_56']) or pd.isnull(df['col_57']):
                if pd.isnull(df['col_59']):
                    return 'Probable Reporting Error(9.1.1 is blank)'
                elif pd.isnull(float(df['col_56']) + float(df['col_57'])):
                    return 'Inconsistent'
            elif float(df['col_103']) > float(df['col_56']) + float(df['col_57']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.1.1(23) <= 1.1(22)
        def res2(df):
            if pd.isnull(df['col_23']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_23']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_23']):
                    return 'Probable Reporting Error(1.1.1 is blank)'
                elif pd.isnull(df['col_22']):
                    return 'Inconsistent'
            elif float(df['col_23']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.2.4(27) <= 1.1(22)
        def res3(df):
            if pd.isnull(df['col_27']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_27']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_27']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_27'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # 1.2.5(28) <= 1.1(22)
        def res4(df):
            if pd.isnull(df['col_28']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_28']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_28']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_28'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # 1.2.7(30) <= 1.1(22)
        def res5(df):
            if pd.isnull(df['col_30']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_30']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_30']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_30'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # 2.1.2(49) <= 2.1.1.a(47) + 2.1.1.b(48)
        def res6(df):
            if pd.isnull(df['col_49']) and pd.isnull(df['col_47']) and pd.isnull(df['col_48']):
                return 'Blank'
            elif pd.isnull(df['col_49']) or pd.isnull(df['col_47']) or pd.isnull(df['col_48']):
                if pd.isnull(df['col_49']):
                    return 'Probable Reporting Error(2.1.2 is blank)'
                elif pd.isnull(float(df['col_47']) + float(df['col_48'])):
                    return 'Inconsistent'
            elif float(df['col_49']) > float(df['col_47']) + float(df['col_48']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 2.1.3(50) <= 2.1.1.a(47) + 2.1.1.b(48)
        def res7(df):
            if pd.isnull(df['col_50']) and pd.isnull(df['col_47']) and pd.isnull(df['col_48']):
                return 'Blank'
            elif pd.isnull(df['col_50']) or pd.isnull(df['col_47']) or pd.isnull(df['col_48']):
                if pd.isnull(df['col_50']) and not pd.isnull(float(df['col_47'])) and pd.isnull(float(df['col_48'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_47'])
                rhs_value = float(df['col_48'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # 2.2.1(52) <= 2.2(51)
        def res8(df):
            if pd.isnull(df['col_52']) and pd.isnull(df['col_51']):
                return 'Blank'
            elif pd.isnull(df['col_52']) or pd.isnull(df['col_51']):
                if pd.isnull(df['col_52']):
                    return 'Probable Reporting Error(2.2.1 is blank)'
                elif pd.isnull(df['col_56']):
                    return 'Inconsistent'
            elif float(df['col_103']) > float(df['col_56']) + float(df['col_57']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 2.2(51) >= 1.3.2(34)
        def res9(df):
            if pd.isnull(df['col_51']) and pd.isnull(df['col_34']):
                return 'Blank'
            elif pd.isnull(df['col_51']) or pd.isnull(df['col_34']):
                if pd.isnull(df['col_51']):
                    return 'Probable Reporting Error(2.2 is blank)'
                elif pd.isnull(df['col_34']):
                    return 'Inconsistent'
            elif float(df['col_51']) < float(df['col_34']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.4.4(38) >= 1.4.3(37)
        def res10(df):
            if pd.isnull(df['col_38']) and pd.isnull(df['col_37']):
                return 'Blank'
            elif pd.isnull(df['col_38']) or pd.isnull(df['col_37']):
                if pd.isnull(df['col_38']) and not pd.isnull(float(df['col_37'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_38'])
                rhs_value = float(df['col_37'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # 1.5.1(39) <= 1.1(22)
        def res11(df):
            if pd.isnull(df['col_39']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_39']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_39']):
                    return 'Probable Reporting Error(1.5.1 is blank)'
                elif pd.isnull(df['col_22']):
                    return 'Inconsistent'
            elif float(df['col_39']) < float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.5.2(40) <= 1.5.1(39)
        def res12(df):
            if pd.isnull(df['col_40']) and pd.isnull(df['col_39']):
                return 'Blank'
            elif pd.isnull(df['col_40']) or pd.isnull(df['col_39']):
                if pd.isnull(df['col_40']):
                    return 'Probable Reporting Error(1.5.1 is blank)'
                elif pd.isnull(df['col_39']):
                    return 'Inconsistent'
            elif float(df['col_40']) < float(df['col_39']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.5.3(41) <= 1.5.2(40)
        def res13(df):
            if pd.isnull(df['col_41']) and pd.isnull(df['col_40']):
                return 'Blank'
            elif pd.isnull(df['col_41']) or pd.isnull(df['col_40']):
                if pd.isnull(df['col_41']):
                    return 'Probable Reporting Error(1.5.3 is blank)'
                elif pd.isnull(df['col_40']):
                    return 'Inconsistent'
            elif float(df['col_41']) < float(df['col_40']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.6.1.a(42) <= 1.1(22)
        def res14(df):
            if pd.isnull(df['col_42']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_42']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_42']):
                    return 'Probable Reporting Error(1.6.1.a is blank)'
                elif pd.isnull(df['col_22']):
                    return 'Inconsistent'
            elif float(df['col_42']) < float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.6.1.b(43) <= 1.6.1.a(42)
        def res15(df):
            if pd.isnull(df['col_43']) and pd.isnull(df['col_42']):
                return 'Blank'
            elif pd.isnull(df['col_43']) or pd.isnull(df['col_42']):
                if pd.isnull(df['col_43']):
                    return 'Probable Reporting Error(1.6.1.a is blank)'
                elif pd.isnull(df['col_42']):
                    return 'Inconsistent'
            elif float(df['col_43']) < float(df['col_42']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.6.1.c(44) <= 1.6.1.b(43)
        def res16(df):
            if pd.isnull(df['col_44']) and pd.isnull(df['col_43']):
                return 'Blank'
            elif pd.isnull(df['col_44']) or pd.isnull(df['col_43']):
                if pd.isnull(df['col_44']):
                    return 'Probable Reporting Error(1.6.1.a is blank)'
                elif pd.isnull(df['col_43']):
                    return 'Inconsistent'
            elif float(df['col_44']) > float(df['col_43']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.6.1.e(46) <= 1.6.1.d(45)
        def res17(df):
            if pd.isnull(df['col_46']) and pd.isnull(df['col_45']):
                return 'Blank'
            elif pd.isnull(df['col_46']) or pd.isnull(df['col_45']):
                if pd.isnull(df['col_46']):
                    return 'Probable Reporting Error(1.6.1.e is blank)'
                elif pd.isnull(df['col_45']):
                    return 'Inconsistent'
            elif float(df['col_46']) > float(df['col_45']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 3.1.1(55) <= 2.2(51)
        def res18(df):
            if pd.isnull(df['col_55']) and pd.isnull(df['col_51']):
                return 'Blank'
            elif pd.isnull(df['col_55']) or pd.isnull(df['col_51']):
                if pd.isnull(df['col_55']):
                    return 'Probable Reporting Error(3.1.1 is blank)'
                elif pd.isnull(df['col_51']):
                    return 'Inconsistent'
            elif float(df['col_55']) < float(df['col_51']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 3.1(54) <= 2.2(51)
        def res19(df):
            if pd.isnull(df['col_54']) and pd.isnull(df['col_51']):
                return 'Blank'
            elif pd.isnull(df['col_54']) or pd.isnull(df['col_51']):
                if pd.isnull(df['col_54']):
                    return 'Probable Reporting Error(1.6.1.a is blank)'
                elif pd.isnull(df['col_51']):
                    return 'Inconsistent'
            elif float(df['col_54']) < float(df['col_51']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 4.1.2(58) <= 4.1.1.a(56) + 4.1.1.b(57)
        def res20(df):
            if pd.isnull(df['col_58']) and pd.isnull(df['col_56']) and pd.isnull(df['col_57']):
                return 'Blank'
            elif pd.isnull(df['col_58']) or pd.isnull(df['col_56']) or pd.isnull(df['col_57']):
                if pd.isnull(df['col_58']):
                    return 'Probable Reporting Error(4.1.2 is blank)'
                elif pd.isnull(float(df['col_56']) + float(df['col_57'])):
                    return 'Inconsistent'
            elif float(df['col_58']) > float(df['col_56']) + float(df['col_57']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 4.1.1.a(56) + 4.1.1.b(57) + 4.1.3(59) >= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)
        def res21(df):
            if pd.isnull(df['col_56']) and pd.isnull(df['col_57']) and pd.isnull(df['col_59']) and pd.isnull(df['col_47']) and pd.isnull(df['col_48']) and pd.isnull(df['col_51']):
                return 'Blank'
            elif pd.isnull(df['col_56']) or pd.isnull(df['col_57']) or pd.isnull(df['col_59']) or pd.isnull(df['col_47']) or pd.isnull(df['col_48']) or pd.isnull(df['col_51']):
                if pd.isnull(df['col_56']) + pd.isnull(df['col_57']) + pd.isnull(df['col_59']):
                    return 'Inconsistent'
                elif pd.isnull(float(df['col_47']) + float(df['col_48']) + float(df['col_51'])):
                    return 'Probable Reporting Error'
            elif float(df['col_56']) + float(df['col_57']) + float(df['col_59']) <  float(df['col_47']) + float(df['col_48']) + float(df['col_51']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 4.3.2.a(63) <= 4.3.1.a(61) + 4.3.1.b(62) + 4.2(60)
        def res22(df):
            if pd.isnull(df['col_63']) and pd.isnull(df['col_61']) and pd.isnull(df['col_62']) and pd.isnull(df['col_60']):
                return 'Blank'
            elif pd.isnull(df['col_63']) or pd.isnull(df['col_61']) or pd.isnull(df['col_62']) or pd.isnull(df['col_60']):
                if pd.isnull(df['col_63']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_61']) + float(df['col_62']) + float(df['col_60']))):
                    return 'Inconsistent'
            elif float(df['col_63']) > float(df['col_61']) + float(df['col_62']) + float(df['col_60']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 4.3.2.b(64) <= 4.3.2.a(63)
        def res23(df):
            if pd.isnull(df['col_64']) and pd.isnull(df['col_63']):
                return 'Blank'
            elif pd.isnull(df['col_64']) or pd.isnull(df['col_63']):
                if pd.isnull(df['col_64']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_63']):
                    return 'Inconsistent'
            elif float(df['col_64']) > float(df['col_63']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 4.3.3(65) <= 4.3.1.a(61) + 4.3.1.b(62) + 4.2(60)
        def res24(df):
            if pd.isnull(df['col_65']) and pd.isnull(df['col_61']) and pd.isnull(df['col_62']) and pd.isnull(df['col_60']):
                return 'Blank'
            elif pd.isnull(df['col_65']) or pd.isnull(df['col_61']) or pd.isnull(df['col_62']) or pd.isnull(df['col_60']):
                if pd.isnull(df['col_65']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_61']) + float(df['col_62']) + float(df['col_60']))):
                    return 'Inconsistent'
            elif float(df['col_65']) > float(df['col_61']) + float(df['col_62']) + float(df['col_60']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 4.4.1(66) <= 4.1.1.a(56) + 4.1.1.b(57)
        def res25(df):
            if pd.isnull(df['col_66']) and pd.isnull(df['col_56']) and pd.isnull(df['col_57']):
                return 'Blank'
            elif pd.isnull(df['col_66']) or pd.isnull(df['col_56']) or pd.isnull(df['col_57']):
                if pd.isnull(df['col_66']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_56']) + float(df['col_57']))):
                    return 'Inconsistent'
            elif float(df['col_66']) > float(df['col_56']) + float(df['col_57']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 4.4.2(67) <= 4.4.1(66)
        def res26(df):
            if pd.isnull(df['col_67']) and pd.isnull(df['col_66']):
                return 'Blank'
            elif pd.isnull(df['col_67']) or pd.isnull(df['col_66']):
                if pd.isnull(df['col_67']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_66']):
                    return 'Inconsistent'
            elif float(df['col_67']) > float(df['col_66']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 4.4.3(68) <= 4.1.1.a(56) + 4.1.1.b(57)
        def res27(df):
            if pd.isnull(df['col_68']) and pd.isnull(df['col_56']) and pd.isnull(df['col_57']):
                return 'Blank'
            elif pd.isnull(df['col_68']) or pd.isnull(df['col_56']) or pd.isnull(df['col_57']):
                if pd.isnull(df['col_68']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_56']) + float(df['col_57']))):
                    return 'Inconsistent'
            elif float(df['col_68']) > float(df['col_56']) + float(df['col_57']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 6.1(70) <= 2.1.1.a(47) + 2.1.1.b(48)
        def res28(df):
            if pd.isnull(df['col_70']) and pd.isnull(df['col_47']) and pd.isnull(df['col_48']):
                return 'Blank'
            elif pd.isnull(df['col_70']) or pd.isnull(df['col_47']) or pd.isnull(df['col_48']):
                if pd.isnull(df['col_70']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_47'])) + (float(df['col_48']))):
                    return 'Inconsistent'
            elif float(df['col_70']) > float(df['col_47']) + float(df['col_48']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 6.3(72) <= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)
        def res29(df):
            if pd.isnull(df['col_72']) and pd.isnull(df['col_47']) and pd.isnull(df['col_48']) and pd.isnull(df['col_51']):
                return 'Blank'
            elif pd.isnull(df['col_72']) or pd.isnull(df['col_47']) or pd.isnull(df['col_48']) or pd.isnull(df['col_51']):
                if pd.isnull(df['col_72']) and not pd.isnull(float(df['col_47'])) and pd.isnull(float(df['col_48'])) and pd.isnull(float(df['col_51'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_72'])
                rhs_value = float(df['col_47']) + float(df['col_48']) + float(df['col_51']) 

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # 6.4(73) <= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)
        def res30(df):
            if pd.isnull(df['col_73']) and pd.isnull(df['col_47']) and pd.isnull(df['col_48']) and pd.isnull(df['col_51']):
                return 'Blank'
            elif pd.isnull(df['col_73']) or pd.isnull(df['col_47']) or pd.isnull(df['col_48']) or pd.isnull(df['col_51']):
                if pd.isnull(df['col_73']) and not pd.isnull(float(df['col_47'])) and pd.isnull(float(df['col_48'])) and pd.isnull(float(df['col_51'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_73'])
                rhs_value = float(df['col_47']) + float(df['col_48']) + float(df['col_51']) 

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # 7.2.1(76) <= 7.1.1(74)
        def res31(df):
            if pd.isnull(df['col_76']) and pd.isnull(df['col_74']):
                return 'Blank'
            elif pd.isnull(df['col_76']) or pd.isnull(df['col_74']):
                if pd.isnull(df['col_76']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_74']):
                    return 'Inconsistent'
            elif float(df['col_76']) > float(df['col_74']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 7.2.2(77) <= 7.1.2(75)
        def res32(df):
            if pd.isnull(df['col_77']) and pd.isnull(df['col_75']):
                return 'Blank'
            elif pd.isnull(df['col_77']) or pd.isnull(df['col_75']):
                if pd.isnull(df['col_77']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_75']):
                    return 'Inconsistent'
            elif float(df['col_77']) > float(df['col_75']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 8.2.3(81) <= 2.2(51)
        def res33(df):
            if pd.isnull(df['col_81']) and pd.isnull(df['col_51']):
                return 'Blank'
            elif pd.isnull(df['col_81']) or pd.isnull(df['col_51']):
                if pd.isnull(df['col_81']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_51']):
                    return 'Inconsistent'
            elif float(df['col_81']) > float(df['col_51']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 8.4(84) <= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)
        def res34(df):
            if pd.isnull(df['col_84']) and pd.isnull(df['col_61']) and pd.isnull(df['col_62']) and pd.isnull(df['col_60']):
                return 'Blank'
            elif pd.isnull(df['col_84']) or pd.isnull(df['col_61']) or pd.isnull(df['col_62']) or pd.isnull(df['col_60']):
                if pd.isnull(df['col_84']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_61']) + float(df['col_62']) + float(df['col_60']))):
                    return 'Inconsistent'
            elif float(df['col_84']) > float(df['col_61']) + float(df['col_62']) + float(df['col_60']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 8.7(87) <= 8.3(83) + 8.4(84) + 8.5(85)
        def res35(df):
            if pd.isnull(df['col_87']) and pd.isnull(df['col_83']) and pd.isnull(df['col_84']) and pd.isnull(df['col_85']):
                return 'Blank'
            elif pd.isnull(df['col_87']) or pd.isnull(df['col_83']) or pd.isnull(df['col_84']) or pd.isnull(df['col_85']):
                if pd.isnull(df['col_87']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_83']) + float(df['col_84']) + float(df['col_85']))):
                    return 'Inconsistent'
            elif float(df['col_87']) > float(df['col_83']) + float(df['col_84']) + float(df['col_85']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 8.17.1(97) <= 8.1.1(78)
        def res36(df):
            if pd.isnull(df['col_97']) and pd.isnull(df['col_78']):
                return 'Blank'
            elif pd.isnull(df['col_97']) or pd.isnull(df['col_78']):
                if pd.isnull(df['col_97']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_78']):
                    return 'Inconsistent'
            elif float(df['col_97']) > float(df['col_78']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 8.17.2(98) <= 8.2.1(79) + 8.2.2(80) + 8.2.3(81) + 8.2.4(82)
        def res37(df):
            if pd.isnull(df['col_98']) and pd.isnull(df['col_79']) and pd.isnull(df['col_80']) and pd.isnull(df['col_81']) and pd.isnull(df['col_82']):
                return 'Blank'
            elif pd.isnull(df['col_98']) or pd.isnull(df['col_79']) or pd.isnull(df['col_80']) or pd.isnull(df['col_81']) or pd.isnull(df['col_82']):
                if pd.isnull(df['col_98']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_79']) + float(df['col_80']) + float(df['col_81']) + float(df['col_82']))):
                    return 'Inconsistent'
            elif float(df['col_98']) > float(df['col_79']) + float(df['col_80']) + float(df['col_81']) + float(df['col_82']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 9.1.9(111) <= 4.1.1.a(56) + 4.1.1.b(57)
        def res38(df):
            if pd.isnull(df['col_111']) and pd.isnull(df['col_56']) and pd.isnull(df['col_57']):
                return 'Blank'
            elif pd.isnull(df['col_111']) or pd.isnull(df['col_56']) or pd.isnull(df['col_57']):
                if pd.isnull(df['col_111']):
                    return 'Probable Reporting Error'
                elif pd.isnull(float(df['col_56']) + float(df['col_57'])):
                    return 'Inconsistent'
            elif float(df['col_111']) > float(df['col_56']) + float(df['col_57']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 9.1.13(115) <= 4.1.1.a(56) + 4.1.1.b(57)
        def res39(df):
            if pd.isnull(df['col_115']) and pd.isnull(df['col_56']) and pd.isnull(df['col_57']):
                return 'Blank'
            elif pd.isnull(df['col_115']) or pd.isnull(df['col_56']) or pd.isnull(df['col_57']):
                if pd.isnull(df['col_115']):
                    return 'Probable Reporting Error'
                elif pd.isnull(float(df['col_56']) + float(df['col_57'])):
                    return 'Inconsistent'
            elif float(df['col_115']) > float(df['col_56']) + float(df['col_57']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 9.2.4.a(127) + 9.2.4.b(128) <= 9.2.1(124) + 9.2.2(125)
        def res40(df):
            if pd.isnull(df['col_127']) and pd.isnull(df['col_128']) and pd.isnull(df['col_124']) and pd.isnull(df['col_125']):
                return 'Blank'
            elif pd.isnull(df['col_127']) or pd.isnull(df['col_128']) or pd.isnull(df['col_124']) or pd.isnull(df['col_125']):
                if pd.isnull(float(df['col_127']) + float(df['col_128'])):
                    return 'Probable Reporting Error'
                elif pd.isnull(float(df['col_124']) + float(df['col_125'])):
                    return 'Inconsistent'
            elif float(df['col_127']) + float(df['col_128']) > float(df['col_124']) + float(df['col_125']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 11.2.2(175) <= 11.2.1(174)
        def res41(df):
            if pd.isnull(df['col_175']) and pd.isnull(df['col_174']):
                return 'Blank'
            elif pd.isnull(df['col_175']) or pd.isnull(df['col_174']):
                if pd.isnull(df['col_175']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_174']):
                    return 'Inconsistent'
            elif float(df['col_175']) > float(df['col_174']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 12.1.2.a(180) <= 12.1.1.a(178)
        def res42(df):
            if pd.isnull(df['col_180']) and pd.isnull(df['col_178']):
                return 'Blank'
            elif pd.isnull(df['col_180']) or pd.isnull(df['col_178']):
                if pd.isnull(df['col_180']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_178']):
                    return 'Inconsistent'
            elif float(df['col_180']) > float(df['col_178']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 12.1.2.b(181) <= 12.1.1.b(179)
        def res43(df):
            if pd.isnull(df['col_181']) and pd.isnull(df['col_179']):
                return 'Blank'
            elif pd.isnull(df['col_181']) or pd.isnull(df['col_179']):
                if pd.isnull(df['col_181']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_179']):
                    return 'Inconsistent'
            elif float(df['col_181']) > float(df['col_179']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 12.1.3.a(182) <= 12.1.1.a(178)
        def res44(df):
            if pd.isnull(df['col_182']) and pd.isnull(df['col_178']):
                return 'Blank'
            elif pd.isnull(df['col_182']) or pd.isnull(df['col_178']):
                if pd.isnull(df['col_182']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_178']):
                    return 'Inconsistent'
            elif float(df['col_182']) > float(df['col_178']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 12.1.3.b(183) <= 12.1.1.b(179)
        def res45(df):
            if pd.isnull(df['col_183']) and pd.isnull(df['col_179']):
                return 'Blank'
            elif pd.isnull(df['col_183']) or pd.isnull(df['col_179']):
                if pd.isnull(df['col_183']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_179']):
                    return 'Inconsistent'
            elif float(df['col_183']) > float(df['col_179']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.2.1(194) + 14.2.2(195) >= 14.1.1(186) + 14.1.2(187) + 14.1.3(188) + 14.1.4(189) + 14.1.5(190) + 14.1.6(191) + 14.1.7(192) + 14.1.8(193)
        def res46(df):
            if pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_186']) and pd.isnull(df['col_187']) and pd.isnull(df['col_188']) and pd.isnull(df['col_189']) and pd.isnull(df['col_190']) and pd.isnull(df['col_191']) and pd.isnull(df['col_192']) and pd.isnull(df['col_193']):
                return 'Blank'
            elif pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_186']) or pd.isnull(df['col_187']) or pd.isnull(df['col_188']) or pd.isnull(df['col_189']) or pd.isnull(df['col_190']) or pd.isnull(df['col_191']) or pd.isnull(df['col_192']) or pd.isnull(df['col_193']):
                if pd.isnull(float(df['col_194']) + float(df['col_195'])):
                    return 'Inconsistent'
                elif pd.isnull(float(df['col_186']) + float(df['col_187']) + float(df['col_188']) + float(df['col_189']) + float(df['col_190']) + float(df['col_191']) + float(df['col_192']) + float(df['col_193'])):
                    return 'Probable Reporting Error'
            elif float(df['col_194']) + float(df['col_195']) < float(df['col_186']) + float(df['col_187']) + float(df['col_188']) + float(df['col_189']) + float(df['col_190']) + float(df['col_191']) + float(df['col_192']) + float(df['col_193']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.3.3(200) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)
        def res47(df):
            if pd.isnull(df['col_200']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_200']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_82']):
                if pd.isnull(df['col_200']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_200']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.5.2(210) <= 14.5.1(209)
        def res48(df):
            if pd.isnull(df['col_210']) and pd.isnull(df['col_209']):
                return 'Blank'
            elif pd.isnull(df['col_210']) or pd.isnull(df['col_209']):
                if pd.isnull(df['col_210']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_209']):
                    return 'Inconsistent'
            elif float(df['col_210']) > float(df['col_209']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 15.3.3.b(230) <= 15.3.3.a(229)
        def res49(df):
            if pd.isnull(df['col_230']) and pd.isnull(df['col_229']):
                return 'Blank'
            elif pd.isnull(df['col_230']) or pd.isnull(df['col_229']):
                if pd.isnull(df['col_230']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_229']):
                    return 'Inconsistent'
            elif float(df['col_230']) > float(df['col_229']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 15.3.3.c(231) <= 15.3.3.b(230)
        def res50(df):
            if pd.isnull(df['col_231']) and pd.isnull(df['col_230']):
                return 'Blank'
            elif pd.isnull(df['col_231']) or pd.isnull(df['col_230']):
                if pd.isnull(df['col_231']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_230']):
                    return 'Inconsistent'
            elif float(df['col_231']) > float(df['col_230']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 15.4.2(237) <= 15.4.1(236)
        def res51(df):
            if pd.isnull(df['col_237']) and pd.isnull(df['col_236']):
                return 'Blank'
            elif pd.isnull(df['col_237']) or pd.isnull(df['col_236']):
                if pd.isnull(df['col_237']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_236']):
                    return 'Inconsistent'
            elif float(df['col_237']) > float(df['col_236']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 9.6.1(142) <= 9.1.1(103) + 9.1.2(104) + 9.1.3(105) + 9.1.4(106) + 9.1.5(107) + 9.1.6(108) +9.1.7(109) + 9.1.8(110) + 9.1.13(115) + 9.1.14(116) + 9.1.15(117) + 9.1.16(118) + 9.1.17(119) + 9.1.18(120) + 9.1.19(121) + 9.1.20(122) + 9.1.21(123) + 9.2.1(124) + 9.2.2(125) + 9.2.3(126) + 9.3.1(129) + 9.3.2(130) + 9.3.3(131) + 9.4.1(132) + 9.4.2(133) + 9.4.3(134) + 9.4.5(136) + 9.4.6(137) + 9.5.1(138) + 9.5.2(139) + 9.5.3(140) + 9.5.4(141)
        def res52(df):
            if pd.isnull(df['col_142']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_108']) and pd.isnull(df['col_109']) and pd.isnull(df['col_110']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_123']) and pd.isnull(df['col_124']) and pd.isnull(df['col_125']) and pd.isnull(df['col_126']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_131']) and pd.isnull(df['col_132']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']) and pd.isnull(df['col_138']) and pd.isnull(df['col_139']) and pd.isnull(df['col_140']) and pd.isnull(df['col_141']) :
                return 'Blank'
            elif pd.isnull(df['col_142']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_108']) or pd.isnull(df['col_109']) or pd.isnull(df['col_110']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_123']) or pd.isnull(df['col_124']) or pd.isnull(df['col_125']) or pd.isnull(df['col_126']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_131']) or pd.isnull(df['col_132']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']) or pd.isnull(df['col_138']) or pd.isnull(df['col_139']) or pd.isnull(df['col_140']) or pd.isnull(df['col_141']) :
                if pd.isnull(df['col_142']):
                    return 'Probable Reporting Error'
                elif pd.isnull(float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_124']) + float(df['col_125']) + float(df['col_126']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138']) + float(df['col_139']) + float(df['col_140']) + float(df['col_141'])):
                    return 'Inconsistent'
            elif float(df['col_142']) > float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_124']) + float(df['col_125']) + float(df['col_126']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138']) + float(df['col_139']) + float(df['col_140']) + float(df['col_141']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df


        # 9.6.2(143) <= 9.1.1(103) + 9.1.2(104) + 9.1.3(105) + 9.1.4(106) + 9.1.5(107) + 9.1.6(108) +9.1.7(109) + 9.1.8(110) + 9.1.13(115) + 9.1.14(116) + 9.1.15(117) + 9.1.16(118) + 9.1.17(119) + 9.1.18(120) + 9.1.19(121) + 9.1.20(122) + 9.1.21(123) + 9.2.1(124) + 9.2.2(125) + 9.2.3(126) + 9.3.1(129) + 9.3.2(130) + 9.3.3(131) + 9.4.1(132) + 9.4.2(133) + 9.4.3(134) + 9.4.5(136) + 9.4.6(137) + 9.5.1(138) + 9.5.2(139) + 9.5.3(140) + 9.5.4(141)
        def res53(df):
            if pd.isnull(df['col_143']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_108']) and pd.isnull(df['col_109']) and pd.isnull(df['col_110']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_123']) and pd.isnull(df['col_124']) and pd.isnull(df['col_125']) and pd.isnull(df['col_126']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_131']) and pd.isnull(df['col_132']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']) and pd.isnull(df['col_138']) and pd.isnull(df['col_139']) and pd.isnull(df['col_140']) and pd.isnull(df['col_141']) :
                return 'Blank'
            elif pd.isnull(df['col_143']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_108']) or pd.isnull(df['col_109']) or pd.isnull(df['col_110']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_123']) or pd.isnull(df['col_124']) or pd.isnull(df['col_125']) or pd.isnull(df['col_126']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_131']) or pd.isnull(df['col_132']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']) or pd.isnull(df['col_138']) or pd.isnull(df['col_139']) or pd.isnull(df['col_140']) or pd.isnull(df['col_141']) :
                if pd.isnull(df['col_143']):
                    return 'Probable Reporting Error'
                elif pd.isnull(float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_124']) + float(df['col_125']) + float(df['col_126']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138']) + float(df['col_139']) + float(df['col_140']) + float(df['col_141'])):
                    return 'Inconsistent'
            elif float(df['col_143']) > float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_124']) + float(df['col_125']) + float(df['col_126']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138']) + float(df['col_139']) + float(df['col_140']) + float(df['col_141']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 9.6.3(144) <= 9.1.1(103) + 9.1.2(104) + 9.1.3(105) + 9.1.4(106) + 9.1.5(107) + 9.1.6(108) +9.1.7(109) + 9.1.8(110) + 9.1.13(115) + 9.1.14(116) + 9.1.15(117) + 9.1.16(118) + 9.1.17(119) + 9.1.18(120) + 9.1.19(121) + 9.1.20(122) + 9.1.21(123) + 9.2.1(124) + 9.2.2(125) + 9.2.3(126) + 9.3.1(129) + 9.3.2(130) + 9.3.3(131) + 9.4.1(132) + 9.4.2(133) + 9.4.3(134) + 9.4.5(136) + 9.4.6(137) + 9.5.1(138) + 9.5.2(139) + 9.5.3(140) + 9.5.4(141)
        def res54(df):
            if pd.isnull(df['col_144']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_108']) and pd.isnull(df['col_109']) and pd.isnull(df['col_110']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_123']) and pd.isnull(df['col_124']) and pd.isnull(df['col_125']) and pd.isnull(df['col_126']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_131']) and pd.isnull(df['col_132']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']) and pd.isnull(df['col_138']) and pd.isnull(df['col_139']) and pd.isnull(df['col_140']) and pd.isnull(df['col_141']) :
                return 'Blank'
            elif pd.isnull(df['col_144']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_108']) or pd.isnull(df['col_109']) or pd.isnull(df['col_110']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_123']) or pd.isnull(df['col_124']) or pd.isnull(df['col_125']) or pd.isnull(df['col_126']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_131']) or pd.isnull(df['col_132']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']) or pd.isnull(df['col_138']) or pd.isnull(df['col_139']) or pd.isnull(df['col_140']) or pd.isnull(df['col_141']) :
                if pd.isnull(df['col_144']):
                    return 'Probable Reporting Error'
                elif pd.isnull(float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_124']) + float(df['col_125']) + float(df['col_126']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138']) + float(df['col_139']) + float(df['col_140']) + float(df['col_141'])):
                    return 'Inconsistent'
            elif float(df['col_144']) > float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_124']) + float(df['col_125']) + float(df['col_126']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138']) + float(df['col_139']) + float(df['col_140']) + float(df['col_141']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.3.1.a(33) <= 1.3.1(32)
        def res55(df):
            if pd.isnull(df['col_33']) and pd.isnull(df['col_32']):
                return 'Blank'
            elif pd.isnull(df['col_33']) or pd.isnull(df['col_32']):
                if pd.isnull(df['col_33']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_32']):
                    return 'Inconsistent'
            elif float(df['col_33']) > float(df['col_32']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 9.7.2(146) <= 9.7.1(145)
        def res56(df):
            if pd.isnull(df['col_146']) and pd.isnull(df['col_145']):
                return 'Blank'
            elif pd.isnull(df['col_146']) or pd.isnull(df['col_145']):
                if pd.isnull(df['col_146']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_145']):
                    return 'Inconsistent'
            elif float(df['col_146']) > float(df['col_145']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 9.7.3(147) <= 9.7.2(146)
        def res57(df):
            if pd.isnull(df['col_147']) and pd.isnull(df['col_146']):
                return 'Blank'
            elif pd.isnull(df['col_147']) or pd.isnull(df['col_146']):
                if pd.isnull(df['col_147']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_146']):
                    return 'Inconsistent'
            elif float(df['col_147']) > float(df['col_146']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 11.1.1.b(169) <= 11.1.1.a(168)
        def res58(df):
            if pd.isnull(df['col_169']) and pd.isnull(df['col_168']):
                return 'Blank'
            elif pd.isnull(df['col_169']) or pd.isnull(df['col_168']):
                if pd.isnull(df['col_169']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_168']):
                    return 'Inconsistent'
            elif float(df['col_169']) > float(df['col_168']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 11.1.1.c(170) <= 11.1.1.a(168)
        def res59(df):
            if pd.isnull(df['col_170']) and pd.isnull(df['col_168']):
                return 'Blank'
            elif pd.isnull(df['col_170']) or pd.isnull(df['col_168']):
                if pd.isnull(df['col_170']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_168']):
                    return 'Inconsistent'
            elif float(df['col_170']) > float(df['col_168']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 11.1.2.b(171) <= 11.1.1.a(168)
        def res60(df):
            if pd.isnull(df['col_170']) and pd.isnull(df['col_168']):
                return 'Blank'
            elif pd.isnull(df['col_170']) or pd.isnull(df['col_168']):
                if pd.isnull(df['col_170']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_168']):
                    return 'Inconsistent'
            elif float(df['col_170']) > float(df['col_168']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 11.1.2.c(173) <= 11.1.2.a(171)
        def res61(df):
            if pd.isnull(df['col_173']) and pd.isnull(df['col_171']):
                return 'Blank'
            elif pd.isnull(df['col_173']) or pd.isnull(df['col_171']):
                if pd.isnull(df['col_173']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_171']):
                    return 'Inconsistent'
            elif float(df['col_173']) > float(df['col_171']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.4.1(201) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)
        def res62(df):
            if pd.isnull(df['col_201']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_201']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_201']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_201']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.4.2(202) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)
        def res63(df):
            if pd.isnull(df['col_202']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_202']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_202']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_202']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.4.3(203) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)
        def res64(df):
            if pd.isnull(df['col_203']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_203']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_203']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_203']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.4.4(204) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)
        def res65(df):
            if pd.isnull(df['col_204']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_204']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_204']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_204']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.4.5(205) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)
        def res66(df):
            if pd.isnull(df['col_205']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_205']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_205']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_205']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.4.6(206) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)
        def res67(df):
            if pd.isnull(df['col_206']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_206']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_206']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_206']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.4.7(207) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199))
        def res68(df):
            if pd.isnull(df['col_207']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_207']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_207']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_207']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.4.8(208) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)
        def res69(df):
            if pd.isnull(df['col_208']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_208']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_208']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_208']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.6.1(214) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)
        def res70(df):
            if pd.isnull(df['col_214']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_214']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_214']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_214']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.6.2(215) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)
        def res71(df):
            if pd.isnull(df['col_215']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_215']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_215']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_215']) > (float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 14.9.2(219) <= 14.9.1(218)
        def res72(df):
            if pd.isnull(df['col_219']) and pd.isnull(df['col_218']):
                return 'Blank'
            elif pd.isnull(df['col_219']) or pd.isnull(df['col_218']):
                if pd.isnull(df['col_219']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_218']):
                    return 'Inconsistent'
            elif float(df['col_219']) > float(df['col_218']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 15.2.2(224) <= 15.2.1(223)
        def res73(df):
            if pd.isnull(df['col_224']) and pd.isnull(df['col_223']):
                return 'Blank'
            elif pd.isnull(df['col_224']) or pd.isnull(df['col_223']):
                if pd.isnull(df['col_224']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_223']):
                    return 'Inconsistent'
            elif float(df['col_224']) > float(df['col_223']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 15.3.1.b(226) <= 15.3.1.a(225)
        def res74(df):
            if pd.isnull(df['col_226']) and pd.isnull(df['col_225']):
                return 'Blank'
            elif pd.isnull(df['col_226']) or pd.isnull(df['col_225']):
                if pd.isnull(df['col_226']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_225']):
                    return 'Inconsistent'
            elif float(df['col_226']) > float(df['col_225']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 15.3.2.b(228) <= 15.3.2.a(227)
        def res75(df):
            if pd.isnull(df['col_228']) and pd.isnull(df['col_227']):
                return 'Blank'
            elif pd.isnull(df['col_228']) or pd.isnull(df['col_227']):
                if pd.isnull(df['col_228']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_227']):
                    return 'Inconsistent'
            elif float(df['col_228']) > float(df['col_227']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 15.3.4.b(233) <= 15.3.4.a(232)
        def res76(df):
            if pd.isnull(df['col_233']) and pd.isnull(df['col_232']):
                return 'Blank'
            elif pd.isnull(df['col_233']) or pd.isnull(df['col_232']):
                if pd.isnull(df['col_233']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_232']):
                    return 'Inconsistent'
            elif float(df['col_232']) > float(df['col_233']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 15.3.4.d(235) <= 15.3.4.c(234)
        def res77(df):
            if pd.isnull(df['col_235']) and pd.isnull(df['col_234']):
                return 'Blank'
            elif pd.isnull(df['col_235']) or pd.isnull(df['col_234']):
                if pd.isnull(df['col_235']):
                    return 'Probable Reporting Error'
                elif pd.isnull(df['col_234']):
                    return 'Inconsistent'
            elif float(df['col_235']) > float(df['col_234']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 9.1.2(104) <= 4.1.1.a(56) + 4.1.1.b(57)
        def res78(df):
            if pd.isnull(df['col_104']) and pd.isnull(df['col_56']) and pd.isnull(df['col_57']):
                return 'Blank'
            elif pd.isnull(df['col_104']) or pd.isnull(df['col_56']) or pd.isnull(df['col_57']):
                if pd.isnull(df['col_104']):
                    return 'Probable Reporting Error'
                elif pd.isnull((float(df['col_56']) + float(df['col_57']))):
                    return 'Inconsistent'
            elif float(df['col_104']) > (float(df['col_56']) + float(df['col_57'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df



        # To count summary of the Modified Checks
        # =======================================

        df['9.1.1(103) <= 4.1.1.a(56) + 4.1.1.b(57)'] = df.apply(res1, axis=1)
        df['1.1.1(23) <= 1.1(22)'] = df.apply(res2, axis=1)
        df['1.2.4(27) <= 1.1(22)'] = df.apply(res3, axis=1)
        df['1.2.5(28) <= 1.1(22)'] = df.apply(res4, axis=1)
        df['1.2.7(30) <= 1.1(22)'] = df.apply(res5, axis=1)
        df['2.1.2(49) <= 2.1.1.a(47) + 2.1.1.b(48)'] = df.apply(res6, axis=1)
        df['2.1.3(50) <= 2.1.1.a(47) + 2.1.1.b(48)'] = df.apply(res7, axis=1)
        df['2.2.1(52) <= 2.2(51)'] = df.apply(res8, axis=1)
        df['2.2(51) >= 1.3.2(34)'] = df.apply(res9, axis=1)
        df['1.4.4(38) >= 1.4.3(37)'] = df.apply(res10, axis=1)
        df['1.5.1(39) <= 1.1(22)'] = df.apply(res11, axis=1)
        df['1.5.2(40) <= 1.5.1(39)'] = df.apply(res12, axis=1)
        df['1.5.3(41) <= 1.5.2(40)'] = df.apply(res13, axis=1)
        df['1.6.1.a(42) <= 1.1(22)'] = df.apply(res14, axis=1)
        df['1.6.1.b(43) <= 1.6.1.a(42)'] = df.apply(res15, axis=1)
        df['1.6.1.c(44) <= 1.6.1.b(43))'] = df.apply(res16, axis=1)
        df['1.6.1.e(46) <= 1.6.1.d(45)'] = df.apply(res17, axis=1)
        df['3.1.1(55) <= 2.2(51)'] = df.apply(res18, axis=1)
        df['3.1(54) <= 2.2(51)'] = df.apply(res19, axis=1)
        df['4.1.2(58) <= 4.1.1.a(56) + 4.1.1.b(57)'] = df.apply(res20, axis=1)
        df['4.1.1.a(56) + 4.1.1.b(57) + 4.1.3(59) >= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)'] = df.apply(res21, axis=1)
        df['4.3.2.a(63) <= 4.3.1.a(61) + 4.3.1.b(62) + 4.2(60)'] = df.apply(res22, axis=1)
        df['4.3.2.b(64) <= 4.3.2.a(63)'] = df.apply(res23, axis=1)
        df['4.3.3(65) <= 4.3.1.a(61) + 4.3.1.b(62) + 4.2(60)'] = df.apply(res24, axis=1)
        df['4.4.1(66) <= 4.1.1.a(56) + 4.1.1.b(57)'] = df.apply(res25, axis=1)
        df['4.4.2(67) <= 4.4.1(66)'] = df.apply(res26, axis=1)
        df['4.4.3(68) <= 4.1.1.a(56) + 4.1.1.b(57)'] = df.apply(res27, axis=1)
        df['6.1(70) <= 2.1.1.a(47) + 2.1.1.b(48)'] = df.apply(res28, axis=1)
        df['6.3(72) <= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)'] = df.apply(res29, axis=1)
        df['6.4(73) <= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)'] = df.apply(res30, axis=1)
        df['7.2.1(76) <= 7.1.1(74)'] = df.apply(res31, axis=1)
        df['7.2.2(77) <= 7.1.2(75)'] = df.apply(res32, axis=1)
        df['8.2.3(81) <= 2.2(51)'] = df.apply(res33, axis=1)
        df['8.4(84) <= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)'] = df.apply(res34, axis=1)
        df['8.7(87) <= 8.3(83) + 8.4(84) + 8.5(85)'] = df.apply(res35, axis=1)
        df['8.17.1(97) <= 8.1.1(78)'] = df.apply(res36, axis=1)
        df['8.17.2(98) <= 8.2.1(79) + 8.2.2(80) + 8.2.3(81) + 8.2.4(82)'] = df.apply(res37, axis=1)
        df['9.1.9(111) <= 4.1.1.a(56) + 4.1.1.b(57)'] = df.apply(res38, axis=1)
        df['9.1.13(115) <= 4.1.1.a(56) + 4.1.1.b(57)'] = df.apply(res39, axis=1)
        df['9.2.4.a(127) + 9.2.4.b(128) <= 9.2.1(124) + 9.2.2(125)'] = df.apply(res40, axis=1)
        df['11.2.2(175) <= 11.2.1(174)'] = df.apply(res41, axis=1)
        df['12.1.2.a(180) <= 12.1.1.a(178)'] = df.apply(res42, axis=1)
        df['12.1.2.b(181) <= 12.1.1.b(179)'] = df.apply(res43, axis=1)
        df['12.1.3.a(182) <= 12.1.1.a(178)'] = df.apply(res44, axis=1)
        df['12.1.3.b(183) <= 12.1.1.b(179)'] = df.apply(res45, axis=1)
        df['14.2.1(194) + 14.2.2(195) >= 14.1.1(186) + 14.1.2(187) + 14.1.3(188) + 14.1.4(189) + 14.1.5(190) + 14.1.6(191) + 14.1.7(192) + 14.1.8(193)'] = df.apply(res46, axis=1)
        df['14.3.3(200) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'] = df.apply(res47, axis=1)
        df['14.5.2(210) <= 14.5.1(209)'] = df.apply(res48, axis=1)
        df['15.3.3.b(230) <= 15.3.3.a(229)'] = df.apply(res49, axis=1)
        df['15.3.3.c(231) <= 15.3.3.b(230))'] = df.apply(res50, axis=1)
        df['15.4.2(237) <= 15.4.1(236)'] = df.apply(res51, axis=1)
        df['9.6.1(142) <= 9.1.1(103) + 9.1.2(104) + 9.1.3(105) + 9.1.4(106) + 9.1.5(107) + 9.1.6(108) +9.1.7(109) + 9.1.8(110) + 9.1.13(115) + 9.1.14(116) + 9.1.15(117) + 9.1.16(118) + 9.1.17(119) + 9.1.18(120) + 9.1.19(121) + 9.1.20(122) + 9.1.21(123) + 9.2.1(124) + 9.2.2(125) + 9.2.3(126) + 9.3.1(129) + 9.3.2(130) + 9.3.3(131) + 9.4.1(132) + 9.4.2(133) + 9.4.3(134) + 9.4.5(136) + 9.4.6(137) + 9.5.1(138) + 9.5.2(139) + 9.5.3(140) + 9.5.4(141)'] = df.apply(res52, axis=1)
        df['9.6.2(143) <= 9.1.1(103) + 9.1.2(104) + 9.1.3(105) + 9.1.4(106) + 9.1.5(107) + 9.1.6(108) +9.1.7(109) + 9.1.8(110) + 9.1.13(115) + 9.1.14(116) + 9.1.15(117) + 9.1.16(118) + 9.1.17(119) + 9.1.18(120) + 9.1.19(121) + 9.1.20(122) + 9.1.21(123) + 9.2.1(124) + 9.2.2(125) + 9.2.3(126) + 9.3.1(129) + 9.3.2(130) + 9.3.3(131) + 9.4.1(132) + 9.4.2(133) + 9.4.3(134) + 9.4.5(136) + 9.4.6(137) + 9.5.1(138) + 9.5.2(139) + 9.5.3(140) + 9.5.4(141)'] = df.apply(res53, axis=1)
        df['9.6.3(144) <= 9.1.1(103) + 9.1.2(104) + 9.1.3(105) + 9.1.4(106) + 9.1.5(107) + 9.1.6(108) +9.1.7(109) + 9.1.8(110) + 9.1.13(115) + 9.1.14(116) + 9.1.15(117) + 9.1.16(118) + 9.1.17(119) + 9.1.18(120) + 9.1.19(121) + 9.1.20(122) + 9.1.21(123) + 9.2.1(124) + 9.2.2(125) + 9.2.3(126) + 9.3.1(129) + 9.3.2(130) + 9.3.3(131) + 9.4.1(132) + 9.4.2(133) + 9.4.3(134) + 9.4.5(136) + 9.4.6(137) + 9.5.1(138) + 9.5.2(139) + 9.5.3(140) + 9.5.4(141)'] = df.apply(res54, axis=1)
        df['1.3.1.a(33) <= 1.3.1(32)'] = df.apply(res55, axis=1)
        df['9.7.2(146) <= 9.7.1(145)'] = df.apply(res56, axis=1)
        df['9.7.3(147) <= 9.7.2(146)'] = df.apply(res57, axis=1)
        df['11.1.1.b(169) <= 11.1.1.a(168)'] = df.apply(res58, axis=1)
        df['11.1.1.c(170) <= 11.1.1.a(168)'] = df.apply(res59, axis=1)
        df['11.1.2.b(171) <= 11.1.1.a(168)'] = df.apply(res60, axis=1)
        df['11.1.2.c(173) <= 11.1.2.a(171)'] = df.apply(res61, axis=1)
        df['14.4.1(201) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'] = df.apply(res62, axis=1)
        df['14.4.2(202) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199))'] = df.apply(res63, axis=1)
        df['14.4.3(203) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'] = df.apply(res64, axis=1)
        df['14.4.4(204) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'] = df.apply(res65, axis=1)
        df['14.4.5(205) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'] = df.apply(res66, axis=1)
        df['14.4.6(206) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'] = df.apply(res67, axis=1)
        df['14.4.7(207) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'] = df.apply(res68, axis=1)
        df['14.4.8(208) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'] = df.apply(res69, axis=1)
        df['14.6.1(214) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'] = df.apply(res70, axis=1)
        df['14.6.2(215) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'] = df.apply(res71, axis=1)
        df['14.9.2(219) <= 14.9.1(218)'] = df.apply(res72, axis=1)
        df['15.2.2(224) <= 15.2.1(223)'] = df.apply(res73, axis=1)
        df['15.3.1.b(226) <= 15.3.1.a(225)'] = df.apply(res74, axis=1)
        df['15.3.2.b(228) <= 15.3.2.a(227))'] = df.apply(res75, axis=1)
        df['15.3.4.b(233) <= 15.3.4.a(232)'] = df.apply(res76, axis=1)
        df['15.3.4.d(235) <= 15.3.4.c(234)'] = df.apply(res77, axis=1)
        df['9.1.2(104) <= 4.1.1.a(56) + 4.1.1.b(57)'] = df.apply(res78, axis=1)

        # Concatenating above renamed columns
        # ===================================
        df = pd.concat([df['9.1.1(103) <= 4.1.1.a(56) + 4.1.1.b(57)'],
                        df['1.1.1(23) <= 1.1(22)'],
                        df['1.2.4(27) <= 1.1(22)'],
                        df['1.2.5(28) <= 1.1(22)'],
                        df['1.2.7(30) <= 1.1(22)'],
                        df['2.1.2(49) <= 2.1.1.a(47) + 2.1.1.b(48)'],
                            df['2.1.3(50) <= 2.1.1.a(47) + 2.1.1.b(48)'],
                            df['2.2.1(52) <= 2.2(51)'],
                            df['2.2(51) >= 1.3.2(34)'],
                            df['1.4.4(38) >= 1.4.3(37)'],
                            df['1.5.1(39) <= 1.1(22)'],
                            df['1.5.2(40) <= 1.5.1(39)'],
                            df['1.5.3(41) <= 1.5.2(40)'],
                            df['1.6.1.a(42) <= 1.1(22)'],
                            df['1.6.1.b(43) <= 1.6.1.a(42)'],
                                df['1.6.1.c(44) <= 1.6.1.b(43))'],
                                df['1.6.1.e(46) <= 1.6.1.d(45)'],
                                df['3.1.1(55) <= 2.2(51)'],
                                df['3.1(54) <= 2.2(51)'],
                                df['4.1.2(58) <= 4.1.1.a(56) + 4.1.1.b(57)'],
                                df['4.1.1.a(56) + 4.1.1.b(57) + 4.1.3(59) >= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)'],
                                df['4.3.2.a(63) <= 4.3.1.a(61) + 4.3.1.b(62) + 4.2(60)'],
                                df['4.3.2.b(64) <= 4.3.2.a(63)'],
                                df['4.3.3(65) <= 4.3.1.a(61) + 4.3.1.b(62) + 4.2(60)'],
                                df['4.4.1(66) <= 4.1.1.a(56) + 4.1.1.b(57)'],
                                df['4.4.2(67) <= 4.4.1(66)'],
                                    df['4.4.3(68) <= 4.1.1.a(56) + 4.1.1.b(57)'],
                                    df['6.1(70) <= 2.1.1.a(47) + 2.1.1.b(48)'],
                                    df['6.3(72) <= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)'],
                                    df['6.4(73) <= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)'],
                                    df['7.2.1(76) <= 7.1.1(74)'],
                                    df['7.2.2(77) <= 7.1.2(75)'],
                                    df['8.2.3(81) <= 2.2(51)'],
                                    df['8.4(84) <= 2.1.1.a(47) + 2.1.1.b(48) + 2.2(51)'],
                                    df['8.7(87) <= 8.3(83) + 8.4(84) + 8.5(85)'],
                                    df['8.17.1(97) <= 8.1.1(78)'],
                                    df['8.17.2(98) <= 8.2.1(79) + 8.2.2(80) + 8.2.3(81) + 8.2.4(82)'],
                                        df['9.1.9(111) <= 4.1.1.a(56) + 4.1.1.b(57)'],
                                        df['9.1.13(115) <= 4.1.1.a(56) + 4.1.1.b(57)'],
                                        df['9.2.4.a(127) + 9.2.4.b(128) <= 9.2.1(124) + 9.2.2(125)'],
                                        df['11.2.2(175) <= 11.2.1(174)'],
                                        df['12.1.2.a(180) <= 12.1.1.a(178)'],
                                        df['12.1.2.b(181) <= 12.1.1.b(179)'],
                                        df['12.1.3.a(182) <= 12.1.1.a(178)'],
                                        df['12.1.3.b(183) <= 12.1.1.b(179)'],
                                        df['14.2.1(194) + 14.2.2(195) >= 14.1.1(186) + 14.1.2(187) + 14.1.3(188) + 14.1.4(189) + 14.1.5(190) + 14.1.6(191) + 14.1.7(192) + 14.1.8(193)'],
                                        df['14.3.3(200) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'],
                                        df['14.5.2(210) <= 14.5.1(209)'],
                                        df['15.3.3.b(230) <= 15.3.3.a(229)'],
                                        df['15.3.3.c(231) <= 15.3.3.b(230))'],
                                            df['15.4.2(237) <= 15.4.1(236)'],
                                            df['9.6.1(142) <= 9.1.1(103) + 9.1.2(104) + 9.1.3(105) + 9.1.4(106) + 9.1.5(107) + 9.1.6(108) +9.1.7(109) + 9.1.8(110) + 9.1.13(115) + 9.1.14(116) + 9.1.15(117) + 9.1.16(118) + 9.1.17(119) + 9.1.18(120) + 9.1.19(121) + 9.1.20(122) + 9.1.21(123) + 9.2.1(124) + 9.2.2(125) + 9.2.3(126) + 9.3.1(129) + 9.3.2(130) + 9.3.3(131) + 9.4.1(132) + 9.4.2(133) + 9.4.3(134) + 9.4.5(136) + 9.4.6(137) + 9.5.1(138) + 9.5.2(139) + 9.5.3(140) + 9.5.4(141)'],
                                            df['9.6.2(143) <= 9.1.1(103) + 9.1.2(104) + 9.1.3(105) + 9.1.4(106) + 9.1.5(107) + 9.1.6(108) +9.1.7(109) + 9.1.8(110) + 9.1.13(115) + 9.1.14(116) + 9.1.15(117) + 9.1.16(118) + 9.1.17(119) + 9.1.18(120) + 9.1.19(121) + 9.1.20(122) + 9.1.21(123) + 9.2.1(124) + 9.2.2(125) + 9.2.3(126) + 9.3.1(129) + 9.3.2(130) + 9.3.3(131) + 9.4.1(132) + 9.4.2(133) + 9.4.3(134) + 9.4.5(136) + 9.4.6(137) + 9.5.1(138) + 9.5.2(139) + 9.5.3(140) + 9.5.4(141)'],
                                            df['9.6.3(144) <= 9.1.1(103) + 9.1.2(104) + 9.1.3(105) + 9.1.4(106) + 9.1.5(107) + 9.1.6(108) +9.1.7(109) + 9.1.8(110) + 9.1.13(115) + 9.1.14(116) + 9.1.15(117) + 9.1.16(118) + 9.1.17(119) + 9.1.18(120) + 9.1.19(121) + 9.1.20(122) + 9.1.21(123) + 9.2.1(124) + 9.2.2(125) + 9.2.3(126) + 9.3.1(129) + 9.3.2(130) + 9.3.3(131) + 9.4.1(132) + 9.4.2(133) + 9.4.3(134) + 9.4.5(136) + 9.4.6(137) + 9.5.1(138) + 9.5.2(139) + 9.5.3(140) + 9.5.4(141)'],
                                            df['1.3.1.a(33) <= 1.3.1(32)'],
                                            df['9.7.2(146) <= 9.7.1(145)'],
                                            df['9.7.3(147) <= 9.7.2(146)'],
                                            df['11.1.1.b(169) <= 11.1.1.a(168)'],
                                            df['11.1.1.c(170) <= 11.1.1.a(168)'],
                                            df['11.1.2.b(171) <= 11.1.1.a(168)'],
                                            df['11.1.2.c(173) <= 11.1.2.a(171)'],
                                            df['14.4.1(201) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'],
                                                df['14.4.2(202) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199))'],
                                                df['14.4.3(203) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'],
                                                df['14.4.4(204) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'],
                                                df['14.4.5(205) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'],
                                                df['14.4.6(206) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'],
                                                df['14.4.7(207) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'],
                                                df['14.4.8(208) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'],
                                                df['14.6.1(214) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'],
                                                df['14.6.2(215) <= 14.3.1.a(196) + 14.3.1.b(197) + 14.3.2.a(198) + 14.3.2.b(199)'],
                                                df['14.9.2(219) <= 14.9.1(218)'],
                                                    df['15.2.2(224) <= 15.2.1(223)'],
                                                    df['15.3.1.b(226) <= 15.3.1.a(225)'],
                                                    df['15.3.2.b(228) <= 15.3.2.a(227))'],
                                                    df['15.3.4.b(233) <= 15.3.4.a(232)'],
                                                    df['15.3.4.d(235) <= 15.3.4.c(234)'],
                                                    df['9.1.2(104) <= 4.1.1.a(56) + 4.1.1.b(57)']], axis=1)

        # Mergining current result of modified checks with original dataframe and displaying it on screen
        frames = [df_, df]
        print(frames)
        df = pd.concat(frames, axis=1, sort=False)
        #df = df.dropna(axis=0, subset=['col_2'])
        self.tableView.setModel(PandasModel(df))

        msg = QMessageBox()
        msg.setWindowTitle("Validation Completion Message")
        msg.setText("Public Health Center Validation Complete")
        msg.setIcon(QMessageBox.Information)
        msg.exec()

        return df


    # Sub District Hospital Validation Rules Function
    # ===============================================
    def SDH_Validate(self):
        global df

        df = self.loadFile(df_)

        filterString = self.comboBox.currentText()
        
        df = df_.loc[df_['col_12'] == filterString]
        print(df)

        print('Entered SDH_Validate')

        # Modified Checks of SDH

        #1.1.1(23) <= 1.1(22)
        def res1(df):
            if pd.isnull(df['col_23']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_23']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_23']):
                    return 'Probable Reporting Error(1.1.1 is blank)'
                elif pd.isnull(float(df['col_22'])):
                    return 'Inconsistent'
            elif float(df['col_23']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #15.3.1.b(239) <= 15.3.1.a(238)
        def res2(df):
            if pd.isnull(df['col_239']) and pd.isnull(df['col_238']):
                return 'Blank'
            elif pd.isnull(df['col_239']) or pd.isnull(df['col_238']):
                if pd.isnull(df['col_239']):
                    return 'Probable Reporting Error(15.3.1.b is blank)'
                elif pd.isnull(float(df['col_238'])):
                    return 'Inconsistent'
            elif float(df['col_239']) > float(df['col_238']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.2.4(27) <= 1.1(22)
        def res3(df):
            if pd.isnull(df['col_27']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_27']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_27']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_27'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df
        
        #1.2.5(28) <= 1.1(22)
        def res4(df):
            if pd.isnull(df['col_28']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_28']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_28']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
            # If value exists for all the elements
            else:

                lhs_value = float(df['col_28'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df   
        
        #1.2.7(30) <= 1.1(22)
        def res5(df):
            if pd.isnull(df['col_30']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_30']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_30']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
            # If value exists for all the elements
            else:

                lhs_value = float(df['col_30'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df    

        #1.3.1.a(33) <= 1.3.1(32)
        def res6(df):
            if pd.isnull(df['col_33']) and pd.isnull(df['col_32']):
                return 'Blank'
            elif pd.isnull(df['col_33']) or pd.isnull(df['col_32']):
                if pd.isnull(df['col_33']):
                    return 'Probable Reporting Error(1.3.1.a is blank)'
                elif pd.isnull(float(df['col_32'])):
                    return 'Inconsistent'
            elif float(df['col_33']) > float(df['col_32']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.3.2(34) <= 2.1(47)
        def res7(df):
            if pd.isnull(df['col_34']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_34']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_34']):
                    return 'Probable Reporting Error(1.3.2 is blank)'
                elif pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
            elif float(df['col_34']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.4.4(38) <= 1.4.3(37)
        def res8(df):
            if pd.isnull(df['col_38']) and pd.isnull(df['col_37']):
                return 'Blank'
            elif pd.isnull(df['col_38']) or pd.isnull(df['col_37']):
                if pd.isnull(df['col_38']) and not pd.isnull(float(df['col_37'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
            # If value exists for all the elements
            else:

                lhs_value = float(df['col_38'])
                rhs_value = float(df['col_37'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df  
        
        #1.5.1(39) <= 1.1(22)
        def res9(df):
            if pd.isnull(df['col_39']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_39']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_39']):
                    return 'Probable Reporting Error(1.5.1 is blank)'
                elif pd.isnull(float(df['col_22'])):
                    return 'Inconsistent'
            elif float(df['col_39']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.5.2(40) <= 1.5.1(39)
        def res10(df):
            if pd.isnull(df['col_40']) and pd.isnull(df['col_39']):
                return 'Blank'
            elif pd.isnull(df['col_40']) or pd.isnull(df['col_39']):
                if pd.isnull(df['col_40']):
                    return 'Probable Reporting Error(1.5.2 is blank)'
                elif pd.isnull(float(df['col_39'])):
                    return 'Inconsistent'
            elif float(df['col_40']) > float(df['col_39']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        
        #1.5.3(41) <= 1.5.2(40)
        def res11(df):
            if pd.isnull(df['col_41']) and pd.isnull(df['col_40']):
                return 'Blank'
            elif pd.isnull(df['col_41']) or pd.isnull(df['col_40']):
                if pd.isnull(df['col_41']):
                    return 'Probable Reporting Error(1.5.3 is blank)'
                elif pd.isnull(float(df['col_40'])):
                    return 'Inconsistent'
            elif float(df['col_41']) > float(df['col_40']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        
        #1.6.1.a(42) <= 1.1(22)
        def res12(df):
            if pd.isnull(df['col_42']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_42']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_42']):
                    return 'Probable Reporting Error(1.6.1 is blank)'
                elif pd.isnull(float(df['col_22'])):
                    return 'Inconsistent'
            elif float(df['col_42']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.6.1.b(43) <= 1.6.1.a(42)
        def res13(df):
            if pd.isnull(df['col_43']) and pd.isnull(df['col_42']):
                return 'Blank'
            elif pd.isnull(df['col_43']) or pd.isnull(df['col_42']):
                if pd.isnull(df['col_43']):
                    return 'Probable Reporting Error(1.6.1.b is blank)'
                elif pd.isnull(float(df['col_42'])):
                    return 'Inconsistent'
            elif float(df['col_43']) > float(df['col_42']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.6.1.c(44) <= 1.6.1.b(43)
        def res14(df):
            if pd.isnull(df['col_44']) and pd.isnull(df['col_43']):
                return 'Blank'
            elif pd.isnull(df['col_44']) or pd.isnull(df['col_43']):
                if pd.isnull(df['col_44']):
                    return 'Probable Reporting Error(1.6.1.c is blank)'
                elif pd.isnull(float(df['col_43'])):
                    return 'Inconsistent'
            elif float(df['col_44']) > float(df['col_43']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.6.1.e(46) <= 1.6.1.d(45)
        def res15(df):
            if pd.isnull(df['col_46']) and pd.isnull(df['col_45']):
                return 'Blank'
            elif pd.isnull(df['col_46']) or pd.isnull(df['col_45']):
                if pd.isnull(df['col_46']):
                    return 'Probable Reporting Error(1.6.1.e is blank)'
                elif pd.isnull(float(df['col_45'])):
                    return 'Inconsistent'
            elif float(df['col_46']) > float(df['col_45']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #2.1.1(48) <= 2.1(47)
        def res16(df):
            if pd.isnull(df['col_48']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_48']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_48']):
                    return 'Probable Reporting Error(2.1.1 is blank)'
                elif pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
            elif float(df['col_48']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #3.1(50) <= 2.1(47)
        def res17(df):
            if pd.isnull(df['col_50']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_50']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_50']):
                    return 'Probable Reporting Error(3.1 is blank)'
                elif pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
            elif float(df['col_50']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #3.1.1(51) <= 3.1(50)
        def res18(df):
            if pd.isnull(df['col_51']) and pd.isnull(df['col_50']):
                return 'Blank'
            elif pd.isnull(df['col_51']) or pd.isnull(df['col_50']):
                if pd.isnull(df['col_51']):
                    return 'Probable Reporting Error(3.1.1 is blank)'
                elif pd.isnull(float(df['col_50'])):
                    return 'Inconsistent'
            elif float(df['col_51']) > float(df['col_50']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.1.1.a(52) + 4.1.1.b(53) + 4.1.3(55)<= 2.1(47)
        def res19(df):
            if pd.isnull(df['col_52']) and pd.isnull(df['col_53']) and pd.isnull(df['col_55']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_52']) or pd.isnull(df['col_53']) or pd.isnull(df['col_55']) or pd.isnull(df['col_47']):
                if pd.isnull((float(df['col_52'])) + (float(df['col_53'])) + (float(df['col_55']))) and not pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
                elif not pd.isnull((float(df['col_52'])) + (float(df['col_53'])) + (float(df['col_55']))) and pd.isnull(float(df['col_47'])):
                    return 'Probable Reporting Error'
            elif float(df['col_47']) < float(df['col_52']) + float(df['col_53']) + float(df['col_55']) :
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.1.2(54) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res20(df):
            if pd.isnull(df['col_54']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_54']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_54']):
                    return 'Probable Reporting Error(4.1.2 is blank)'
                elif pd.isnull((float(df['col_52']) + float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_54']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.3.2.a(59) <= 4.3.1.a(57) + 4.3.1.b(58) + 4.2(56)
        def res21(df):
            if pd.isnull(df['col_59']) and pd.isnull(df['col_57']) and pd.isnull(df['col_58']) and pd.isnull(df['col_56']):
                return 'Blank'
            elif pd.isnull(df['col_59']) or pd.isnull(df['col_57']) or pd.isnull(df['col_58']) or pd.isnull(df['col_56']):
                if pd.isnull(df['col_59']):
                    return 'Probable Reporting Error(4.3.2.a is blank)'
                elif pd.isnull((float(df['col_57']) + float(df['col_58']) + float(df['col_56']))):
                    return 'Inconsistent'
            elif float(df['col_59']) > float(df['col_57']) + float(df['col_58']) + float(df['col_56']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.3.2.b(60) <= 4.3.2.a(59)
        def res22(df):
            if pd.isnull(df['col_60']) and pd.isnull(df['col_59']):
                return 'Blank'
            elif pd.isnull(df['col_60']) or pd.isnull(df['col_59']):
                if pd.isnull(df['col_60']):
                    return 'Probable Reporting Error(4.3.2.b is blank)'
                elif pd.isnull((float(df['col_59']))):
                    return 'Inconsistent'
            elif float(df['col_60']) > float(df['col_59']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.3.3(61) <= 4.3.1.a(57) + 4.3.1.b(58) + 4.2(56)
        def res23(df):
            if pd.isnull(df['col_61']) and pd.isnull(df['col_57']) and pd.isnull(df['col_58']) and pd.isnull(df['col_56']):
                return 'Blank'
            elif pd.isnull(df['col_61']) or pd.isnull(df['col_57']) or pd.isnull(df['col_58']) or pd.isnull(df['col_56']):
                if pd.isnull(df['col_61']):
                    return 'Probable Reporting Error(4.3.3 is blank)'
                elif pd.isnull((float(df['col_57'])) + (float(df['col_58'])) + (float(df['col_56']))):
                    return 'Inconsistent'
            elif float(df['col_61']) > float(df['col_57']) + float(df['col_58']) + float(df['col_56']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.4.1(62) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res24(df):
            if pd.isnull(df['col_62']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_62']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_62']):
                    return 'Probable Reporting Error(4.4.1 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_62']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.4.2(63) <= 4.4.1(62)
        def res25(df):
            if pd.isnull(df['col_63']) and pd.isnull(df['col_62']):
                return 'Blank'
            elif pd.isnull(df['col_63']) or pd.isnull(df['col_62']):
                if pd.isnull(df['col_63']):
                    return 'Probable Reporting Error(4.4.2 is blank)'
                elif pd.isnull((float(df['col_62']))):
                    return 'Inconsistent'
            elif float(df['col_63']) > float(df['col_62']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.4.3(64) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res26(df):
            if pd.isnull(df['col_64']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_64']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_64']):
                    return 'Probable Reporting Error(4.4.3 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_64']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #5.2(66) <= 5.1(65)
        def res27(df):
            if pd.isnull(df['col_66']) and pd.isnull(df['col_65']):
                return 'Blank'
            elif pd.isnull(df['col_66']) or pd.isnull(df['col_65']):
                if pd.isnull(df['col_66']):
                    return 'Probable Reporting Error(5.2 is blank)'
                elif pd.isnull((float(df['col_65']))):
                    return 'Inconsistent'
            elif float(df['col_66']) > float(df['col_65']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #6.3(69) <= 2.1(47)
        def res28(df):
            if pd.isnull(df['col_69']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_69']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_69']) and not pd.isnull(float(df['col_47'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
            # If value exists for all the elements
            else:

                lhs_value = float(df['col_69'])
                rhs_value = float(df['col_47'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df 
        
        #6.4(70) <= 2.1(47)
        def res29(df):
            if pd.isnull(df['col_70']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_70']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_70']) and not pd.isnull(float(df['col_47'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
            # If value exists for all the elements
            else:

                lhs_value = float(df['col_70'])
                rhs_value = float(df['col_47'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df 
        
        #7.2.1(73) <= 7.1.1(71)
        def res30(df):
            if pd.isnull(df['col_73']) and pd.isnull(df['col_71']):
                return 'Blank'
            elif pd.isnull(df['col_73']) or pd.isnull(df['col_71']):
                if pd.isnull(df['col_73']):
                    return 'Probable Reporting Error(7.2.1 is blank)'
                elif pd.isnull((float(df['col_71']))):
                    return 'Inconsistent'
            elif float(df['col_73']) > float(df['col_71']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #7.2.2(74) <= 7.1.2(72)
        def res31(df):
            if pd.isnull(df['col_74']) and pd.isnull(df['col_72']):
                return 'Blank'
            elif pd.isnull(df['col_74']) or pd.isnull(df['col_72']):
                if pd.isnull(df['col_74']):
                    return 'Probable Reporting Error(7.2.2 is blank)'
                elif pd.isnull((float(df['col_72']))):
                    return 'Inconsistent'
            elif float(df['col_74']) > float(df['col_72']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.2.3(78) <= 2.1(47)
        def res32(df):
            if pd.isnull(df['col_78']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_78']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_78']):
                    return 'Probable Reporting Error(8.2.3 is blank)'
                elif pd.isnull((float(df['col_47']))):
                    return 'Inconsistent'
            elif float(df['col_78']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.4(81) <= 2.1(47)
        def res33(df):
            if pd.isnull(df['col_81']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_81']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_81']):
                    return 'Probable Reporting Error(8.4 is blank)'
                elif pd.isnull((float(df['col_47']))):
                    return 'Inconsistent'
            elif float(df['col_81']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.7(84) <= 8.3(80) + 8.4(81) + 8.5(82)
        def res34(df):
            if pd.isnull(df['col_84']) and pd.isnull(df['col_80']) and pd.isnull(df['col_81']) and pd.isnull(df['col_82']):
                return 'Blank'
            elif pd.isnull(df['col_84']) or pd.isnull(df['col_80']) or pd.isnull(df['col_81']) or pd.isnull(df['col_82']):
                if pd.isnull(df['col_84']):
                    return 'Probable Reporting Error(8.7 is blank)'
                elif pd.isnull((float(df['col_80'])) + (float(df['col_81'])) + (float(df['col_82']))):
                    return 'Inconsistent'
            elif float(df['col_84']) > float(df['col_80']) + float(df['col_81']) + float(df['col_82']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.17.1(94) <= 8.1.1(75)
        def res35(df):
            if pd.isnull(df['col_94']) and pd.isnull(df['col_75']):
                return 'Blank'
            elif pd.isnull(df['col_94']) or pd.isnull(df['col_75']):
                if pd.isnull(df['col_94']):
                    return 'Probable Reporting Error(8.17.1 is blank)'
                elif pd.isnull((float(df['col_75']))):
                    return 'Inconsistent'
            elif float(df['col_94']) > float(df['col_75']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.17.2(95) <= 8.2.1(76) + 8.2.2(77) + 8.2.3(78) + 8.2.4(79)
        def res36(df):
            if pd.isnull(df['col_95']) and pd.isnull(df['col_76']) and pd.isnull(df['col_77']) and pd.isnull(df['col_78']) and pd.isnull(df['col_79']):
                return 'Blank'
            elif pd.isnull(df['col_95']) or pd.isnull(df['col_76']) or pd.isnull(df['col_77']) or pd.isnull(df['col_78']) or pd.isnull(df['col_79']):
                if pd.isnull(df['col_95']):
                    return 'Probable Reporting Error(8.17.2 is blank)'
                elif pd.isnull((float(df['col_76'])) + (float(df['col_77'])) + (float(df['col_78'])) + (float(df['col_79']))):
                    return 'Inconsistent'
            elif float(df['col_95']) > float(df['col_76']) + float(df['col_77']) + float(df['col_78']) + float(df['col_79']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.1(100) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res37(df):
            if pd.isnull(df['col_100']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_100']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_100']):
                    return 'Probable Reporting Error(9.1.1 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_100']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.2(101) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res38(df):
            if pd.isnull(df['col_101']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_101']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_101']):
                    return 'Probable Reporting Error(9.1.2 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_101']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.9(108) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res39(df):
            if pd.isnull(df['col_108']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_108']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_108']):
                    return 'Probable Reporting Error(9.1.9 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_108']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.13(112) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res40(df):
            if pd.isnull(df['col_112']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_112']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_112']):
                    return 'Probable Reporting Error(9.1.13 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_112']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.2.4.a(124) + 9.2.4.b(125) <= 9.2.1(121) + 9.2.2(122)
        def res41(df):
            if pd.isnull(df['col_124']) and pd.isnull(df['col_125']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']):
                return 'Blank'
            elif pd.isnull(df['col_124']) or pd.isnull(df['col_125']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']):
                if pd.isnull((float(df['col_124'])) + (float(df['col_125']))) and not pd.isnull(float(df['col_121']) + float(df['col_122'])):
                    return 'Probable Reporting Error'
                elif not pd.isnull((float(df['col_124'])) + (float(df['col_125']))) and pd.isnull(float(df['col_121']) + float(df['col_122'])):
                    return 'Inconsistent'
            elif (float(df['col_124']) + float(df['col_125'])) > (float(df['col_121']) + float(df['col_122'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #11.2.2(171) <= 11.2.1(170)
        def res42(df):
            if pd.isnull(df['col_171']) and pd.isnull(df['col_170']):
                return 'Blank'
            elif pd.isnull(df['col_171']) or pd.isnull(df['col_170']):
                if pd.isnull(df['col_171']):
                    return 'Probable Reporting Error(11.2.2 is blank)'
                elif pd.isnull((float(df['col_170']))):
                    return 'Inconsistent'
            elif float(df['col_171']) > float(df['col_170']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #11.4.2(176) <= 11.4.1(175)
        def res43(df):
            if pd.isnull(df['col_176']) and pd.isnull(df['col_175']):
                return 'Blank'
            elif pd.isnull(df['col_176']) or pd.isnull(df['col_175']):
                if pd.isnull(df['col_176']):
                    return 'Probable Reporting Error(11.4.2 is blank)'
                elif pd.isnull((float(df['col_175']))):
                    return 'Inconsistent'
            elif float(df['col_176']) > float(df['col_175']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #12.1.2.a(179) <= 12.1.1.a(177)
        def res44(df):
            if pd.isnull(df['col_179']) and pd.isnull(df['col_177']):
                return 'Blank'
            elif pd.isnull(df['col_179']) or pd.isnull(df['col_177']):
                if pd.isnull(df['col_179']):
                    return 'Probable Reporting Error(12.1.2.a is blank)'
                elif pd.isnull((float(df['col_177']))):
                    return 'Inconsistent'
            elif float(df['col_179']) > float(df['col_177']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #12.1.2.b(180) <= 12.1.1.b(178)
        def res45(df):
            if pd.isnull(df['col_180']) and pd.isnull(df['col_178']):
                return 'Blank'
            elif pd.isnull(df['col_180']) or pd.isnull(df['col_178']):
                if pd.isnull(df['col_180']):
                    return 'Probable Reporting Error(12.1.2.b is blank)'
                elif pd.isnull((float(df['col_178']))):
                    return 'Inconsistent'
            elif float(df['col_180']) > float(df['col_178']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #12.1.3.a(181) <= 12.1.1.a(177)
        def res46(df):
            if pd.isnull(df['col_181']) and pd.isnull(df['col_177']):
                return 'Blank'
            elif pd.isnull(df['col_181']) or pd.isnull(df['col_177']):
                if pd.isnull(df['col_181']):
                    return 'Probable Reporting Error(12.1.3.a is blank)'
                elif pd.isnull((float(df['col_177']))):
                    return 'Inconsistent'
            elif float(df['col_181']) > float(df['col_177']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #12.1.3.b(182) <= 12.1.1.b(178)
        def res47(df):
            if pd.isnull(df['col_182']) and pd.isnull(df['col_178']):
                return 'Blank'
            elif pd.isnull(df['col_182']) or pd.isnull(df['col_178']):
                if pd.isnull(df['col_182']):
                    return 'Probable Reporting Error(12.1.3.b is blank)'
                elif pd.isnull((float(df['col_178']))):
                    return 'Inconsistent'
            elif float(df['col_182']) > float(df['col_178']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.2.1(194) +14.2.2(195) >= 14.1.1(185) +14.1.2(186) +14.1.3(187) +14.1.4(188) +14.1.5(189) +14.1.6(190) +14.1.7(191) +14.1.8(192) +14.1.9(193)
        def res48(df):
            if pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_185']) and pd.isnull(df['col_186']) and pd.isnull(df['col_187']) and pd.isnull(df['col_188']) and pd.isnull(df['col_189']) and pd.isnull(df['col_190']) and pd.isnull(df['col_191']) and pd.isnull(df['col_192']) and pd.isnull(df['col_193']):
                return 'Blank'
            elif pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_185']) or pd.isnull(df['col_186']) or pd.isnull(df['col_187']) or pd.isnull(df['col_188']) or pd.isnull(df['col_189']) or pd.isnull(df['col_190']) or pd.isnull(df['col_191']) or pd.isnull(df['col_192']) or pd.isnull(df['col_193']):
                if pd.isnull((float(df['col_194'])) + (float(df['col_195']))):
                    return 'Inconsistent'
                elif pd.isnull(float(df['col_185']) + float(df['col_186']) + float(df['col_187']) + float(df['col_188']) + float(df['col_189']) + float(df['col_190'])+ float(df['col_191']) + float(df['col_192']) + float(df['col_193'])):
                    return 'Probable Reporting Error'
            elif (float(df['col_194']) + float(df['col_195'])) < (float(df['col_185']) + float(df['col_186']) + float(df['col_187'])
                                                                                                + float(df['col_188']) + float(df['col_189']) + float(df['col_190'])
                                                                                                + float(df['col_191']) + float(df['col_192']) + float(df['col_193'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.3.3(200) <= 14.3.1.a(196) +14.3.1.b(197) +14.3.2.a(198) +14.3.2.b(199)
        def res49(df):
            if pd.isnull(df['col_200']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_200']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_200']):
                    return 'Probable Reporting Error(14.3.3 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198'])) + (float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_200']) > float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.4.1(201) <=14.3.1.a(196) +14.3.1.b(197) +14.3.2.a(198) +14.3.2.b(199)
        def res50(df):
            if pd.isnull(df['col_201']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_201']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_201']):
                    return 'Probable Reporting Error(14.4.1 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198'])) + (float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_201']) > float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.4.2 (202)<=14.3.1.a (196)+14.3.1.b (197)+14.3.2.a (198)+14.3.2.b (199)
        def res51(df):
            if pd.isnull(df['col_202']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_202']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_202']):
                    return 'Probable Reporting Error(14.4.2 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198'])) + (float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_202']) > float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df


        #14.4.3 (203)<=14.3.1.a(196) +14.3.1.b(197) +14.3.2.a(198) +14.3.2.b(199)
        def res52(df):
            if pd.isnull(df['col_203']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_203']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_203']):
                    return 'Probable Reporting Error(14.4.3 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198'])) + (float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_203']) > float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        
        
        #14.4.4 (204)<=14.3.1.a (196)+14.3.1.b (197)+14.3.2.a (198)+14.3.2.b (199)
        def res53(df):
            if pd.isnull(df['col_204']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_204']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_204']):
                    return 'Probable Reporting Error(14.4.4 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198'])) + (float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_204']) > float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df



        #14.4.5 (205)<=14.3.1.a (196)+14.3.1.b (197)+14.3.2.a (198)+14.3.2.b (199)
        def res54(df):
            if pd.isnull(df['col_205']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_205']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_205']):
                    return 'Probable Reporting Error(14.4.5 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198'])) + (float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_205']) > float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        
        
        #14.4.6 (206)<=14.3.1.a (196)+14.3.1.b (197)+14.3.2.a (198)+14.3.2.b (199)
        def res55(df):
            if pd.isnull(df['col_206']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_206']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_206']):
                    return 'Probable Reporting Error(14.4.6 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198'])) + (float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_206']) > float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df        
        


        #14.4.7 (207)<=14.3.1.a (196)+14.3.1.b (197)+14.3.2.a (198)+14.3.2.b (199)
        def res56(df):
            if pd.isnull(df['col_207']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_207']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_207']):
                    return 'Probable Reporting Error(14.4.7 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198'])) + (float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_207']) > float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df        
        

        #14.4.8 (208)<=14.3.1.a (196)+14.3.1.b (197)+14.3.2.a (198)+14.3.2.b (199)
        def res57(df):
            if pd.isnull(df['col_208']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_208']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_208']):
                    return 'Probable Reporting Error(14.4.8 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197'])) + (float(df['col_198'])) + (float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_208']) > float(df['col_196']) + float(df['col_197']) + float(df['col_198']) + float(df['col_199']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df        
        
        

        #14.6.1(210) <= 14.5(209)
        def res58(df):
            if pd.isnull(df['col_210']) and pd.isnull(df['col_209']):
                return 'Blank'
            elif pd.isnull(df['col_210']) or pd.isnull(df['col_209']):
                if pd.isnull(df['col_210']):
                    return 'Probable Reporting Error(14.6.1 is blank)'
                elif pd.isnull(float(df['col_209'])):
                    return 'Inconsistent'
            elif float(df['col_210']) > float(df['col_209']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df        
        
        

        #14.6.2(211) <= 14.5(209)
        def res59(df):
            if pd.isnull(df['col_211']) and pd.isnull(df['col_209']):
                return 'Blank'
            elif pd.isnull(df['col_211']) or pd.isnull(df['col_209']):
                if pd.isnull(df['col_211']):
                    return 'Probable Reporting Error(14.6.2 is blank)'
                elif pd.isnull(float(df['col_209'])):
                    return 'Inconsistent'
            elif float(df['col_211']) > float(df['col_209']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df          



        #14.6.3(212) <= 14.5(209)
        def res60(df):
            if pd.isnull(df['col_212']) and pd.isnull(df['col_209']):
                return 'Blank'
            elif pd.isnull(df['col_212']) or pd.isnull(df['col_209']):
                if pd.isnull(df['col_212']):
                    return 'Probable Reporting Error(14.6.3 is blank)'
                elif pd.isnull(float(df['col_209'])):
                    return 'Inconsistent'
            elif float(df['col_212']) > float(df['col_209']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df         
        


        #14.6.4(213) <= 14.5(209)
        def res61(df):
            if pd.isnull(df['col_213']) and pd.isnull(df['col_209']):
                return 'Blank'
            elif pd.isnull(df['col_213']) or pd.isnull(df['col_209']):
                if pd.isnull(df['col_213']):
                    return 'Probable Reporting Error(14.6.4 is blank)'
                elif pd.isnull(float(df['col_209'])):
                    return 'Inconsistent'
            elif float(df['col_213']) > float(df['col_209']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df              
        


        #14.6.5(214) <= 14.5(209)
        def res62(df):
            if pd.isnull(df['col_214']) and pd.isnull(df['col_209']):
                return 'Blank'
            elif pd.isnull(df['col_214']) or pd.isnull(df['col_209']):
                if pd.isnull(df['col_214']):
                    return 'Probable Reporting Error(14.6.5 is blank)'
                elif pd.isnull(float(df['col_209'])):
                    return 'Inconsistent'
            elif float(df['col_214']) > float(df['col_209']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df    



        #14.6.6(215) <= 14.5(209)
        def res63(df):
            if pd.isnull(df['col_215']) and pd.isnull(df['col_209']):
                return 'Blank'
            elif pd.isnull(df['col_215']) or pd.isnull(df['col_209']):
                if pd.isnull(df['col_215']):
                    return 'Probable Reporting Error(14.6.6 is blank)'
                elif pd.isnull(float(df['col_209'])):
                    return 'Inconsistent'
            elif float(df['col_215']) > float(df['col_209']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df    



        #14.7(216) <= 14.5(209)
        def res64(df):
            if pd.isnull(df['col_216']) and pd.isnull(df['col_209']):
                return 'Blank'
            elif pd.isnull(df['col_216']) or pd.isnull(df['col_209']):
                if pd.isnull(df['col_216']):
                    return 'Probable Reporting Error(14.7 is blank)'
                elif pd.isnull(float(df['col_209'])):
                    return 'Inconsistent'
            elif float(df['col_216']) > float(df['col_209']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df



        #14.14.2(233) <= 14.14.1(232)
        def res65(df):
            if pd.isnull(df['col_233']) and pd.isnull(df['col_232']):
                return 'Blank'
            elif pd.isnull(df['col_233']) or pd.isnull(df['col_232']):
                if pd.isnull(df['col_233']):
                    return 'Probable Reporting Error(14.14.2 is blank)'
                elif pd.isnull(float(df['col_232'])):
                    return 'Inconsistent'
            elif float(df['col_233']) > float(df['col_232']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df


        #15.2.2(237) <= 15.2.1(236)
        def res66(df):
            if pd.isnull(df['col_237']) and pd.isnull(df['col_236']):
                return 'Blank'
            elif pd.isnull(df['col_237']) or pd.isnull(df['col_236']):
                if pd.isnull(df['col_237']):
                    return 'Probable Reporting Error(15.2.2 is blank)'
                elif pd.isnull(float(df['col_236'])):
                    return 'Inconsistent'
            elif float(df['col_237']) > float(df['col_236']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df



        #15.3.2b(241) <= 15.3.2a(240)
        def res67(df):
            if pd.isnull(df['col_241']) and pd.isnull(df['col_240']):
                return 'Blank'
            elif pd.isnull(df['col_241']) or pd.isnull(df['col_240']):
                if pd.isnull(df['col_241']):
                    return 'Probable Reporting Error(15.3.2b is blank)'
                elif pd.isnull(float(df['col_240'])):
                    return 'Inconsistent'
            elif float(df['col_241']) > float(df['col_240']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df



        #15.3.3b(243) <= 15.3.3a(242)
        def res68(df):
            if pd.isnull(df['col_243']) and pd.isnull(df['col_242']):
                return 'Blank'
            elif pd.isnull(df['col_243']) or pd.isnull(df['col_242']):
                if pd.isnull(df['col_243']):
                    return 'Probable Reporting Error(15.3.3b is blank)'
                elif pd.isnull(float(df['col_242'])):
                    return 'Inconsistent'
            elif float(df['col_243']) > float(df['col_242']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df



        #15.3.3c(244) <= 15.3.3b(243)
        def res69(df):
            if pd.isnull(df['col_244']) and pd.isnull(df['col_243']):
                return 'Blank'
            elif pd.isnull(df['col_244']) or pd.isnull(df['col_243']):
                if pd.isnull(df['col_244']):
                    return 'Probable Reporting Error(15.3.3c is blank)'
                elif pd.isnull(float(df['col_243'])):
                    return 'Inconsistent'
            elif float(df['col_244']) > float(df['col_243']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df



        #15.4.2(250) <= 15.4.1(249)
        def res70(df):
            if pd.isnull(df['col_250']) and pd.isnull(df['col_249']):
                return 'Blank'
            elif pd.isnull(df['col_250']) or pd.isnull(df['col_249']):
                if pd.isnull(df['col_250']):
                    return 'Probable Reporting Error(15.4.2 is blank)'
                elif pd.isnull(float(df['col_249'])):
                    return 'Inconsistent'
            elif float(df['col_250']) > float(df['col_249']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #9.6.1 (139) <=9.1.1 (100)+9.1.2 (101)+9.1.3 (102)+9.1.4 (103)+9.1.5 (104)+9.1.6 (105)+9.1.7 (106)+9.1.8 (107)+9.1.13 (112)+9.1.14 (113)+9.1.15 (114)+9.1.16 (115)+9.1.17 (116)+9.1.18 (117)+9.1.19 (118)+9.1.20 (119)+9.1.21 (120)+9.2.1 (121)+9.2.2 (122)+9.2.3 (123)+9.3.1 (126)+9.3.2 (127)+9.3.3 (128)+9.4.1 (129)+9.4.2 (130)+9.4.3 (131)+9.4.5 (133)+9.4.6 (134)+9.5.1 (135)+9.5.2 (136)+9.5.3 (137)+9.5.4 (138)
        def res71(df):
            if pd.isnull(df['col_139']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_112']) and pd.isnull(df['col_113']) and pd.isnull(df['col_114']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_123']) and pd.isnull(df['col_126']) and pd.isnull(df['col_127']) and pd.isnull(df['col_128']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_131']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_135']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']) and pd.isnull(df['col_138']):
                return 'Blank'
            elif pd.isnull(df['col_139']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_112']) or pd.isnull(df['col_113']) or pd.isnull(df['col_114']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_123']) or pd.isnull(df['col_126']) or pd.isnull(df['col_127']) or pd.isnull(df['col_128']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_131']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_135']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']) or pd.isnull(df['col_138']):
                if pd.isnull(df['col_139']):
                    return 'Probable Reporting Error(9.6.1 is blank)'
                elif pd.isnull((float(df['col_100'])) + (float(df['col_101'])) + (float(df['col_102'])) + (float(df['col_103'])) + (float(df['col_104'])) + (float(df['col_105'])) + (float(df['col_106'])) + (float(df['col_107'])) + (float(df['col_112'])) + (float(df['col_113'])) + (float(df['col_114'])) + (float(df['col_115'])) + (float(df['col_116'])) + (float(df['col_117'])) + (float(df['col_118'])) + (float(df['col_119'])) + (float(df['col_120'])) + (float(df['col_121'])) + (float(df['col_122'])) + (float(df['col_123'])) + (float(df['col_126'])) + (float(df['col_127'])) + (float(df['col_128'])) + (float(df['col_129'])) + (float(df['col_130'])) + (float(df['col_131'])) + (float(df['col_133'])) + (float(df['col_134'])) + (float(df['col_135'])) + (float(df['col_136'])) + (float(df['col_137'])) + (float(df['col_138']))):
                    return 'Inconsistent'
            elif float(df['col_139']) > float(df['col_100']) > + (float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_112']) + float(df['col_113']) + float(df['col_114']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_126']) + float(df['col_127']) + float(df['col_128']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #9.6.2 (140) <=9.1.1 (100)+9.1.2 (101)+9.1.3 (102)+9.1.4 (103)+9.1.5 (104)+9.1.6 (105)+9.1.7 (106)+9.1.8 (107)+9.1.13 (112)+9.1.14 (113)+9.1.15 (114)+9.1.16 (115)+9.1.17 (116)+9.1.18 (117)+9.1.19 (118)+9.1.20 (119)+9.1.21 (120)+9.2.1 (121)+9.2.2 (122)+9.2.3 (123)+9.3.1 (126)+9.3.2 (127)+9.3.3 (128)+9.4.1 (129)+9.4.2 (130)+9.4.3 (131)+9.4.5 (133)+9.4.6 (134)+9.5.1 (135)+9.5.2 (136)+9.5.3 (137)+9.5.4 (138)
        def res72(df):
            if pd.isnull(df['col_140']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_112']) and pd.isnull(df['col_113']) and pd.isnull(df['col_114']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_123']) and pd.isnull(df['col_126']) and pd.isnull(df['col_127']) and pd.isnull(df['col_128']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_131']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_135']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']) and pd.isnull(df['col_138']):
                return 'Blank'
            elif pd.isnull(df['col_140']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_112']) or pd.isnull(df['col_113']) or pd.isnull(df['col_114']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_123']) or pd.isnull(df['col_126']) or pd.isnull(df['col_127']) or pd.isnull(df['col_128']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_131']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_135']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']) or pd.isnull(df['col_138']):
                if pd.isnull(df['col_140']):
                    return 'Probable Reporting Error(9.6.2 is blank)'
                elif pd.isnull((float(df['col_100'])) + (float(df['col_101'])) + (float(df['col_102'])) + (float(df['col_103'])) + (float(df['col_104'])) + (float(df['col_105'])) + (float(df['col_106'])) + (float(df['col_107'])) + (float(df['col_112'])) + (float(df['col_113'])) + (float(df['col_114'])) + (float(df['col_115'])) + (float(df['col_116'])) + (float(df['col_117'])) + (float(df['col_118'])) + (float(df['col_119'])) + (float(df['col_120'])) + (float(df['col_121'])) + (float(df['col_122'])) + (float(df['col_123'])) + (float(df['col_126'])) + (float(df['col_127'])) + (float(df['col_128'])) + (float(df['col_129'])) + (float(df['col_130'])) + (float(df['col_131'])) + (float(df['col_133'])) + (float(df['col_134'])) + (float(df['col_135'])) + (float(df['col_136'])) + (float(df['col_137'])) + (float(df['col_138']))):
                    return 'Inconsistent'
            elif float(df['col_140']) > float(df['col_100']) > + (float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_112']) + float(df['col_113']) + float(df['col_114']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_126']) + float(df['col_127']) + float(df['col_128']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #9.6.3 (141) <=9.1.1 (100)+9.1.2 (101)+9.1.3 (102)+9.1.4 (103)+9.1.5 (104)+9.1.6 (105)+9.1.7 (106)+9.1.8 (107)+9.1.13 (112)+9.1.14 (113)+9.1.15 (114)+9.1.16 (115)+9.1.17 (116)+9.1.18 (117)+9.1.19 (118)+9.1.20 (119)+9.1.21 (120)+9.2.1 (121)+9.2.2 (122)+9.2.3 (123)+9.3.1 (126)+9.3.2 (127)+9.3.3 (128)+9.4.1 (129)+9.4.2 (130)+9.4.3 (131)+9.4.5 (133)+9.4.6 (134)+9.5.1 (135)+9.5.2 (136)+9.5.3 (137)+9.5.4 (138)
        def res73(df):
            if pd.isnull(df['col_141']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_112']) and pd.isnull(df['col_113']) and pd.isnull(df['col_114']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_123']) and pd.isnull(df['col_126']) and pd.isnull(df['col_127']) and pd.isnull(df['col_128']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_131']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_135']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']) and pd.isnull(df['col_138']):
                return 'Blank'
            elif pd.isnull(df['col_141']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_112']) or pd.isnull(df['col_113']) or pd.isnull(df['col_114']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_123']) or pd.isnull(df['col_126']) or pd.isnull(df['col_127']) or pd.isnull(df['col_128']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_131']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_135']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']) or pd.isnull(df['col_138']):
                if pd.isnull(df['col_141']):
                    return 'Probable Reporting Error(9.6.3 is blank)'
                elif pd.isnull((float(df['col_100'])) + (float(df['col_101'])) + (float(df['col_102'])) + (float(df['col_103'])) + (float(df['col_104'])) + (float(df['col_105'])) + (float(df['col_106'])) + (float(df['col_107'])) + (float(df['col_112'])) + (float(df['col_113'])) + (float(df['col_114'])) + (float(df['col_115'])) + (float(df['col_116'])) + (float(df['col_117'])) + (float(df['col_118'])) + (float(df['col_119'])) + (float(df['col_120'])) + (float(df['col_121'])) + (float(df['col_122'])) + (float(df['col_123'])) + (float(df['col_126'])) + (float(df['col_127'])) + (float(df['col_128'])) + (float(df['col_129'])) + (float(df['col_130'])) + (float(df['col_131'])) + (float(df['col_133'])) + (float(df['col_134'])) + (float(df['col_135'])) + (float(df['col_136'])) + (float(df['col_137'])) + (float(df['col_138']))):
                    return 'Inconsistent'
            elif float(df['col_141']) > float(df['col_100']) > + (float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_112']) + float(df['col_113']) + float(df['col_114']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_126']) + float(df['col_127']) + float(df['col_128']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df 

        #9.7.2(143) <= 9.7.1(142)
        def res74(df):
            if pd.isnull(df['col_143']) and pd.isnull(df['col_142']):
                return 'Blank'
            elif pd.isnull(df['col_143']) or pd.isnull(df['col_142']):
                if pd.isnull(df['col_143']):
                    return 'Probable Reporting Error(9.7.2 is blank)'
                elif pd.isnull(float(df['col_142'])):
                    return 'Inconsistent'
            elif float(df['col_143']) > float(df['col_142']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #9.7.3(144) <= 9.7.2(143)
        def res75(df):
            if pd.isnull(df['col_144']) and pd.isnull(df['col_143']):
                return 'Blank'
            elif pd.isnull(df['col_144']) or pd.isnull(df['col_143']):
                if pd.isnull(df['col_144']):
                    return 'Probable Reporting Error(9.7.3 is blank)'
                elif pd.isnull(float(df['col_143'])):
                    return 'Inconsistent'
            elif float(df['col_144']) > float(df['col_143']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.1b (165) <= 11.1.1a (164)
        def res76(df):
            if pd.isnull(df['col_165']) and pd.isnull(df['col_164']):
                return 'Blank'
            elif pd.isnull(df['col_165']) or pd.isnull(df['col_164']):
                if pd.isnull(df['col_165']):
                    return 'Probable Reporting Error(11.1.1b is blank)'
                elif pd.isnull(float(df['col_164'])):
                    return 'Inconsistent'
            elif float(df['col_165']) > float(df['col_164']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.1c (166) <= 11.1.1a (164)
        def res77(df):
            if pd.isnull(df['col_166']) and pd.isnull(df['col_164']):
                return 'Blank'
            elif pd.isnull(df['col_166']) or pd.isnull(df['col_164']):
                if pd.isnull(df['col_166']):
                    return 'Probable Reporting Error(11.1.1c is blank)'
                elif pd.isnull(float(df['col_164'])):
                    return 'Inconsistent'
            elif float(df['col_166']) > float(df['col_164']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.2b (168) <= 11.1.2a (167)
        def res78(df):
            if pd.isnull(df['col_168']) and pd.isnull(df['col_167']):
                return 'Blank'
            elif pd.isnull(df['col_168']) or pd.isnull(df['col_167']):
                if pd.isnull(df['col_168']):
                    return 'Probable Reporting Error(11.1.2b is blank)'
                elif pd.isnull(float(df['col_167'])):
                    return 'Inconsistent'
            elif float(df['col_168']) > float(df['col_167']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

    #11.1.2c (169) <= 11.1.2a (167)
        def res79(df):
            if pd.isnull(df['col_169']) and pd.isnull(df['col_167']):
                return 'Blank'
            elif pd.isnull(df['col_169']) or pd.isnull(df['col_167']):
                if pd.isnull(df['col_169']):
                    return 'Probable Reporting Error(11.1.2c is blank)'
                elif pd.isnull(float(df['col_167'])):
                    return 'Inconsistent'
            elif float(df['col_169']) > float(df['col_167']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #14.9.1 (222)<=14.3.1.a (196)+14.3.1.b (197)
        def res80(df):
            if pd.isnull(df['col_222']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_222']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_222']):
                    return 'Probable Reporting Error(14.9.1 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_222']) > float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df      

        #14.9.2 (223)<=14.3.2.a (198)+14.3.2.b (199)
        def res81(df):
            if pd.isnull(df['col_223']) and pd.isnull(df['col_198']) and pd.isnull(df['col_199']):
                return 'Blank'
            elif pd.isnull(df['col_223']) or pd.isnull(df['col_198']) or pd.isnull(df['col_199']):
                if pd.isnull(df['col_223']):
                    return 'Probable Reporting Error(14.9.2 is blank)'
                elif pd.isnull((float(df['col_198'])) + (float(df['col_199']))):
                    return 'Inconsistent'
            elif float(df['col_223']) > float(df['col_198']) + float(df['col_199']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df     

        #14.13 (231)<=14.12.1 (226)+14.12.2 (227)+14.12.3 (228)+14.12.4 (229)+14.12.5 (230)
        def res82(df):
            if pd.isnull(df['col_231']) and pd.isnull(df['col_226']) and pd.isnull(df['col_227']) and pd.isnull(df['col_228']) and pd.isnull(df['col_229']) and pd.isnull(df['col_230']):
                return 'Blank'
            elif pd.isnull(df['col_231']) or pd.isnull(df['col_226']) or pd.isnull(df['col_227']) or pd.isnull(df['col_228']) or pd.isnull(df['col_229']) or pd.isnull(df['col_230']):
                if pd.isnull(df['col_231']):
                    return 'Probable Reporting Error(14.13 is blank)'
                elif pd.isnull((float(df['col_226'])) + (float(df['col_227'])) + (float(df['col_228'])) + (float(df['col_229'])) + (float(df['col_230']))):
                    return 'Inconsistent'
            elif float(df['col_231']) > float(df['col_226']) + float(df['col_227']) + float(df['col_228']) + float(df['col_229']) + float(df['col_230']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df      

        #14.8.2(218) <= 14.8.1(217)
        def res83(df):
            if pd.isnull(df['col_218']) or pd.isnull(df['col_217']):
                return 'Blank'
            elif pd.isnull(df['col_218']) or pd.isnull(df['col_217']):
                if pd.isnull(df['col_218']):
                    return 'Probable Reporting Error(14.8.2 is blank)'
                elif pd.isnull(float(df['col_217'])):
                    return 'Inconsistent'
            elif float(df['col_218']) > float(df['col_217']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #15.3.4b(246) <= 15.3.4a(245)
        def res84(df):
            if pd.isnull(df['col_246']) or pd.isnull(df['col_245']):
                return 'Blank'
            elif pd.isnull(df['col_246']) or pd.isnull(df['col_245']):
                if pd.isnull(df['col_246']):
                    return 'Probable Reporting Error(15.3.4b is blank)'
                elif pd.isnull(float(df['col_245'])):
                    return 'Inconsistent'
            elif float(df['col_246']) > float(df['col_245']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df       
        
        #15.3.4d(248) <= 15.3.4c(247)
        def res85(df):
            if pd.isnull(df['col_248']) or pd.isnull(df['col_247']):
                return 'Blank'
            elif pd.isnull(df['col_248']) or pd.isnull(df['col_247']):
                if pd.isnull(df['col_248']):
                    return 'Probable Reporting Error(15.3.4d is blank)'
                elif pd.isnull(float(df['col_247'])):
                    return 'Inconsistent'
            elif float(df['col_248']) > float(df['col_247']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df       

        # Renaming column names=================SDH

        df['1.1.1<=1.1'] = df.apply(res1, axis=1)
        df['15.3.1.b<=15.3.1.a'] = df.apply(res2, axis=1)
        df['1.2.4<=1.1'] = df.apply(res3, axis=1)
        df['1.2.5<=1.1'] = df.apply(res4, axis=1)
        df['1.2.7<=1.1'] = df.apply(res5, axis=1)
        df['1.3.1.a<=1.3.1'] = df.apply(res6, axis=1)
        df['1.3.2<=2.1'] = df.apply(res7, axis=1)
        df['1.4.4>=1.4.3'] = df.apply(res8, axis=1)
        df['1.5.1<=1.1'] = df.apply(res9, axis=1)
        df['1.5.2<=1.5.1'] = df.apply(res10, axis=1)
        df['1.5.3<=1.5.2'] = df.apply(res11, axis=1)
        df['1.6.1.a<=1.1'] = df.apply(res12, axis=1)
        df['1.6.1.b<=1.6.1.a'] = df.apply(res13, axis=1)
        df['1.6.1.c<=1.6.1.b'] = df.apply(res14, axis=1)
        df['1.6.1.e<=1.6.1.d'] = df.apply(res15, axis=1)
        df['2.1.1<=2.1'] = df.apply(res16, axis=1)
        df['3.1<=2.1'] = df.apply(res17, axis=1)
        df['3.1.1<=3.1'] = df.apply(res18, axis=1)
        df['4.1.1.a+4.1.1.b+4.1.3>=2.1'] = df.apply(res19, axis=1)
        df['4.1.2<=4.1.1.a+4.1.1.b'] = df.apply(res20, axis=1)
        df['4.3.2.a<=4.3.1.a+4.3.1.b+4.2'] = df.apply(res21, axis=1)
        df['4.3.2.b<=4.3.2.a'] = df.apply(res22, axis=1)
        df['4.3.3<=4.3.1.a+4.3.1.b+4.2'] = df.apply(res23, axis=1)
        df['4.4.1<=4.1.1.a+4.1.1.b'] = df.apply(res24, axis=1)
        df['4.4.2<=4.4.1'] = df.apply(res25, axis=1)
        df['4.4.3<=4.1.1.a+4.1.1.b'] = df.apply(res26, axis=1)
        df['5.2<=5.1'] = df.apply(res27, axis=1)
        df['6.3<=2.1'] = df.apply(res28, axis=1)
        df['6.4<=2.1'] = df.apply(res29, axis=1)
        df['7.2.1<=7.1.1'] = df.apply(res30, axis=1)
        df['7.2.2<=7.1.2'] = df.apply(res31, axis=1)
        df['8.2.3<=2.1'] = df.apply(res32, axis=1)
        df['8.4<=2.1'] = df.apply(res33, axis=1)
        df['8.7<=8.3+8.4+8.5'] = df.apply(res34, axis=1)
        df['8.17.1<=8.1.1'] = df.apply(res35, axis=1)
        df['8.17.2<=8.2.1+8.2.2+8.2.3+8.2.4'] = df.apply(res36, axis=1)
        df['9.1.1<=4.1.1.a+4.1.1.b'] = df.apply(res37, axis=1)
        df['9.1.2<=4.1.1.a+4.1.1.b'] = df.apply(res38, axis=1)
        df['9.1.9<=4.1.1.a+4.1.1.b'] = df.apply(res39, axis=1)
        df['9.1.13<=4.1.1.a+4.1.1.b'] = df.apply(res40, axis=1)
        df['9.2.4.a+9.2.4.b<=9.2.1+ 9.2.2'] = df.apply(res41, axis=1)
        df['11.2.2<=11.2.1'] = df.apply(res42, axis=1)
        df['11.4.2<=11.4.1'] = df.apply(res43, axis=1)
        df['12.1.2.a<=12.1.1.a'] = df.apply(res44, axis=1)
        df['12.1.2.b<=12.1.1.b'] = df.apply(res45, axis=1)
        df['12.1.3.a<=12.1.1.a'] = df.apply(res46, axis=1)
        df['12.1.3.b<=12.1.1.b'] = df.apply(res47, axis=1)
        df['14.2.1+14.2.2>=14.1.1+14.1.2+14.1.3+14.1.4+14.1.5+14.1.6+14.1.7+14.1.8+14.1.9'] = df.apply(res48, axis=1)
        df['14.3.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res49, axis=1)
        df['14.4.1<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res50, axis=1)
        df['14.4.2<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res51, axis=1)
        df['14.4.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res52, axis=1)
        df['14.4.4<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res53, axis=1)
        df['14.4.5<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res54, axis=1)
        df['14.4.6<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res55, axis=1)
        df['14.4.7<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res56, axis=1)
        df['14.4.8<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res57, axis=1)
        df['14.6.1<=14.5'] = df.apply(res58, axis=1)
        df['14.6.2<=14.5'] = df.apply(res59, axis=1)
        df['14.6.3<=14.5'] = df.apply(res60, axis=1)
        df['14.6.4<=14.5'] = df.apply(res61, axis=1)
        df['14.6.5<=14.5'] = df.apply(res62, axis=1)
        df['14.6.6<=14.5'] = df.apply(res63, axis=1)
        df['14.7<=14.5'] = df.apply(res64, axis=1)
        df['14.14.2<=14.14.1'] = df.apply(res65, axis=1)
        df['15.2.2<=15.2.1'] = df.apply(res66, axis=1)
        df['15.3.2.b<=15.3.2.a'] = df.apply(res67, axis=1)
        df['15.3.3.b<=15.3.3.a'] = df.apply(res68, axis=1)
        df['15.3.3.c<=15.3.3.b'] = df.apply(res69, axis=1)
        df['15.4.2<=15.4.1'] = df.apply(res70, axis=1)
        df['9.6.1<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'] = df.apply(res71, axis=1)
        df['9.6.2<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'] = df.apply(res72, axis=1)
        df['9.6.3<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'] = df.apply(res73, axis=1)
        df['9.7.2<=9.7.1'] = df.apply(res74, axis=1)
        df['9.7.3<=9.7.2'] = df.apply(res75, axis=1)
        df['11.1.1.b<=11.1.1.a'] = df.apply(res76, axis=1)
        df['11.1.1.c<=11.1.1.a'] = df.apply(res77, axis=1)
        df['11.1.2.b<=11.1.2.a'] = df.apply(res78, axis=1)
        df['11.1.2.c<=11.1.2.a'] = df.apply(res79, axis=1)
        df['14.9.1<=14.3.1.a+14.3.1.b'] = df.apply(res80, axis=1)
        df['14.9.2<=14.3.2.a+14.3.2.b'] = df.apply(res81, axis=1)
        df['14.13<=14.12.1+14.12.2+14.12.3+14.12.4+14.12.5'] = df.apply(res82, axis=1)
        df['14.8.2<=14.8.1'] = df.apply(res83, axis=1)
        df['15.3.4.b<=15.3.4.a'] = df.apply(res84, axis=1)
        df['15.3.4.d<=15.3.4.c'] = df.apply(res85, axis=1)


        # Merging all the renamed columns
        # ===============================
        df = pd.concat([df ['1.1.1<=1.1'],
                        df ['15.3.1.b<=15.3.1.a'],
                        df ['1.2.4<=1.1'],
                        df ['1.2.5<=1.1'],
                        df ['1.2.7<=1.1'],
                        df ['1.3.1.a<=1.3.1'],
                        df ['1.3.2<=2.1'],
                        df ['1.4.4>=1.4.3'],
                        df ['1.5.1<=1.1'],
                            df ['1.5.2<=1.5.1'],
                            df ['1.5.3<=1.5.2'],
                            df ['1.6.1.a<=1.1'],
                            df ['1.6.1.b<=1.6.1.a'],
                            df ['1.6.1.c<=1.6.1.b'],
                            df ['1.6.1.e<=1.6.1.d'],
                            df ['2.1.1<=2.1'],
                            df ['3.1<=2.1'],
                            df ['3.1.1<=3.1'],
                            df ['4.1.1.a+4.1.1.b+4.1.3>=2.1'],
                            df ['4.1.2<=4.1.1.a+4.1.1.b'],
                                df ['4.3.2.a<=4.3.1.a+4.3.1.b+4.2'],
                                df ['4.3.2.b<=4.3.2.a'],
                                df ['4.3.3<=4.3.1.a+4.3.1.b+4.2'],
                                df ['4.4.1<=4.1.1.a+4.1.1.b'],
                                df ['4.4.2<=4.4.1'],
                                df ['4.4.3<=4.1.1.a+4.1.1.b'],
                                df ['5.2<=5.1'],
                                df ['6.3<=2.1'],
                                df ['6.4<=2.1'],
                                df ['7.2.1<=7.1.1'],
                                df ['7.2.2<=7.1.2'],
                                df ['8.2.3<=2.1'],
                                    df ['8.4<=2.1'],
                                    df ['8.7<=8.3+8.4+8.5'],
                                    df ['8.17.1<=8.1.1'],
                                    df ['8.17.2<=8.2.1+8.2.2+8.2.3+8.2.4'],
                                    df ['9.1.1<=4.1.1.a+4.1.1.b'],
                                    df ['9.1.2<=4.1.1.a+4.1.1.b'],
                                    df ['9.1.9<=4.1.1.a+4.1.1.b'],
                                    df ['9.1.13<=4.1.1.a+4.1.1.b'],
                                    df ['9.2.4.a+9.2.4.b<=9.2.1+ 9.2.2'],
                                    df ['11.2.2<=11.2.1'],
                                    df ['11.4.2<=11.4.1'],
                                    df ['12.1.2.a<=12.1.1.a'],
                                        df ['12.1.2.b<=12.1.1.b'],
                                        df ['12.1.3.a<=12.1.1.a'],
                                        df ['12.1.3.b<=12.1.1.b'],
                                        df ['14.2.1+14.2.2>=14.1.1+14.1.2+14.1.3+14.1.4+14.1.5+14.1.6+14.1.7+14.1.8+14.1.9'],
                                        df ['14.3.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                        df ['14.4.1<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                        df ['14.4.2<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                        df ['14.4.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                        df ['14.4.4<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                        df ['14.4.5<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                        df ['14.4.6<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.4.7<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.4.8<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.6.1<=14.5'],
                                            df ['14.6.2<=14.5'],
                                            df ['14.6.3<=14.5'],
                                            df ['14.6.4<=14.5'],
                                            df ['14.6.5<=14.5'],
                                            df ['14.6.6<=14.5'],
                                            df ['14.7<=14.5'],
                                            df ['14.14.2<=14.14.1'],
                                            df ['15.2.2<=15.2.1'],
                                                df ['15.3.2.b<=15.3.2.a'],
                                                df ['15.3.3.b<=15.3.3.a'],
                                                df ['15.3.3.c<=15.3.3.b'],
                                                df ['15.4.2<=15.4.1'],
                                                df ['9.6.1<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'],
                                                df ['9.6.2<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'],
                                                df ['9.6.3<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'],
                                                df ['9.7.2<=9.7.1'],
                                                df ['9.7.3<=9.7.2'],
                                                df ['11.1.1.b<=11.1.1.a'],
                                                df ['11.1.1.c<=11.1.1.a'],
                                                df ['11.1.2.b<=11.1.2.a'],
                                                    df ['11.1.2.c<=11.1.2.a'],
                                                    df ['14.9.1<=14.3.1.a+14.3.1.b'],
                                                    df ['14.9.2<=14.3.2.a+14.3.2.b'],
                                                    df ['14.13<=14.12.1+14.12.2+14.12.3+14.12.4+14.12.5'],
                                                    df ['14.8.2<=14.8.1'],
                                                    df ['15.3.4.b<=15.3.4.a'],
                                                    df ['15.3.4.d<=15.3.4.c']], axis=1)

        # Mergining current result of modified checks with original dataframe and displaying it on screen
        frames = [df_, df]
        print(frames)
        df = pd.concat(frames, axis=1, sort=False)
        #df = df.dropna(axis=0, subset=['col_2'])
        self.tableView.setModel(PandasModel(df))

        msg = QMessageBox()
        msg.setWindowTitle("Validation Completion Message")
        msg.setText("Sub District Hospital Validation Complete")
        msg.setIcon(QMessageBox.Information)
        msg.exec()

        return df


    # Community Health Centers Validation Rules Function
    # ===============================================
    def CHC_Validate(self):
        global df

        df = self.loadFile(df_)

        filterString = self.comboBox.currentText()
        
        df = df_.loc[df_['col_12'] == filterString]
        print(df)

        print('Entered CHC_Validate')

        # Modified Checks of SDH

        #1.1.1(23) <= 1.1(22)
        def res1(df):
            if pd.isnull(df['col_23']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_23']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_23']):
                    return 'Probable Reporting Error(1.1.1 is blank)'
                elif pd.isnull(float(df['col_22'])):
                    return 'Inconsistent'
            elif float(df['col_23']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #15.3.1.b(237) <= 15.3.1.a(236)
        def res2(df):
            if pd.isnull(df['col_237']) and pd.isnull(df['col_236']):
                return 'Blank'
            elif pd.isnull(df['col_237']) or pd.isnull(df['col_236']):
                if pd.isnull(df['col_237']):
                    return 'Probable Reporting Error(15.3.1.b is blank)'
                elif pd.isnull(float(df['col_236'])):
                    return 'Inconsistent'
            elif float(df['col_237']) > float(df['col_236']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        # 1.2.4(27) <= 1.1(22)
        def res3(df):
            if pd.isnull(df['col_27']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_27']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_27']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_27'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df
        
        
        #1.2.5(28) <= 1.1(22)
        def res4(df):
            if pd.isnull(df['col_28']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_28']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_28']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
             # If value exists for all the elements
            else:

                lhs_value = float(df['col_28'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df 
        
        #1.2.7(30) <= 1.1(22)
        def res5(df):
            if pd.isnull(df['col_30']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_30']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_30']) and not pd.isnull(float(df['col_22'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
             # If value exists for all the elements
            else:

                lhs_value = float(df['col_30'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df 
        
        #1.3.1.a(33) <= 1.3.1(32)
        def res6(df):
            if pd.isnull(df['col_33']) and pd.isnull(df['col_32']):
                return 'Blank'
            elif pd.isnull(df['col_33']) or pd.isnull(df['col_32']):
                if pd.isnull(df['col_33']):
                    return 'Probable Reporting Error(1.3.1.a is blank)'
                elif pd.isnull(float(df['col_32'])):
                    return 'Inconsistent'
            elif float(df['col_33']) > float(df['col_32']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.3.2(34) <= 2.1(47)
        def res7(df):
            if pd.isnull(df['col_34']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_34']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_34']):
                    return 'Probable Reporting Error(1.3.2 is blank)'
                elif pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
            elif float(df['col_34']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.4.4(38) <= 1.4.3(37)
        def res8(df):
            if pd.isnull(df['col_38']) and pd.isnull(df['col_37']):
                return 'Blank'
            elif pd.isnull(df['col_38']) or pd.isnull(df['col_37']):
                if pd.isnull(df['col_38']) and not pd.isnull(float(df['col_37'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
             # If value exists for all the elements
            else:

                lhs_value = float(df['col_38'])
                rhs_value = float(df['col_37'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df
        
        #1.5.2(40) <= 1.5.1(39)
        def res9(df):
            if pd.isnull(df['col_40']) and pd.isnull(df['col_39']):
                return 'Blank'
            elif pd.isnull(df['col_40']) or pd.isnull(df['col_39']):
                if pd.isnull(df['col_40']):
                    return 'Probable Reporting Error(1.5.2 is blank)'
                elif pd.isnull(float(df['col_39'])):
                    return 'Inconsistent'
            elif float(df['col_40']) > float(df['col_39']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        
        #1.5.1(39) <= 1.1(22)
        def res10(df):
            if pd.isnull(df['col_39']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_39']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_39']):
                    return 'Probable Reporting Error(1.5.1 is blank)'
                elif pd.isnull(float(df['col_22'])):
                    return 'Inconsistent'
            elif float(df['col_39']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.5.3(41) <= 1.5.2(40)
        def res11(df):
            if pd.isnull(df['col_41']) and pd.isnull(df['col_40']):
                return 'Blank'
            elif pd.isnull(df['col_41']) or pd.isnull(df['col_40']):
                if pd.isnull(df['col_41']):
                    return 'Probable Reporting Error(1.5.3 is blank)'
                elif pd.isnull(float(df['col_40'])):
                    return 'Inconsistent'
            elif float(df['col_41']) > float(df['col_40']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.6.1.b(43) <= 1.6.1.a(42)
        def res12(df):
            if pd.isnull(df['col_43']) and pd.isnull(df['col_42']):
                return 'Blank'
            elif pd.isnull(df['col_43']) or pd.isnull(df['col_42']):
                if pd.isnull(df['col_43']):
                    return 'Probable Reporting Error(1.6.1.b is blank)'
                elif pd.isnull(float(df['col_42'])):
                    return 'Inconsistent'
            elif float(df['col_43']) > float(df['col_42']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.6.1.a(42) <= 1.1(22)
        def res13(df):
            if pd.isnull(df['col_42']) and pd.isnull(df['col_22']):
                return 'Blank'
            elif pd.isnull(df['col_42']) or pd.isnull(df['col_22']):
                if pd.isnull(df['col_42']):
                    return 'Probable Reporting Error(1.6.1 is blank)'
                elif pd.isnull(float(df['col_22'])):
                    return 'Inconsistent'
            elif float(df['col_42']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.6.1.c(44) <= 1.6.1.b(43)
        def res14(df):
            if pd.isnull(df['col_44']) and pd.isnull(df['col_43']):
                return 'Blank'
            elif pd.isnull(df['col_44']) or pd.isnull(df['col_43']):
                if pd.isnull(df['col_44']):
                    return 'Probable Reporting Error(1.6.1.c is blank)'
                elif pd.isnull(float(df['col_43'])):
                    return 'Inconsistent'
            elif float(df['col_44']) > float(df['col_43']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #1.6.1.e(46) <= 1.6.1.d(45)
        def res15(df):
            if pd.isnull(df['col_46']) and pd.isnull(df['col_45']):
                return 'Blank'
            elif pd.isnull(df['col_46']) or pd.isnull(df['col_45']):
                if pd.isnull(df['col_46']):
                    return 'Probable Reporting Error(1.6.1.e is blank)'
                elif pd.isnull(float(df['col_45'])):
                    return 'Inconsistent'
            elif float(df['col_46']) > float(df['col_45']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #2.1.1(48) <= 2.1(47)
        def res16(df):
            if pd.isnull(df['col_48']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_48']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_48']):
                    return 'Probable Reporting Error(2.1.1 is blank)'
                elif pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
            elif float(df['col_48']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #3.1(50) <= 2.1(47)
        def res17(df):
            if pd.isnull(df['col_50']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_50']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_50']):
                    return 'Probable Reporting Error(3.1 is blank)'
                elif pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
            elif float(df['col_50']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #3.1.1(51) <= 3.1(50)
        def res18(df):
            if pd.isnull(df['col_51']) and pd.isnull(df['col_50']):
                return 'Blank'
            elif pd.isnull(df['col_51']) or pd.isnull(df['col_50']):
                if pd.isnull(df['col_51']):
                    return 'Probable Reporting Error(3.1.1 is blank)'
                elif pd.isnull(float(df['col_50'])):
                    return 'Inconsistent'
            elif float(df['col_51']) > float(df['col_50']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.1.1.a(52) + 4.1.1.b(53) + 4.1.3(55)<= 2.1(47)
        def res19(df):
            if pd.isnull(df['col_52']) and pd.isnull(df['col_53']) and pd.isnull(df['col_55']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_52']) or pd.isnull(df['col_53']) or pd.isnull(df['col_55']) or pd.isnull(df['col_47']):
                if pd.isnull((float(df['col_52'])) + (float(df['col_53'])) + (float(df['col_55']))) and not pd.isnull(float(df['col_47'])):
                    return 'Inconsistent'
                elif not pd.isnull((float(df['col_52'])) + (float(df['col_53'])) + (float(df['col_55']))) and pd.isnull(float(df['col_47'])):
                    return 'Probable Reporting Error'
            elif float(df['col_47']) < float(df['col_52']) + float(df['col_53']) + float(df['col_55']) :
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.1.2(54) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res20(df):
            if pd.isnull(df['col_54']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_54']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_54']):
                    return 'Probable Reporting Error(4.1.2 is blank)'
                elif pd.isnull((float(df['col_52']) + float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_54']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.3.2.a(59) <= 4.3.1.a(57) + 4.3.1.b(58) + 4.2(56)
        def res21(df):
            if pd.isnull(df['col_59']) and pd.isnull(df['col_57']) and pd.isnull(df['col_58']) and pd.isnull(df['col_56']):
                return 'Blank'
            elif pd.isnull(df['col_59']) or pd.isnull(df['col_57']) or pd.isnull(df['col_58']) or pd.isnull(df['col_56']):
                if pd.isnull(df['col_59']):
                    return 'Probable Reporting Error(4.3.2.a is blank)'
                elif pd.isnull((float(df['col_57']) + float(df['col_58']) + float(df['col_56']))):
                    return 'Inconsistent'
            elif float(df['col_59']) > float(df['col_57']) + float(df['col_58']) + float(df['col_56']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.3.2.b(60) <= 4.3.2.a(59)
        def res22(df):
            if pd.isnull(df['col_60']) and pd.isnull(df['col_59']):
                return 'Blank'
            elif pd.isnull(df['col_60']) or pd.isnull(df['col_59']):
                if pd.isnull(df['col_60']):
                    return 'Probable Reporting Error(4.3.2.b is blank)'
                elif pd.isnull((float(df['col_59']))):
                    return 'Inconsistent'
            elif float(df['col_60']) > float(df['col_59']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.3.3(61) <= 4.3.1.a(57) + 4.3.1.b(58) + 4.2(56)
        def res23(df):
            if pd.isnull(df['col_61']) and pd.isnull(df['col_57']) and pd.isnull(df['col_58']) and pd.isnull(df['col_56']):
                return 'Blank'
            elif pd.isnull(df['col_61']) or pd.isnull(df['col_57']) or pd.isnull(df['col_58']) or pd.isnull(df['col_56']):
                if pd.isnull(df['col_61']):
                    return 'Probable Reporting Error(4.3.3 is blank)'
                elif pd.isnull((float(df['col_57'])) + (float(df['col_58'])) + (float(df['col_56']))):
                    return 'Inconsistent'
            elif float(df['col_61']) > float(df['col_57']) + float(df['col_58']) + float(df['col_56']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.4.1(62) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res24(df):
            if pd.isnull(df['col_62']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_62']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_62']):
                    return 'Probable Reporting Error(4.4.1 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_62']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.4.2(63) <= 4.4.1(62)
        def res25(df):
            if pd.isnull(df['col_63']) and pd.isnull(df['col_62']):
                return 'Blank'
            elif pd.isnull(df['col_63']) or pd.isnull(df['col_62']):
                if pd.isnull(df['col_63']):
                    return 'Probable Reporting Error(4.4.2 is blank)'
                elif pd.isnull((float(df['col_62']))):
                    return 'Inconsistent'
            elif float(df['col_63']) > float(df['col_62']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #4.4.3(64) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res26(df):
            if pd.isnull(df['col_64']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_64']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_64']):
                    return 'Probable Reporting Error(4.4.3 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_64']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #5.2(66) <= 5.1(65)
        def res27(df):
            if pd.isnull(df['col_66']) and pd.isnull(df['col_65']):
                return 'Blank'
            elif pd.isnull(df['col_66']) or pd.isnull(df['col_65']):
                if pd.isnull(df['col_66']):
                    return 'Probable Reporting Error(5.2 is blank)'
                elif pd.isnull((float(df['col_65']))):
                    return 'Inconsistent'
            elif float(df['col_66']) > float(df['col_65']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #6.3(69) <= 2.1(47)
        def res28(df):
            if pd.isnull(df['col_69']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_69']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_69']) and not pd.isnull(float(df['col_47'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
             # If value exists for all the elements
            else:

                lhs_value = float(df['col_69'])
                rhs_value = float(df['col_47'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df 
        
        #6.4(70) <= 2.1(47)
        def res29(df):
            if pd.isnull(df['col_70']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_70']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_70']) and not pd.isnull(float(df['col_47'])):
                    return 'Probable Reporting Error'
                else:
                    return 'Probable Reporting Error'
                
             # If value exists for all the elements
            else:

                lhs_value = float(df['col_70'])
                rhs_value = float(df['col_47'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df 
        
        #7.2.1(73) <= 7.1.1(71)
        def res30(df):
            if pd.isnull(df['col_73']) and pd.isnull(df['col_71']):
                return 'Blank'
            elif pd.isnull(df['col_73']) or pd.isnull(df['col_71']):
                if pd.isnull(df['col_73']):
                    return 'Probable Reporting Error(7.2.1 is blank)'
                elif pd.isnull((float(df['col_71']))):
                    return 'Inconsistent'
            elif float(df['col_73']) > float(df['col_71']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #7.2.2(74) <= 7.1.2(72)
        def res31(df):
            if pd.isnull(df['col_74']) and pd.isnull(df['col_72']):
                return 'Blank'
            elif pd.isnull(df['col_74']) or pd.isnull(df['col_72']):
                if pd.isnull(df['col_74']):
                    return 'Probable Reporting Error(7.2.2 is blank)'
                elif pd.isnull((float(df['col_72']))):
                    return 'Inconsistent'
            elif float(df['col_74']) > float(df['col_72']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.2.3(78) <= 2.1(47)
        def res32(df):
            if pd.isnull(df['col_78']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_78']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_78']):
                    return 'Probable Reporting Error(8.2.3 is blank)'
                elif pd.isnull((float(df['col_47']))):
                    return 'Inconsistent'
            elif float(df['col_78']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
         #8.4(81) <= 2.1(47)
        def res33(df):
            if pd.isnull(df['col_81']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_81']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_81']):
                    return 'Probable Reporting Error(8.4 is blank)'
                elif pd.isnull((float(df['col_47']))):
                    return 'Inconsistent'
            elif float(df['col_81']) > float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.7(84) <= 8.3(80) + 8.4(81) + 8.5(82)
        def res34(df):
            if pd.isnull(df['col_84']) and pd.isnull(df['col_80']) and pd.isnull(df['col_81']) and pd.isnull(df['col_82']):
                return 'Blank'
            elif pd.isnull(df['col_84']) or pd.isnull(df['col_80']) or pd.isnull(df['col_81']) or pd.isnull(df['col_82']):
                if pd.isnull(df['col_84']):
                    return 'Probable Reporting Error(8.7 is blank)'
                elif pd.isnull((float(df['col_80'])) + (float(df['col_81'])) + (float(df['col_82']))):
                    return 'Inconsistent'
            elif float(df['col_84']) > float(df['col_80']) + float(df['col_81']) + float(df['col_82']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.17.1(94) <= 8.1.1(75)
        def res35(df):
            if pd.isnull(df['col_94']) and pd.isnull(df['col_75']):
                return 'Blank'
            elif pd.isnull(df['col_94']) or pd.isnull(df['col_75']):
                if pd.isnull(df['col_94']):
                    return 'Probable Reporting Error(8.17.1 is blank)'
                elif pd.isnull((float(df['col_75']))):
                    return 'Inconsistent'
            elif float(df['col_94']) > float(df['col_75']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #8.17.2(95) <= 8.2.1(76) + 8.2.2(77) + 8.2.3(78) + 8.2.4(79)
        def res36(df):
            if pd.isnull(df['col_95']) and pd.isnull(df['col_76']) and pd.isnull(df['col_77']) and pd.isnull(df['col_78']) and pd.isnull(df['col_79']):
                return 'Blank'
            elif pd.isnull(df['col_95']) or pd.isnull(df['col_76']) or pd.isnull(df['col_77']) or pd.isnull(df['col_78']) or pd.isnull(df['col_79']):
                if pd.isnull(df['col_95']):
                    return 'Probable Reporting Error(8.17.2 is blank)'
                elif pd.isnull((float(df['col_76'])) + (float(df['col_77'])) + (float(df['col_78'])) + (float(df['col_79']))):
                    return 'Inconsistent'
            elif float(df['col_95']) > float(df['col_76']) + float(df['col_77']) + float(df['col_78']) + float(df['col_79']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.1(100) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res37(df):
            if pd.isnull(df['col_100']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_100']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_100']):
                    return 'Probable Reporting Error(9.1.1 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_100']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.2(101) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res38(df):
            if pd.isnull(df['col_101']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_101']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_101']):
                    return 'Probable Reporting Error(9.1.2 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_101']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.9(108) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res39(df):
            if pd.isnull(df['col_108']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_108']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_108']):
                    return 'Probable Reporting Error(9.1.9 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_108']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.1.13(112) <= 4.1.1.a(52) + 4.1.1.b(53)
        def res40(df):
            if pd.isnull(df['col_112']) and pd.isnull(df['col_52']) and pd.isnull(df['col_53']):
                return 'Blank'
            elif pd.isnull(df['col_112']) or pd.isnull(df['col_52']) or pd.isnull(df['col_53']):
                if pd.isnull(df['col_112']):
                    return 'Probable Reporting Error(9.1.13 is blank)'
                elif pd.isnull((float(df['col_52'])) + (float(df['col_53']))):
                    return 'Inconsistent'
            elif float(df['col_112']) > float(df['col_52']) + float(df['col_53']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #9.2.4.a(124) + 9.2.4.b(125) <= 9.2.1(121) + 9.2.2(122)
        def res41(df):
            if pd.isnull(df['col_124']) and pd.isnull(df['col_125']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']):
                return 'Blank'
            elif pd.isnull(df['col_124']) or pd.isnull(df['col_125']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']):
                if pd.isnull((float(df['col_124'])) + (float(df['col_125']))) and not pd.isnull(float(df['col_121']) + float(df['col_122'])):
                    return 'Probable Reporting Error'
                elif not pd.isnull((float(df['col_124'])) + (float(df['col_125']))) and pd.isnull(float(df['col_121']) + float(df['col_122'])):
                    return 'Inconsistent'
            elif (float(df['col_124']) + float(df['col_125'])) > (float(df['col_121']) + float(df['col_122'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #11.2.2(172) <= 11.2.1(171)
        def res42(df):
            if pd.isnull(df['col_172']) and pd.isnull(df['col_171']):
                return 'Blank'
            elif pd.isnull(df['col_172']) or pd.isnull(df['col_171']):
                if pd.isnull(df['col_172']):
                    return 'Probable Reporting Error(11.2.2 is blank)'
                elif pd.isnull((float(df['col_171']))):
                    return 'Inconsistent'
            elif float(df['col_172']) > float(df['col_171']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
         #12.1.2.a(177) <= 12.1.1.a(175)
        def res43(df):
            if pd.isnull(df['col_177']) and pd.isnull(df['col_175']):
                return 'Blank'
            elif pd.isnull(df['col_177']) or pd.isnull(df['col_175']):
                if pd.isnull(df['col_177']):
                    return 'Probable Reporting Error(12.1.2.a is blank)'
                elif pd.isnull((float(df['col_175']))):
                    return 'Inconsistent'
            elif float(df['col_177']) > float(df['col_175']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #12.1.2.b(178) <= 12.1.1.b(176)
        def res44(df):
            if pd.isnull(df['col_178']) and pd.isnull(df['col_176']):
                return 'Blank'
            elif pd.isnull(df['col_178']) or pd.isnull(df['col_176']):
                if pd.isnull(df['col_178']):
                    return 'Probable Reporting Error(12.1.2.b is blank)'
                elif pd.isnull((float(df['col_176']))):
                    return 'Inconsistent'
            elif float(df['col_178']) > float(df['col_176']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #12.1.3.a(179) <= 12.1.1.a(175)
        def res45(df):
            if pd.isnull(df['col_179']) and pd.isnull(df['col_175']):
                return 'Blank'
            elif pd.isnull(df['col_179']) or pd.isnull(df['col_175']):
                if pd.isnull(df['col_179']):
                    return 'Probable Reporting Error(12.1.3.a is blank)'
                elif pd.isnull((float(df['col_175']))):
                    return 'Inconsistent'
            elif float(df['col_179']) > float(df['col_175']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #12.1.3.b(180) <= 12.1.1.b(176)
        def res46(df):
            if pd.isnull(df['col_180']) and pd.isnull(df['col_176']):
                return 'Blank'
            elif pd.isnull(df['col_180']) or pd.isnull(df['col_176']):
                if pd.isnull(df['col_180']):
                    return 'Probable Reporting Error(12.1.3.b is blank)'
                elif pd.isnull((float(df['col_176']))):
                    return 'Inconsistent'
            elif float(df['col_180']) > float(df['col_176']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.2.1(192) +14.2.2(193) >= 14.1.1(183) +14.1.2(184) +14.1.3(185) +14.1.4(186) +14.1.5(187) +14.1.6(188) +14.1.7(189) +14.1.8(190) +14.1.9(191)
        def res47(df):
            if pd.isnull(df['col_192']) and pd.isnull(df['col_193']) and pd.isnull(df['col_183']) and pd.isnull(df['col_184']) and pd.isnull(df['col_185']) and pd.isnull(df['col_186']) and pd.isnull(df['col_187']) and pd.isnull(df['col_188']) and pd.isnull(df['col_189']) and pd.isnull(df['col_190']) and pd.isnull(df['col_191']):
                return 'Blank'
            elif pd.isnull(df['col_192']) or pd.isnull(df['col_193']) or pd.isnull(df['col_183']) or pd.isnull(df['col_184']) or pd.isnull(df['col_185']) or pd.isnull(df['col_186']) or pd.isnull(df['col_187']) or pd.isnull(df['col_188']) or pd.isnull(df['col_189']) or pd.isnull(df['col_190']) or pd.isnull(df['col_191']):
                if pd.isnull((float(df['col_192'])) + (float(df['col_193']))):
                    return 'Inconsistent'
                elif pd.isnull(float(df['col_183']) + float(df['col_184']) + float(df['col_185'])+ float(df['col_186']) + float(df['col_187']) + float(df['col_188'])+ float(df['col_189']) + float(df['col_190']) + float(df['col_191'])):
                    return 'Probable Reporting Error'
            elif (float(df['col_192']) + float(df['col_193'])) < (float(df['col_183']) + float(df['col_184']) + float(df['col_185'])
                                                                                                + float(df['col_186']) + float(df['col_187']) + float(df['col_188'])
                                                                                                + float(df['col_189']) + float(df['col_190']) + float(df['col_191'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.3.3(198) <= 14.3.1.a(194) +14.3.1.b(195) +14.3.2.a(196) +14.3.2.b(197)
        def res48(df):
            if pd.isnull(df['col_198']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_198']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_198']):
                    return 'Probable Reporting Error(14.3.3 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_198']) > float(df['col_194']) + float(df['col_195']) + float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.4.1(199) <=14.3.1.a(194) +14.3.1.b(195) +14.3.2.a(196) +14.3.2.b(197)
        def res49(df):
            if pd.isnull(df['col_199']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_199']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_199']):
                    return 'Probable Reporting Error(14.4.1 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_199']) > float(df['col_194']) + float(df['col_195']) + float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.4.2(200) <=14.3.1.a(194)+ 14.3.1.b(195) +14.3.2.a(196) +14.3.2.b(197)
        def res50(df):
            if pd.isnull(df['col_200']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_200']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_200']):
                    return 'Probable Reporting Error(14.4.2 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_200']) > float(df['col_194']) + float(df['col_195']) + float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.4.2 (200)<=14.3.1.a (194)+14.3.1.b (195)+14.3.2.a (196)+14.3.2.b (197)
        def res50(df):
            if pd.isnull(df['col_200']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_200']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_200']):
                    return 'Probable Reporting Error(14.4.2 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_200']) > float(df['col_194']) + float(df['col_195']) + float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #14.4.3 (201)<=14.3.1.a (194)+14.3.1.b (195)+14.3.2.a (196)+14.3.2.b (197)
        def res51(df):
            if pd.isnull(df['col_201']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_201']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_201']):
                    return 'Probable Reporting Error(14.4.3 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_201']) > float(df['col_194']) + float(df['col_195']) + float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.4.4 (202)<=14.3.1.a (194)+14.3.1.b (195)+14.3.2.a (196)+14.3.2.b (197)
        def res52(df):
            if pd.isnull(df['col_202']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_202']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_202']):
                    return 'Probable Reporting Error(14.4.4 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_202']) > float(df['col_194']) + float(df['col_195']) + float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #14.4.5 (203)<=14.3.1.a (194)+14.3.1.b (195)+14.3.2.a (196)+14.3.2.b (197)
        def res53(df):
            if pd.isnull(df['col_203']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_203']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_203']):
                    return 'Probable Reporting Error(14.4.5 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_203']) > float(df['col_194']) + float(df['col_195']) + float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
        
        #14.4.6 (204)<=14.3.1.a (194)+14.3.1.b (195)+14.3.2.a (196)+14.3.2.b (197)
        def res54(df):
            if pd.isnull(df['col_204']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_204']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_204']):
                    return 'Probable Reporting Error(14.4.6 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_204']) > float(df['col_194']) + float(df['col_195']) + float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df        
        
        #14.4.7 (205)<=14.3.1.a (194)+14.3.1.b (195)+14.3.2.a (196)+14.3.2.b (197)
        def res55(df):
            if pd.isnull(df['col_205']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_205']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_205']):
                    return 'Probable Reporting Error(14.4.7 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_205']) > float(df['col_194']) + float(df['col_195']) + float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df      
        
        #14.4.8 (206)<=14.3.1.a (194)+14.3.1.b (195)+14.3.2.a (196)+14.3.2.b (197)
        def res56(df):
            if pd.isnull(df['col_206']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_206']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_206']):
                    return 'Probable Reporting Error(14.4.8 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195'])) + (float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_206']) > float(df['col_194']) + float(df['col_195']) + float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df    

        #14.6.1(208) <= 14.5(207)
        def res57(df):
            if pd.isnull(df['col_208']) and pd.isnull(df['col_207']):
                return 'Blank'
            elif pd.isnull(df['col_208']) or pd.isnull(df['col_207']):
                if pd.isnull(df['col_208']):
                    return 'Probable Reporting Error(14.6.1 is blank)'
                elif pd.isnull(float(df['col_207'])):
                    return 'Inconsistent'
            elif float(df['col_208']) > float(df['col_207']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df        
        
        #14.6.2(209) <= 14.5(207)
        def res58(df):
            if pd.isnull(df['col_209']) and pd.isnull(df['col_207']):
                return 'Blank'
            elif pd.isnull(df['col_209']) or pd.isnull(df['col_207']):
                if pd.isnull(df['col_209']):
                    return 'Probable Reporting Error(14.6.2 is blank)'
                elif pd.isnull(float(df['col_207'])):
                    return 'Inconsistent'
            elif float(df['col_209']) > float(df['col_207']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #14.6.3(210) <= 14.5(207)
        def res59(df):
            if pd.isnull(df['col_210']) and pd.isnull(df['col_207']):
                return 'Blank'
            elif pd.isnull(df['col_210']) or pd.isnull(df['col_207']):
                if pd.isnull(df['col_210']):
                    return 'Probable Reporting Error(14.6.3 is blank)'
                elif pd.isnull(float(df['col_207'])):
                    return 'Inconsistent'
            elif float(df['col_210']) > float(df['col_207']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #14.6.4(211) <= 14.5(207)
        def res60(df):
            if pd.isnull(df['col_211']) and pd.isnull(df['col_207']):
                return 'Blank'
            elif pd.isnull(df['col_211']) or pd.isnull(df['col_207']):
                if pd.isnull(df['col_211']):
                    return 'Probable Reporting Error(14.6.4 is blank)'
                elif pd.isnull(float(df['col_207'])):
                    return 'Inconsistent'
            elif float(df['col_211']) > float(df['col_207']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df              
        
        #14.6.5(212) <= 14.5(207)
        def res61(df):
            if pd.isnull(df['col_212']) and pd.isnull(df['col_207']):
                return 'Blank'
            elif pd.isnull(df['col_212']) or pd.isnull(df['col_207']):
                if pd.isnull(df['col_212']):
                    return 'Probable Reporting Error(14.6.5 is blank)'
                elif pd.isnull(float(df['col_207'])):
                    return 'Inconsistent'
            elif float(df['col_212']) > float(df['col_207']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #14.7(214) <= 14.5(207)
        def res62(df):
            if pd.isnull(df['col_214']) and pd.isnull(df['col_207']):
                return 'Blank'
            elif pd.isnull(df['col_214']) or pd.isnull(df['col_207']):
                if pd.isnull(df['col_214']):
                    return 'Probable Reporting Error(14.7 is blank)'
                elif pd.isnull(float(df['col_207'])):
                    return 'Inconsistent'
            elif float(df['col_214']) > float(df['col_207']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #14.14.2(231) <= 14.14.1(230)
        def res63(df):
            if pd.isnull(df['col_231']) and pd.isnull(df['col_230']):
                return 'Blank'
            elif pd.isnull(df['col_231']) or pd.isnull(df['col_230']):
                if pd.isnull(df['col_231']):
                    return 'Probable Reporting Error(14.14.2 is blank)'
                elif pd.isnull(float(df['col_230'])):
                    return 'Inconsistent'
            elif float(df['col_231']) > float(df['col_230']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
    
        #15.2.2(235) <= 15.2.1(234)
        def res64(df):
            if pd.isnull(df['col_235']) and pd.isnull(df['col_234']):
                return 'Blank'
            elif pd.isnull(df['col_235']) or pd.isnull(df['col_234']):
                if pd.isnull(df['col_235']):
                    return 'Probable Reporting Error(15.2.2 is blank)'
                elif pd.isnull(float(df['col_234'])):
                    return 'Inconsistent'
            elif float(df['col_235']) > float(df['col_234']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #15.3.2b(240) <= 15.3.2a(239)
        def res65(df):
            if pd.isnull(df['col_240']) and pd.isnull(df['col_239']):
                return 'Blank'
            elif pd.isnull(df['col_240']) or pd.isnull(df['col_239']):
                if pd.isnull(df['col_240']):
                    return 'Probable Reporting Error(15.3.2b is blank)'
                elif pd.isnull(float(df['col_239'])):
                    return 'Inconsistent'
            elif float(df['col_240']) > float(df['col_239']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #15.3.3b(241) <= 15.3.3a(240)
        def res66(df):
            if pd.isnull(df['col_241']) and pd.isnull(df['col_240']):
                return 'Blank'
            elif pd.isnull(df['col_241']) or pd.isnull(df['col_240']):
                if pd.isnull(df['col_241']):
                    return 'Probable Reporting Error(15.3.3b is blank)'
                elif pd.isnull(float(df['col_240'])):
                    return 'Inconsistent'
            elif float(df['col_241']) > float(df['col_240']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #15.3.3c(242) <= 15.3.3b(241)
        def res67(df):
            if pd.isnull(df['col_242']) and pd.isnull(df['col_241']):
                return 'Blank'
            elif pd.isnull(df['col_242']) or pd.isnull(df['col_241']):
                if pd.isnull(df['col_242']):
                    return 'Probable Reporting Error(15.3.3c is blank)'
                elif pd.isnull(float(df['col_241'])):
                    return 'Inconsistent'
            elif float(df['col_242']) > float(df['col_241']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #15.4.2(248) <= 15.4.1(247)
        def res68(df):
            if pd.isnull(df['col_248']) and pd.isnull(df['col_247']):
                return 'Blank'
            elif pd.isnull(df['col_248']) or pd.isnull(df['col_247']):
                if pd.isnull(df['col_250']):
                    return 'Probable Reporting Error(15.4.2 is blank)'
                elif pd.isnull(float(df['col_247'])):
                    return 'Inconsistent'
            elif float(df['col_248']) > float(df['col_247']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
   
        #9.6.1 (139) <=9.1.1 (100)+9.1.2 (101)+9.1.3 (102)+9.1.4 (103)+9.1.5 (104)+9.1.6 (105)+9.1.7 (106)+9.1.8 (107)+9.1.13 (112)+9.1.14 (113)+9.1.15 (114)+9.1.16 (115)+9.1.17 (116)+9.1.18 (117)+9.1.19 (118)+9.1.20 (119)+9.1.21 (120)+9.2.1 (121)+9.2.2 (122)+9.2.3 (123)+9.3.1 (126)+9.3.2 (127)+9.3.3 (128)+9.4.1 (129)+9.4.2 (130)+9.4.3 (131)+9.4.5 (133)+9.4.6 (134)+9.5.1 (135)+9.5.2 (136)+9.5.3 (137)+9.5.4 (138)
        def res69(df):
            if pd.isnull(df['col_139']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_112']) and pd.isnull(df['col_113']) and pd.isnull(df['col_114']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_123']) and pd.isnull(df['col_126']) and pd.isnull(df['col_127']) and pd.isnull(df['col_128']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_131']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_135']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']) and pd.isnull(df['col_138']):
                return 'Blank'
            elif pd.isnull(df['col_139']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_112']) or pd.isnull(df['col_113']) or pd.isnull(df['col_114']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_123']) or pd.isnull(df['col_126']) or pd.isnull(df['col_127']) or pd.isnull(df['col_128']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_131']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_135']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']) or pd.isnull(df['col_138']):
                if pd.isnull(df['col_139']):
                    return 'Probable Reporting Error(9.6.1 is blank)'
                elif pd.isnull((float(df['col_100'])) + (float(df['col_101'])) + (float(df['col_102'])) + (float(df['col_103'])) + (float(df['col_104'])) + (float(df['col_105'])) + (float(df['col_106'])) + (float(df['col_107'])) + (float(df['col_112'])) + (float(df['col_113'])) + (float(df['col_114'])) + (float(df['col_115'])) + (float(df['col_116'])) + (float(df['col_117'])) + (float(df['col_118'])) + (float(df['col_119'])) + (float(df['col_120'])) + (float(df['col_121'])) + (float(df['col_122'])) + (float(df['col_123'])) + (float(df['col_126'])) + (float(df['col_127'])) + (float(df['col_128'])) + (float(df['col_129'])) + (float(df['col_130'])) + (float(df['col_131'])) + (float(df['col_133'])) + (float(df['col_134'])) + (float(df['col_135'])) + (float(df['col_136'])) + (float(df['col_137'])) + (float(df['col_138']))):
                    return 'Inconsistent'
            elif float(df['col_139']) > float(df['col_100']) > + (float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_112']) + float(df['col_113']) + float(df['col_114']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_126']) + float(df['col_127']) + float(df['col_128']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #9.6.2 (140) <=9.1.1 (100)+9.1.2 (101)+9.1.3 (102)+9.1.4 (103)+9.1.5 (104)+9.1.6 (105)+9.1.7 (106)+9.1.8 (107)+9.1.13 (112)+9.1.14 (113)+9.1.15 (114)+9.1.16 (115)+9.1.17 (116)+9.1.18 (117)+9.1.19 (118)+9.1.20 (119)+9.1.21 (120)+9.2.1 (121)+9.2.2 (122)+9.2.3 (123)+9.3.1 (126)+9.3.2 (127)+9.3.3 (128)+9.4.1 (129)+9.4.2 (130)+9.4.3 (131)+9.4.5 (133)+9.4.6 (134)+9.5.1 (135)+9.5.2 (136)+9.5.3 (137)+9.5.4 (138)
        def res70(df):
            if pd.isnull(df['col_140']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_112']) and pd.isnull(df['col_113']) and pd.isnull(df['col_114']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_123']) and pd.isnull(df['col_126']) and pd.isnull(df['col_127']) and pd.isnull(df['col_128']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_131']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_135']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']) and pd.isnull(df['col_138']):
                return 'Blank'
            elif pd.isnull(df['col_140']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_112']) or pd.isnull(df['col_113']) or pd.isnull(df['col_114']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_123']) or pd.isnull(df['col_126']) or pd.isnull(df['col_127']) or pd.isnull(df['col_128']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_131']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_135']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']) or pd.isnull(df['col_138']):
                if pd.isnull(df['col_140']):
                    return 'Probable Reporting Error(9.6.2 is blank)'
                elif pd.isnull((float(df['col_100'])) + (float(df['col_101'])) + (float(df['col_102'])) + (float(df['col_103'])) + (float(df['col_104'])) + (float(df['col_105'])) + (float(df['col_106'])) + (float(df['col_107'])) + (float(df['col_112'])) + (float(df['col_113'])) + (float(df['col_114'])) + (float(df['col_115'])) + (float(df['col_116'])) + (float(df['col_117'])) + (float(df['col_118'])) + (float(df['col_119'])) + (float(df['col_120'])) + (float(df['col_121'])) + (float(df['col_122'])) + (float(df['col_123'])) + (float(df['col_126'])) + (float(df['col_127'])) + (float(df['col_128'])) + (float(df['col_129'])) + (float(df['col_130'])) + (float(df['col_131'])) + (float(df['col_133'])) + (float(df['col_134'])) + (float(df['col_135'])) + (float(df['col_136'])) + (float(df['col_137'])) + (float(df['col_138']))):
                    return 'Inconsistent'
            elif float(df['col_140']) > float(df['col_100']) > + (float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_112']) + float(df['col_113']) + float(df['col_114']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_126']) + float(df['col_127']) + float(df['col_128']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df  

        #9.6.3 (141) <=9.1.1 (100)+9.1.2 (101)+9.1.3 (102)+9.1.4 (103)+9.1.5 (104)+9.1.6 (105)+9.1.7 (106)+9.1.8 (107)+9.1.13 (112)+9.1.14 (113)+9.1.15 (114)+9.1.16 (115)+9.1.17 (116)+9.1.18 (117)+9.1.19 (118)+9.1.20 (119)+9.1.21 (120)+9.2.1 (121)+9.2.2 (122)+9.2.3 (123)+9.3.1 (126)+9.3.2 (127)+9.3.3 (128)+9.4.1 (129)+9.4.2 (130)+9.4.3 (131)+9.4.5 (133)+9.4.6 (134)+9.5.1 (135)+9.5.2 (136)+9.5.3 (137)+9.5.4 (138)
        def res71(df):
            if pd.isnull(df['col_141']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_105']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_112']) and pd.isnull(df['col_113']) and pd.isnull(df['col_114']) and pd.isnull(df['col_115']) and pd.isnull(df['col_116']) and pd.isnull(df['col_117']) and pd.isnull(df['col_118']) and pd.isnull(df['col_119']) and pd.isnull(df['col_120']) and pd.isnull(df['col_121']) and pd.isnull(df['col_122']) and pd.isnull(df['col_123']) and pd.isnull(df['col_126']) and pd.isnull(df['col_127']) and pd.isnull(df['col_128']) and pd.isnull(df['col_129']) and pd.isnull(df['col_130']) and pd.isnull(df['col_131']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_135']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']) and pd.isnull(df['col_138']):
                return 'Blank'
            elif pd.isnull(df['col_141']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_105']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_112']) or pd.isnull(df['col_113']) or pd.isnull(df['col_114']) or pd.isnull(df['col_115']) or pd.isnull(df['col_116']) or pd.isnull(df['col_117']) or pd.isnull(df['col_118']) or pd.isnull(df['col_119']) or pd.isnull(df['col_120']) or pd.isnull(df['col_121']) or pd.isnull(df['col_122']) or pd.isnull(df['col_123']) or pd.isnull(df['col_126']) or pd.isnull(df['col_127']) or pd.isnull(df['col_128']) or pd.isnull(df['col_129']) or pd.isnull(df['col_130']) or pd.isnull(df['col_131']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_135']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']) or pd.isnull(df['col_138']):
                if pd.isnull(df['col_141']):
                    return 'Probable Reporting Error(9.6.3 is blank)'
                elif pd.isnull((float(df['col_100'])) + (float(df['col_101'])) + (float(df['col_102'])) + (float(df['col_103'])) + (float(df['col_104'])) + (float(df['col_105'])) + (float(df['col_106'])) + (float(df['col_107'])) + (float(df['col_112'])) + (float(df['col_113'])) + (float(df['col_114'])) + (float(df['col_115'])) + (float(df['col_116'])) + (float(df['col_117'])) + (float(df['col_118'])) + (float(df['col_119'])) + (float(df['col_120'])) + (float(df['col_121'])) + (float(df['col_122'])) + (float(df['col_123'])) + (float(df['col_126'])) + (float(df['col_127'])) + (float(df['col_128'])) + (float(df['col_129'])) + (float(df['col_130'])) + (float(df['col_131'])) + (float(df['col_133'])) + (float(df['col_134'])) + (float(df['col_135'])) + (float(df['col_136'])) + (float(df['col_137'])) + (float(df['col_138']))):
                    return 'Inconsistent'
            elif float(df['col_141']) > float(df['col_100']) > + (float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_105']) + float(df['col_106']) + float(df['col_107']) + float(df['col_112']) + float(df['col_113']) + float(df['col_114']) + float(df['col_115']) + float(df['col_116']) + float(df['col_117']) + float(df['col_118']) + float(df['col_119']) + float(df['col_120']) + float(df['col_121']) + float(df['col_122']) + float(df['col_123']) + float(df['col_126']) + float(df['col_127']) + float(df['col_128']) + float(df['col_129']) + float(df['col_130']) + float(df['col_131']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138'])):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df 

        #9.7.2(143) <= 9.7.1(142)
        def res72(df):
            if pd.isnull(df['col_143']) and pd.isnull(df['col_142']):
                return 'Blank'
            elif pd.isnull(df['col_143']) or pd.isnull(df['col_142']):
                if pd.isnull(df['col_143']):
                    return 'Probable Reporting Error(9.7.2 is blank)'
                elif pd.isnull(float(df['col_142'])):
                    return 'Inconsistent'
            elif float(df['col_143']) > float(df['col_142']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #9.7.3(144) <= 9.7.2(143)
        def res73(df):
            if pd.isnull(df['col_144']) and pd.isnull(df['col_143']):
                return 'Blank'
            elif pd.isnull(df['col_144']) or pd.isnull(df['col_143']):
                if pd.isnull(df['col_144']):
                    return 'Probable Reporting Error(9.7.3 is blank)'
                elif pd.isnull(float(df['col_143'])):
                    return 'Inconsistent'
            elif float(df['col_144']) > float(df['col_143']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.1b (166) <= 11.1.1a (165)
        def res74(df):
            if pd.isnull(df['col_166']) and pd.isnull(df['col_165']):
                return 'Blank'
            elif pd.isnull(df['col_166']) or pd.isnull(df['col_165']):
                if pd.isnull(df['col_166']):
                    return 'Probable Reporting Error(11.1.1b is blank)'
                elif pd.isnull(float(df['col_165'])):
                    return 'Inconsistent'
            elif float(df['col_166']) > float(df['col_165']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.1c (167) <= 11.1.1a (165)
        def res75(df):
            if pd.isnull(df['col_167']) and pd.isnull(df['col_165']):
                return 'Blank'
            elif pd.isnull(df['col_167']) or pd.isnull(df['col_165']):
                if pd.isnull(df['col_167']):
                    return 'Probable Reporting Error(11.1.1c is blank)'
                elif pd.isnull(float(df['col_165'])):
                    return 'Inconsistent'
            elif float(df['col_167']) > float(df['col_165']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.2b (169) <= 11.1.2a (168)
        def res76(df):
            if pd.isnull(df['col_169']) and pd.isnull(df['col_168']):
                return 'Blank'
            elif pd.isnull(df['col_169']) or pd.isnull(df['col_168']):
                if pd.isnull(df['col_169']):
                    return 'Probable Reporting Error(11.1.2b is blank)'
                elif pd.isnull(float(df['col_168'])):
                    return 'Inconsistent'
            elif float(df['col_169']) > float(df['col_168']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #11.1.2c (170) <= 11.1.2a (168)
        def res77(df):
            if pd.isnull(df['col_170']) and pd.isnull(df['col_168']):
                return 'Blank'
            elif pd.isnull(df['col_170']) or pd.isnull(df['col_168']):
                if pd.isnull(df['col_170']):
                    return 'Probable Reporting Error(11.1.2c is blank)'
                elif pd.isnull(float(df['col_170'])):
                    return 'Inconsistent'
            elif float(df['col_170']) > float(df['col_168']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df     
        
        #14.9.1 (220)<=14.3.1.a (194)+14.3.1.b (195)
        def res78(df):
            if pd.isnull(df['col_220']) and pd.isnull(df['col_194']) and pd.isnull(df['col_195']):
                return 'Blank'
            elif pd.isnull(df['col_220']) or pd.isnull(df['col_194']) or pd.isnull(df['col_195']):
                if pd.isnull(df['col_220']):
                    return 'Probable Reporting Error(14.9.1 is blank)'
                elif pd.isnull((float(df['col_194'])) + (float(df['col_195']))):
                    return 'Inconsistent'
            elif float(df['col_220']) > float(df['col_194']) + float(df['col_195']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df      

        #14.9.2 (221)<=14.3.2.a (196)+14.3.2.b (197)
        def res79(df):
            if pd.isnull(df['col_221']) and pd.isnull(df['col_196']) and pd.isnull(df['col_197']):
                return 'Blank'
            elif pd.isnull(df['col_221']) or pd.isnull(df['col_196']) or pd.isnull(df['col_197']):
                if pd.isnull(df['col_221']):
                    return 'Probable Reporting Error(14.9.2 is blank)'
                elif pd.isnull((float(df['col_196'])) + (float(df['col_197']))):
                    return 'Inconsistent'
            elif float(df['col_221']) > float(df['col_196']) + float(df['col_197']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df     

        #14.13 (229)<=14.12.1 (224)+14.12.2 (225)+14.12.3 (226)+14.12.4 (227)+14.12.5 (228)
        def res80(df):
            if pd.isnull(df['col_229']) and pd.isnull(df['col_224']) and pd.isnull(df['col_225']) and pd.isnull(df['col_226']) and pd.isnull(df['col_227']) and pd.isnull(df['col_228']):
                return 'Blank'
            elif pd.isnull(df['col_229']) or pd.isnull(df['col_224']) or pd.isnull(df['col_225']) or pd.isnull(df['col_226']) or pd.isnull(df['col_227']) or pd.isnull(df['col_228']):
                if pd.isnull(df['col_229']):
                    return 'Probable Reporting Error(14.13 is blank)'
                elif pd.isnull((float(df['col_224'])) + (float(df['col_225'])) + (float(df['col_226'])) + (float(df['col_227'])) + (float(df['col_228']))):
                    return 'Inconsistent'
            elif float(df['col_229']) > float(df['col_224']) + float(df['col_225']) + float(df['col_226']) + float(df['col_227']) + float(df['col_228']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df      

        #14.8.2(216) <= 14.8.1(215)
        def res81(df):
            if pd.isnull(df['col_216']) and pd.isnull(df['col_215']):
                return 'Blank'
            elif pd.isnull(df['col_216']) or pd.isnull(df['col_215']):
                if pd.isnull(df['col_216']):
                    return 'Probable Reporting Error(14.8.2 is blank)'
                elif pd.isnull(float(df['col_215'])):
                    return 'Inconsistent'
            elif float(df['col_216']) > float(df['col_215']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df
                
        #15.3.4b(244) <= 15.3.4a(243)
        def res82(df):
            if pd.isnull(df['col_244']) and pd.isnull(df['col_243']):
                return 'Blank'
            elif pd.isnull(df['col_244']) or pd.isnull(df['col_243']):
                if pd.isnull(df['col_244']):
                    return 'Probable Reporting Error(15.3.4b is blank)'
                elif pd.isnull(float(df['col_243'])):
                    return 'Inconsistent'
            elif float(df['col_244']) > float(df['col_243']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df       
    
       #15.3.4d(246) <= 15.3.4c(245)
        def res83(df):
            if pd.isnull(df['col_246']) and pd.isnull(df['col_245']):
                return 'Blank'
            elif pd.isnull(df['col_246']) or pd.isnull(df['col_245']):
                if pd.isnull(df['col_246']):
                    return 'Probable Reporting Error(15.3.4d is blank)'
                elif pd.isnull(float(df['col_245'])):
                    return 'Inconsistent'
            elif float(df['col_246']) > float(df['col_245']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df 

        # Renaming column names===============================CHC


        df['1.1.1<=1.1'] = df.apply(res1, axis=1)
        df['15.3.1.b<=15.3.1.a'] = df.apply(res2, axis=1)
        df['1.2.4<=1.1'] = df.apply(res3, axis=1)
        df['1.2.5<=1.1'] = df.apply(res4, axis=1)
        df['1.2.7<=1.1'] = df.apply(res5, axis=1)
        df['1.3.1.a<=1.3.1'] = df.apply(res6, axis=1)
        df['1.3.2<=2.1'] = df.apply(res7, axis=1)
        df['1.4.4>=1.4.3'] = df.apply(res8, axis=1)
        df['1.5.2<=1.5.1'] = df.apply(res9, axis=1)
        df['1.5.1<=1.1'] = df.apply(res10, axis=1)
        df['1.5.3<=1.5.2'] = df.apply(res11, axis=1)
        df['1.6.1.b<=1.6.1.a'] = df.apply(res12, axis=1)
        df['1.6.1.a<=1.1'] = df.apply(res13, axis=1)
        df['1.6.1.c<=1.6.1.b'] = df.apply(res14, axis=1)
        df['1.6.1.e<=1.6.1.d'] = df.apply(res15, axis=1)
        df['2.1.1<=2.1'] = df.apply(res16, axis=1)
        df['3.1<=2.1'] = df.apply(res17, axis=1)
        df['3.1.1<=3.1'] = df.apply(res18, axis=1)
        df['4.1.1.a+4.1.1.b+4.1.3>=2.1'] = df.apply(res19, axis=1)
        df['4.1.2<=4.1.1.a+4.1.1.b'] = df.apply(res20, axis=1)
        df['4.3.2.a<=4.3.1.a+4.3.1.b+4.2'] = df.apply(res21, axis=1)
        df['4.3.2.b<=4.3.2.a'] = df.apply(res22, axis=1)
        df['4.3.3<=4.3.1.a+4.3.1.b+4.2'] = df.apply(res23, axis=1)
        df['4.4.1<=4.1.1.a+4.1.1.b'] = df.apply(res24, axis=1)
        df['4.4.2<=4.4.1'] = df.apply(res25, axis=1)
        df['4.4.3<=4.1.1.a+4.1.1.b'] = df.apply(res26, axis=1)
        df['5.2<=5.1'] = df.apply(res27, axis=1)
        df['6.3<=2.1'] = df.apply(res28, axis=1)
        df['6.4<=2.1'] = df.apply(res29, axis=1)
        df['7.2.1<=7.1.1'] = df.apply(res30, axis=1)
        df['7.2.2<=7.1.2'] = df.apply(res31, axis=1)
        df['8.2.3<=2.1'] = df.apply(res32, axis=1)
        df['8.4<=2.1'] = df.apply(res33, axis=1)
        df['8.7<=8.3+8.4+8.5'] = df.apply(res34, axis=1)
        df['8.17.1<=8.1.1'] = df.apply(res35, axis=1)
        df['8.17.2<=8.2.1+8.2.2+8.2.3+8.2.4'] = df.apply(res36, axis=1)
        df['9.1.1<=4.1.1.a+4.1.1.b'] = df.apply(res37, axis=1)
        df['9.1.2<=4.1.1.a+4.1.1.b'] = df.apply(res38, axis=1)
        df['9.1.9<=4.1.1.a+4.1.1.b'] = df.apply(res39, axis=1)
        df['9.1.13<=4.1.1.a+4.1.1.b'] = df.apply(res40, axis=1)
        df['9.2.4.a+9.2.4.b<=9.2.1+ 9.2.2'] = df.apply(res41, axis=1)
        df['11.2.2<=11.2.1'] = df.apply(res42, axis=1)
        df['12.1.2.a<=12.1.1.a'] = df.apply(res43, axis=1)
        df['12.1.2.b<=12.1.1.b'] = df.apply(res44, axis=1)
        df['12.1.3.a<=12.1.1.a'] = df.apply(res45, axis=1)
        df['12.1.3.b<=12.1.1.b'] = df.apply(res46, axis=1)
        df['14.2.1+14.2.2>=14.1.1+14.1.2+14.1.3+14.1.4+14.1.5+14.1.6+14.1.7+14.1.8+14.1.9'] = df.apply(res47, axis=1)
        df['14.3.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res48, axis=1)
        df['14.4.1<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res49, axis=1)
        df['14.4.2<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res50, axis=1)
        df['14.4.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res51, axis=1)
        df['14.4.4<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res52, axis=1)
        df['14.4.5<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res53, axis=1)
        df['14.4.6<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res54, axis=1)
        df['14.4.7<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res55, axis=1)
        df['14.4.8<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'] = df.apply(res56, axis=1)
        df['14.6.1<=14.5'] = df.apply(res57, axis=1)
        df['14.6.2<=14.5'] = df.apply(res58, axis=1)
        df['14.6.3<=14.5'] = df.apply(res59, axis=1)
        df['14.6.4<=14.5'] = df.apply(res60, axis=1)
        df['14.6.5<=14.5'] = df.apply(res61, axis=1)
        df['14.7<=14.5'] = df.apply(res62, axis=1)
        df['14.14.2<=14.14.1'] = df.apply(res63, axis=1)
        df['15.2.2<=15.2.1'] = df.apply(res64, axis=1)
        df['15.3.2.b<=15.3.2.a'] = df.apply(res65, axis=1)
        df['15.3.3.b<=15.3.3.a'] = df.apply(res66, axis=1)
        df['15.3.3.c<=15.3.3.b'] = df.apply(res67, axis=1)
        df['15.4.2<=15.4.1'] = df.apply(res68, axis=1)
        df['9.6.1<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'] = df.apply(res69, axis=1)
        df['9.6.2<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'] = df.apply(res70, axis=1)
        df['9.6.3<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'] = df.apply(res71, axis=1)
        df['9.7.2<=9.7.1'] = df.apply(res72, axis=1)
        df['9.7.3<=9.7.2'] = df.apply(res73, axis=1)
        df['11.1.1.b<=11.1.1.a'] = df.apply(res74, axis=1)
        df['11.1.1.c<=11.1.1.a'] = df.apply(res75, axis=1)
        df['11.1.2.b<=11.1.2.a'] = df.apply(res76, axis=1)
        df['11.1.2.c<=11.1.2.a'] = df.apply(res77, axis=1)
        df['14.9.1<=14.3.1.a+14.3.1.b'] = df.apply(res78, axis=1)
        df['14.9.2<=14.3.2.a+14.3.2.b'] = df.apply(res79, axis=1)
        df['14.13<=14.12.1+14.12.2+14.12.3+14.12.4+14.12.5'] = df.apply(res80, axis=1)
        df['14.8.2<=14.8.1'] = df.apply(res81, axis=1)
        df['15.3.4.b<=15.3.4.a'] = df.apply(res82, axis=1)
        df['15.3.4.d<=15.3.4.c'] = df.apply(res83, axis=1)


        # Merging all the renamed columns
        # ===============================
        df= pd.concat ([df ['1.1.1<=1.1'],
                        df ['15.3.1.b<=15.3.1.a'],
                        df ['1.2.4<=1.1'],
                        df ['1.2.5<=1.1'],
                        df ['1.2.7<=1.1'],
                        df ['1.3.1.a<=1.3.1'],
                        df ['1.3.2<=2.1'],
                        df ['1.4.4>=1.4.3'],
                        df ['1.5.2<=1.5.1'],
                            df ['1.5.1<=1.1'],
                            df ['1.5.3<=1.5.2'],
                            df ['1.6.1.b<=1.6.1.a'],
                            df ['1.6.1.a<=1.1'],
                            df ['1.6.1.c<=1.6.1.b'],
                            df ['1.6.1.e<=1.6.1.d'],
                            df ['2.1.1<=2.1'],
                            df ['3.1<=2.1'],
                            df ['3.1.1<=3.1'],
                                df ['4.1.1.a+4.1.1.b+4.1.3>=2.1'],
                                df ['4.1.2<=4.1.1.a+4.1.1.b'],
                                df ['4.3.2.a<=4.3.1.a+4.3.1.b+4.2'],
                                df ['4.3.2.b<=4.3.2.a'],
                                df ['4.3.3<=4.3.1.a+4.3.1.b+4.2'],
                                df ['4.4.1<=4.1.1.a+4.1.1.b'],
                                df ['4.4.2<=4.4.1'],
                                df ['4.4.3<=4.1.1.a+4.1.1.b'],
                                df ['5.2<=5.1'],
                                df ['6.3<=2.1'],
                                    df ['6.4<=2.1'],
                                    df ['7.2.1<=7.1.1'],
                                    df ['7.2.2<=7.1.2'],
                                    df ['8.2.3<=2.1'],
                                    df ['8.4<=2.1'],
                                    df ['8.7<=8.3+8.4+8.5'],
                                    df ['8.17.1<=8.1.1'],
                                    df ['8.17.2<=8.2.1+8.2.2+8.2.3+8.2.4'],
                                    df ['9.1.1<=4.1.1.a+4.1.1.b'],
                                    df ['9.1.2<=4.1.1.a+4.1.1.b'],
                                        df ['9.1.9<=4.1.1.a+4.1.1.b'],
                                        df ['9.1.13<=4.1.1.a+4.1.1.b'],
                                        df ['9.2.4.a+9.2.4.b<=9.2.1+ 9.2.2'],
                                        df ['11.2.2<=11.2.1'],
                                        df ['12.1.2.a<=12.1.1.a'],
                                        df ['12.1.2.b<=12.1.1.b'],
                                        df ['12.1.3.a<=12.1.1.a'],
                                        df ['12.1.3.b<=12.1.1.b'],
                                        df ['14.2.1+14.2.2>=14.1.1+14.1.2+14.1.3+14.1.4+14.1.5+14.1.6+14.1.7+14.1.8+14.1.9'],
                                        df ['14.3.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                        df ['14.4.1<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.4.2<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.4.3<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.4.4<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.4.5<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.4.6<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.4.7<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.4.8<=14.3.1.a+14.3.1.b+14.3.2.a+14.3.2.b'],
                                            df ['14.6.1<=14.5'],
                                            df ['14.6.2<=14.5'],
                                            df ['14.6.3<=14.5'],
                                            df ['14.6.4<=14.5'],
                                            df ['14.6.5<=14.5'],
                                                df ['14.7<=14.5'],
                                                df ['14.14.2<=14.14.1'],
                                                df ['15.2.2<=15.2.1'],
                                                df ['15.3.2.b<=15.3.2.a'],
                                                df ['15.3.3.b<=15.3.3.a'],
                                                df ['15.3.3.c<=15.3.3.b'],
                                                df ['15.4.2<=15.4.1'],
                                                df ['9.6.1<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'],
                                                df ['9.6.2<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'],
                                                df ['9.6.3<=9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4'],
                                                df ['9.7.2<=9.7.1'],
                                                df ['9.7.3<=9.7.2'],
                                                    df ['11.1.1.b<=11.1.1.a'],
                                                    df ['11.1.1.c<=11.1.1.a'],
                                                    df ['11.1.2.b<=11.1.2.a'],
                                                    df ['11.1.2.c<=11.1.2.a'],
                                                    df ['14.9.1<=14.3.1.a+14.3.1.b'],
                                                    df ['14.9.2<=14.3.2.a+14.3.2.b'],
                                                    df ['14.13<=14.12.1+14.12.2+14.12.3+14.12.4+14.12.5'],
                                                    df ['14.8.2<=14.8.1'],
                                                    df ['15.3.4.b<=15.3.4.a'],
                                                    df ['15.3.4.d<=15.3.4.c']], axis=1)

        # Mergining current result of modified checks with original dataframe and displaying it on screen
        frames = [df_, df]
        print(frames)
        df = pd.concat(frames, axis=1, sort=False)
        #df = df.dropna(axis=0, subset=['col_2'])
        self.tableView.setModel(PandasModel(df))

        msg = QMessageBox()
        msg.setWindowTitle("Validation Completion Message")
        msg.setText("Community Health Centers Validation Complete")
        msg.setIcon(QMessageBox.Information)
        msg.exec()

        return df



    '''
    # Validation for HSC
    '''
    def HSC_Validate(self):
        global df, table_result

        filterString = self.comboBox.currentText()
        
        df = df_.loc[df_['col_12'] == filterString]
        print(df)
       
        print('Entered HSC_Validate')

        # Modified Checks of HSC

        # 4.3 (56) <= 2.1.1.a (39) + 2.1.1.b (40) + 2.2 (43) (For recurring data items)
        def res1(df):

            # If all elements are null
            if pd.isnull(df['col_56']) and pd.isnull(df['col_39']) and pd.isnull(df['col_40']) and pd.isnull(df['col_43']):
                return 'Blank'

            # If any one element is null
            elif pd.isnull(df['col_56']) or pd.isnull(df['col_39']) or pd.isnull(df['col_40']) or pd.isnull(df['col_43']):
                if pd.isnull(df['col_56']) and not pd.isnull(float(df['col_39']) + float(df['col_40']) + float(df['col_43'])):
                    return 'Probable Reporting Error'
                else:
                    if pd.isnull(float(df['col_39']) + float(df['col_40']) + float(df['col_43'])) and not pd.isnull(df['col_56']):
                        return 'Probable Reporting Error'
                    else:
                        return 'Probable Reporting Error'

            # If value exists for all the elements
            else:
                lhs_value = float(df['col_56'])
                rhs_value = float(df['col_39']) + \
                    float(df['col_40']) + float(df['col_43'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # 1.1(col_22) >= 1.1.1(col_23) (for related data items)
        def res2(df):
            if pd.isnull(df['col_22']) and pd.isnull(df['col_23']):
                return 'Blank'
            elif pd.isnull(df['col_22']) or pd.isnull(df['col_23']):
                if pd.isnull(df['col_22']):
                    return 'Inconsistent'
                elif pd.isnull(df['col_23']):
                    return 'Probable Reporting Error (1.1.1 is blank)'
            elif df['col_22'] < df['col_23']:
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.3.1.a(col_33) <= 1.3.1(col_32) (for related data items)

        def res3(df):
            if (df['col_33'] is None) and (df['col_32'] is None):
                return 'Blank'
            elif (df['col_33'] is None) or (df['col_32'] is None):
                if (df['col_33'] is None):
                    return 'Inconsistent'
                elif (df['col_32'] is None):
                    return 'Probable Reporting Error (1.3.1 is blank)'
            if (df['col_33'] is None) > (df['col_32'] is None):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        #  1.2.7 (30) <= 1.1(22) (For recurring data items)
        def res4(df):
            if pd.isnull(float(df['col_30'])) and pd.isnull(float(df['col_22'])):
                return 'Blank'
            elif pd.isnull(float(df['col_30'])) or pd.isnull(float(df['col_22'])):
                if pd.isnull(df['col_30']):
                    return 'Inconsistent'
                elif pd.isnull(df['col_22']):
                    return 'Probable Reporting Error (1.1 is blank)'
            # If value exists for all the elements
            else:

                lhs_value = float(df['col_30'])
                rhs_value = float(df['col_22'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # # 1.5.1.a (37) <= 1.1 (22) (for unrelated data items)
        def res5(df):
            if pd.isnull(float(df['col_37'])) and pd.isnull(float(df['col_22'])):
                return 'Blank'
            elif pd.isnull(float(df['col_37'])) or pd.isnull(float(df['col_22'])):
                if pd.isnull(df['col_37']):
                    return 'Probable Reporting Error(1.5.1.a is blank)'
                elif pd.isnull(df['col_22']):
                    return 'Inconsistent (1.1 is blank)'
            elif float(df['col_37']) > float(df['col_22']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 1.5.1.b (col_38) <= 1.5.1.a (col_37) (for related data items)
        def res6(df):

            if pd.isnull(float(df['col_38'])) and pd.isnull(float(df['col_37'])):
                return 'Blank'
            elif pd.isnull(float(df['col_38'])) or pd.isnull(float(df['col_37'])):
                if pd.isnull(df['col_38']):
                    return 'Probable Reporting Error (1.5.1.b is blank)'
                elif pd.isnull(df['col_37']):
                    return 'Inconsistent (1.5.1.a is blank)'
            elif float(df['col_38']) > float(df['col_37']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

          # 2.1.2 (col_41) <= 2.1.1.a(col_39) + 2.1.1.b(col_40) (for unrelated data items)
        def res7(df):
            if pd.isnull(df['col_41']) and pd.isnull(df['col_39']) and pd.isnull(df['col_40']):
                return 'Blank'
            elif pd.isnull(df['col_41']) or pd.isnull(df['col_39']) or pd.isnull(df['col_40']):
                if pd.isnull(df['col_41']):
                    return 'Probable Reporting Error (2.1.2 is blank)'
                elif pd.isnull(df['col_39']):
                    return 'Inconsistent 2.1.1.a is null'
                elif pd.isnull(df['col_40']):
                    return 'Inconsistent 2.1.1.b is null'
            elif float(df['col_41']) > float(df['col_39']) + float(df['col_40']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

            # 2.1.3 (col_42) <= 2.1.1.a(col_39) + 2.1.1.b(col_40) (For recurring data items)
        def res8(df):

            if pd.isnull(df['col_42']) and pd.isnull(df['col_39']) and pd.isnull(df['col_40']):
                return 'Blank'
            elif pd.isnull(df['col_42']) or pd.isnull(df['col_39']) or pd.isnull(df['col_40']):
                if pd.isnull(df['col_42']) and not pd.isnull(float(df['col_39']) + float(df['col_40'])):
                    return 'Probable Reporting Error'
                else:
                    if pd.isnull(float(df['col_39']) + float(df['col_40'])) and not pd.isnull(df['col_42']):
                        return 'Probable Reporting Error'
                    else:
                        return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_42'])
                rhs_value = float(df['col_39']) + float(df['col_40'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # 2.2.2(col_45) <= 2.2 (col_43) (For recurring data items) ----------need update
        def res9(df):
            if pd.isnull(df['col_45']) and pd.isnull(df['col_43']):
                return 'Blank'
            elif pd.isnull(df['col_45']) or pd.isnull(df['col_43']):
                if pd.isnull(df['col_45']) and not pd.isnull(float(df['col_43'])):
                    return 'Probable Reporting Error'
                else:
                    if pd.isnull(float(df['col_43'])) and not pd.isnull(df['col_45']):
                        return 'Probable Reporting Error'
                    else:
                        return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_45'])
                rhs_value = float(df['col_43'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

        # 4.4(col_57)<= 2.1.1.a(col_39) + 2.1.1.b(col_40) + 2.2(col_43) (For recurring data items)
        def res10(df):
            if pd.isnull(df['col_57']) and pd.isnull(df['col_39']) and pd.isnull(df['col_40']) and pd.isnull(df['col_43']):
                return 'Blank'
            elif pd.isnull(df['col_57']) and pd.isnull(df['col_39']) and pd.isnull(df['col_40']) and pd.isnull(df['col_43']):
                if pd.isnull(df['col_57']) and not pd.isnull(float(df['col_39']) + float(df['col_40']) + float(df['col_43'])):
                    return 'Probable Reporting Error'
                else:
                    if pd.isnull(float(df['col_39']) + float(df['col_40']) + float(df['col_43'])) and not pd.isnull(df['col_42']):
                        return 'Probable Reporting Error'
                    else:
                        return 'Probable Reporting Error'

            # If value exists for all the elements
            else:

                lhs_value = float(df['col_57'])
                rhs_value = float(df['col_39']) + \
                    float(df['col_40']) + float(df['col_43'])

                if lhs_value <= rhs_value:
                    if lhs_value < (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'consistent'
                else:
                    if lhs_value > (0.5*rhs_value):
                        return 'Probable Reporting Error'
                    else:
                        return 'Inconsistent'
            return df

         # 6.1.1(col_73) <= 3.1.1.a(col_46) + 3.1.1.b(col_47)
        def res11(df):
            if pd.isnull(df['col_73']) and pd.isnull(df['col_46']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_73']) or pd.isnull(df['col_46']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_73']):
                    return 'Probable Reporting Error (6.1.1 is blank)'
                elif pd.isnull(df['col_46']):
                    return 'Inconsistent (3.1.1.a is blank)'
                elif pd.isnull(df['col_47']):
                    return 'Inconsistent (3.1.1.b is blank)'
            elif float(df['col_73']) > float(df['col_46']) + float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 6.1.9(col_81) <= 3.1.1.a(col_46) + 3.1.1.b(col_47)
        def res12(df):

            if pd.isnull(df['col_81']) and pd.isnull(df['col_46']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_81']) or pd.isnull(df['col_46']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_81']):
                    return 'Probable Reporting Error (6.1.9 is blank)'
                elif pd.isnull(df['col_46']):
                    return 'Inconsistent (3.1.1.a is blank)'
                elif pd.isnull(df['col_47']):
                    return 'Inconsistent (3.1.1.b is blank)'
            elif float(df['col_81']) > float(df['col_46']) + float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 6.1.13(col_85) <= 3.1.1.a(col_46) + 3.1.1.b(col_47)
        def res13(df):
            if pd.isnull(df['col_85']) and pd.isnull(df['col_46']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_85']) or pd.isnull(df['col_46']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_85']):
                    return 'Probable Reporting Error  (6.1.13 is blank)'
                elif pd.isnull(df['col_46']):
                    return 'Inconsistent (6.1.13 is blank)'
                elif pd.isnull(df['col_47']):
                    return 'Inconsistent (6.1.13 is blank)'
            elif float(df['col_85']) > float(df['col_46']) + float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 2.2.1(col_44) <= 2.2(col_43)
        def res14(df):
            if pd.isnull(df['col_44']) and pd.isnull(df['col_43']):
                return 'Blank'
            elif pd.isnull(df['col_44']) or pd.isnull(df['col_43']):
                if pd.isnull(df['col_44']):
                    return 'Probable Reporting Error (2.2.1 is blank)'
                elif pd.isnull(df['col_43']):
                    return 'Inconsistent (2.2 is blank)'
            elif float(df['col_44']) > float(df['col_43']):
                return 'Inconsistent (check fails)'
            else:
                return 'consistent'
            return df

         # 3.1.2(col_48) <= 3.1.1.a(col_46)+ 3.1.1.b(col_47)
        def res15(df):
            if pd.isnull(df['col_48']) and pd.isnull(df['col_46']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_48']) or pd.isnull(df['col_46']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_48']):
                    return 'Probable Reporting Error (3.1.2 is blank)'
                elif pd.isnull(df['col_46']):
                    return 'Inconsistent(3.1.1.a is blank)'
                elif pd.isnull(df['col_47']):
                    return 'Inconsistent (3.1.1.b is blank)'
            elif float(df['col_48']) > float(df['col_46']) + float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

         # 3.3.1 <= 3.1.1.a + 3.1.1.b
        def res16(df):

            if pd.isnull(df['col_51']) and pd.isnull(df['col_46']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_51']) or pd.isnull(df['col_46']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_51']):
                    return 'Probable Reporting Error (3.3.1 is blank)'
                elif pd.isnull(df['col_46']):
                    return 'Inconsistent (3.3.1.a is blank)'
                elif pd.isnull(df['col_47']):
                    return 'Inconsistent (3.3.1.b is blank)'
            elif float(df['col_51']) > float(df['col_46']) + float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

         # 3.3.2 <= 3.3.1
        def res17(df):
            if pd.isnull(df['col_52']) and pd.isnull(df['col_51']):
                return 'Blank'
            elif pd.isnull(df['col_52']) or pd.isnull(df['col_51']):
                if pd.isnull(df['col_52']):
                    return 'Probable Reporting Error (3.3.2 is blank)'
                elif pd.isnull(df['col_67']):
                    return 'Inconsistent (3.3.1 is blank)'
            if float(df['col_52']) > float(df['col_51']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 3.3.3<=3.1.1.a+3.1.1.b
        def res18(df):

            if pd.isnull(df['col_53']) and pd.isnull(df['col_46']) and pd.isnull(df['col_47']):
                return 'Blank'
            elif pd.isnull(df['col_53']) or pd.isnull(df['col_46']) or pd.isnull(df['col_47']):
                if pd.isnull(df['col_53']):
                    return 'Probable Reporting Error (3.3.3 is blank)'
                elif pd.isnull(float(df['col_46']) + float(df['col_47'])):
                    return 'Inconsistent'
            elif float(df['col_53']) > float(df['col_46']) + float(df['col_47']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

         # 4.1 <= 2.1.1.a + 2.1.1.b
        def res19(df):

            if pd.isnull(df['col_54']) and pd.isnull(df['col_39']) and pd.isnull(df['col_40']):
                return 'Blank'
            elif pd.isnull(df['col_54']) or pd.isnull(df['col_39']) or pd.isnull(df['col_40']):
                if pd.isnull(df['col_54']):
                    return 'Probable Reporting Error(4.1 is blank)'
                elif pd.isnull(float(df['col_39']) + float(df['col_40'])):
                    return 'Inconsistent'
            elif float(df['col_54']) > float(df['col_39']) + float(df['col_40']):
                return 'Inconsistent (check fails)'
            else:
                return 'consistent'
            return df

        # 5.2 <= 2.1.1.a + 2.1.1.b + 2.2
        def res20(df):

            if pd.isnull(df['col_59']) and pd.isnull(df['col_39']) and pd.isnull(df['col_40']) and pd.isnull(df['col_43']):
                return 'Blank'
            elif pd.isnull(df['col_59']) or pd.isnull(df['col_39']) or pd.isnull(df['col_40']) or pd.isnull(df['col_43']):
                if pd.isnull(df['col_59']):
                    return 'Probable Reporting Error(5.2 is blank)'
                elif pd.isnull(float(df['col_39']) + float(df['col_40']) + float(df['col_43'])):
                    return 'Inconsistent'
            elif float(df['col_59']) > float(df['col_39']) + float(df['col_40']) + float(df['col_43']):
                return 'Inconsistent (check fails)'
            else:
                return 'consistent'
            return df

         # 6.2.4.a + 6.2.4.b <= 6.2.1 + 6.2.2
        def res21(df):

            if pd.isnull(df['col_97']) and pd.isnull(df['col_98']) and pd.isnull(df['col_94']) and pd.isnull(df['col_95']):
                return 'Blank'
            elif pd.isnull(df['col_97']) or pd.isnull(df['col_98']) or pd.isnull(df['col_94']) or pd.isnull(df['col_95']):
                if pd.isnull(df['col_97']):
                    return 'Probable Reporting Error (6.2.4.a is blank)'
                elif pd.isnull(float(df['col_94']) + float(df['col_95'])):
                    return 'Probable Reporting Error'
            elif float(df['col_97']) + float(df['col_98']) > float(df['col_94']) + float(df['col_95']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 6.6.1<=6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4
        def res22(df):

            if pd.isnull(df['col_112']) and pd.isnull(df['col_73']) and pd.isnull(df['col_74']) and pd.isnull(df['col_75']) and pd.isnull(df['col_76']) and pd.isnull(df['col_77']) and pd.isnull(df['col_78']) and pd.isnull(df['col_79']) and pd.isnull(df['col_80']) and pd.isnull(df['col_85']) and pd.isnull(df['col_86']) and pd.isnull(df['col_87']) and pd.isnull(df['col_88']) and pd.isnull(df['col_89']) and pd.isnull(df['col_90']) and pd.isnull(df['col_91']) and pd.isnull(df['col_92']) and pd.isnull(df['col_93']) and pd.isnull(df['col_94']) and pd.isnull(df['col_95']) and pd.isnull(df['col_96']) and pd.isnull(df['col_99']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_108']) and pd.isnull(df['col_109']) and pd.isnull(df['col_110']):
                return 'Blank'
            elif pd.isnull(df['col_112']) or pd.isnull(df['col_73']) or pd.isnull(df['col_74']) or pd.isnull(df['col_75']) or pd.isnull(df['col_76']) or pd.isnull(df['col_77']) or pd.isnull(df['col_78']) or pd.isnull(df['col_79']) or pd.isnull(df['col_80']) or pd.isnull(df['col_85']) or pd.isnull(df['col_86']) or pd.isnull(df['col_87']) or pd.isnull(df['col_88']) or pd.isnull(df['col_89']) or pd.isnull(df['col_90']) or pd.isnull(df['col_91']) or pd.isnull(df['col_92']) or pd.isnull(df['col_93']) or pd.isnull(df['col_122']) or pd.isnull(df['col_94']) or pd.isnull(df['col_96']) or pd.isnull(df['col_99']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_108']) or pd.isnull(df['col_109']) or pd.isnull(df['col_110']):
                if pd.isnull(df['col_112']):
                    return 'Probable Reporting Error (6.6.1 is blank)'
                elif pd.isnull(float(df['col_73']) + float(df['col_74']) + float(df['col_75']) + float(df['col_76']) + float(df['col_77']) + float(df['col_78']) + float(df['col_79']) + float(df['col_80']) + float(df['col_85']) + float(df['col_86']) + float(df['col_87']) + float(df['col_88']) + float(df['col_89']) + float(df['col_90']) + float(df['col_91']) + float(df['col_92']) + float(df['col_93']) + float(df['col_94']) + float(df['col_95']) + float(df['col_96']) + float(df['col_99']) + float(df['col_100']) + float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110'])):
                    return 'Inconsistent'
            elif float(df['col_112']) > float(df['col_73']) + float(df['col_74']) + float(df['col_75']) + float(df['col_76']) + float(df['col_77']) + float(df['col_78']) + float(df['col_79']) + float(df['col_80']) + float(df['col_85']) + float(df['col_86']) + float(df['col_87']) + float(df['col_88']) + float(df['col_89']) + float(df['col_90']) + float(df['col_91']) + float(df['col_92']) + float(df['col_93']) + float(df['col_94']) + float(df['col_95']) + float(df['col_96']) + float(df['col_99']) + float(df['col_100']) + float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110']):
                return 'Inconsistent (check fails)'
            else:
                return 'consistent'
            return df

        # 6.6.2<=6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4
        def res23(df):

            if pd.isnull(df['col_113']) and pd.isnull(df['col_73']) and pd.isnull(df['col_74']) and pd.isnull(df['col_75']) and pd.isnull(df['col_76']) and pd.isnull(df['col_77']) and pd.isnull(df['col_78']) and pd.isnull(df['col_79']) and pd.isnull(df['col_80']) and pd.isnull(df['col_85']) and pd.isnull(df['col_86']) and pd.isnull(df['col_87']) and pd.isnull(df['col_88']) and pd.isnull(df['col_89']) and pd.isnull(df['col_90']) and pd.isnull(df['col_91']) and pd.isnull(df['col_92']) and pd.isnull(df['col_93']) and pd.isnull(df['col_94']) and pd.isnull(df['col_95']) and pd.isnull(df['col_96']) and pd.isnull(df['col_99']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_108']) and pd.isnull(df['col_109']) and pd.isnull(df['col_110']):
                return 'Blank'
            elif pd.isnull(df['col_113']) or pd.isnull(df['col_73']) or pd.isnull(df['col_74']) or pd.isnull(df['col_75']) or pd.isnull(df['col_76']) or pd.isnull(df['col_77']) or pd.isnull(df['col_78']) or pd.isnull(df['col_79']) or pd.isnull(df['col_80']) or pd.isnull(df['col_85']) or pd.isnull(df['col_86']) or pd.isnull(df['col_87']) or pd.isnull(df['col_88']) or pd.isnull(df['col_89']) or pd.isnull(df['col_90']) or pd.isnull(df['col_91']) or pd.isnull(df['col_92']) or pd.isnull(df['col_93']) or pd.isnull(df['col_122']) or pd.isnull(df['col_94']) or pd.isnull(df['col_96']) or pd.isnull(df['col_99']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_108']) or pd.isnull(df['col_109']) or pd.isnull(df['col_110']):
                if pd.isnull(df['col_113']):
                    return 'Probable Reporting Error (6.6.2 is blank)'
                elif pd.isnull(float(df['col_73']) + float(df['col_74']) + float(df['col_75']) + float(df['col_76']) + float(df['col_77']) + float(df['col_78']) + float(df['col_79']) + float(df['col_80']) + float(df['col_85']) + float(df['col_86']) + float(df['col_87']) + float(df['col_88']) + float(df['col_89']) + float(df['col_90']) + float(df['col_91']) + float(df['col_92']) + float(df['col_93']) + float(df['col_94']) + float(df['col_95']) + float(df['col_96']) + float(df['col_99']) + float(df['col_100']) + float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110'])):
                    return 'Inconsistent'
            elif float(df['col_113']) > float(df['col_73']) + float(df['col_74']) + float(df['col_75']) + float(df['col_76']) + float(df['col_77']) + float(df['col_78']) + float(df['col_79']) + float(df['col_80']) + float(df['col_85']) + float(df['col_86']) + float(df['col_87']) + float(df['col_88']) + float(df['col_89']) + float(df['col_90']) + float(df['col_91']) + float(df['col_92']) + float(df['col_93']) + float(df['col_94']) + float(df['col_95']) + float(df['col_96']) + float(df['col_99']) + float(df['col_100']) + float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110']):
                return 'Inconsistent (check fails)'
            else:
                return 'consistent'
            return df

         # 6.6.3<=6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4
        def res24(df):

            if pd.isnull(df['col_114']) and pd.isnull(df['col_73']) and pd.isnull(df['col_74']) and pd.isnull(df['col_75']) and pd.isnull(df['col_76']) and pd.isnull(df['col_77']) and pd.isnull(df['col_78']) and pd.isnull(df['col_79']) and pd.isnull(df['col_80']) and pd.isnull(df['col_85']) and pd.isnull(df['col_86']) and pd.isnull(df['col_87']) and pd.isnull(df['col_88']) and pd.isnull(df['col_89']) and pd.isnull(df['col_90']) and pd.isnull(df['col_91']) and pd.isnull(df['col_92']) and pd.isnull(df['col_93']) and pd.isnull(df['col_94']) and pd.isnull(df['col_95']) and pd.isnull(df['col_96']) and pd.isnull(df['col_99']) and pd.isnull(df['col_100']) and pd.isnull(df['col_101']) and pd.isnull(df['col_102']) and pd.isnull(df['col_103']) and pd.isnull(df['col_104']) and pd.isnull(df['col_106']) and pd.isnull(df['col_107']) and pd.isnull(df['col_108']) and pd.isnull(df['col_109']) and pd.isnull(df['col_110']):
                return 'Blank'
            elif pd.isnull(df['col_114']) or pd.isnull(df['col_73']) or pd.isnull(df['col_74']) or pd.isnull(df['col_75']) or pd.isnull(df['col_76']) or pd.isnull(df['col_77']) or pd.isnull(df['col_78']) or pd.isnull(df['col_79']) or pd.isnull(df['col_80']) or pd.isnull(df['col_85']) or pd.isnull(df['col_86']) or pd.isnull(df['col_87']) or pd.isnull(df['col_88']) or pd.isnull(df['col_89']) or pd.isnull(df['col_90']) or pd.isnull(df['col_91']) or pd.isnull(df['col_92']) or pd.isnull(df['col_93']) or pd.isnull(df['col_122']) or pd.isnull(df['col_94']) or pd.isnull(df['col_96']) or pd.isnull(df['col_99']) or pd.isnull(df['col_100']) or pd.isnull(df['col_101']) or pd.isnull(df['col_102']) or pd.isnull(df['col_103']) or pd.isnull(df['col_104']) or pd.isnull(df['col_106']) or pd.isnull(df['col_107']) or pd.isnull(df['col_108']) or pd.isnull(df['col_109']) or pd.isnull(df['col_110']):
                if pd.isnull(df['col_114']):
                    return 'Probable Reporting Error (6.6.1 is blank)'
                elif pd.isnull(float(df['col_73']) + float(df['col_74']) + float(df['col_75']) + float(df['col_76']) + float(df['col_77']) + float(df['col_78']) + float(df['col_79']) + float(df['col_80']) + float(df['col_85']) + float(df['col_86']) + float(df['col_87']) + float(df['col_88']) + float(df['col_89']) + float(df['col_90']) + float(df['col_91']) + float(df['col_92']) + float(df['col_93']) + float(df['col_94']) + float(df['col_95']) + float(df['col_96']) + float(df['col_99']) + float(df['col_100']) + float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110'])):
                    return 'Inconsistent'
            elif float(df['col_114']) > float(df['col_73']) + float(df['col_74']) + float(df['col_75']) + float(df['col_76']) + float(df['col_77']) + float(df['col_78']) + float(df['col_79']) + float(df['col_80']) + float(df['col_85']) + float(df['col_86']) + float(df['col_87']) + float(df['col_88']) + float(df['col_89']) + float(df['col_90']) + float(df['col_91']) + float(df['col_92']) + float(df['col_93']) + float(df['col_94']) + float(df['col_95']) + float(df['col_96']) + float(df['col_99']) + float(df['col_100']) + float(df['col_101']) + float(df['col_102']) + float(df['col_103']) + float(df['col_104']) + float(df['col_106']) + float(df['col_107']) + float(df['col_108']) + float(df['col_109']) + float(df['col_110']):
                return 'Inconsistent (check fails)'
            else:
                return 'consistent'
            return df

        # 6.7.3<=6.7.2
        def res25(df):

            if pd.isnull(df['col_117']) and pd.isnull(df['col_116']):
                return 'Blank'
            elif pd.isnull(df['col_117']) or pd.isnull(df['col_116']):
                if pd.isnull(df['col_117']):
                    return 'Probable Reporting Error (6.7.3 is blank)'
            elif pd.isnull(df['col_116']):
                return 'Inconsistent (6.7.2 is blank)'
            elif float(df['col_117']) > float(df['col_116']):
                return 'Inconsistent (check fails)'
            else:
                return 'consistent'
            return df

        # 10.1.2<=10.1.1
        def res26(df):
            if pd.isnull(df['col_144']) and pd.isnull(df['col_143']):
                return 'Blank'
            elif pd.isnull(df['col_144']) or pd.isnull(df['col_143']):
                if pd.isnull(df['col_144']):
                    return 'Probable Reporting Error (10.1.2 is blank)'
                elif pd.isnull(df['col_143']):
                    return 'Inconsistent (10.1.1 is blank)'
            elif float(df['col_144']) > float(df['col_143']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 10.2.1.b<=10.2.1.a
        def res27(df):

            if pd.isnull(df['col_146']) and pd.isnull(df['col_145']):
                return 'Blank'
            elif pd.isnull(df['col_146']) or pd.isnull(df['col_145']):
                if pd.isnull(df['col_146']):
                    return 'Probable Reporting Error (10.2.1.b is blank)'
                elif pd.isnull(df['col_145']):
                    return 'Inconsistent (10.2.1.a is blank)'
            elif float(df['col_146']) > float(df['col_145']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 3.1.1.a+3.1.1.b+3.1.3 >= 2.1.1.a+2.1.1.b+2.2
        def res28(df):

            if pd.isnull(df['col_46']) and pd.isnull(df['col_47']) and pd.isnull(df['col_49']) and pd.isnull(df['col_39']) and pd.isnull(df['col_40']) and pd.isnull(df['col_43']):
                return 'Blank'
            elif pd.isnull(df['col_46']) or pd.isnull(df['col_47']) or pd.isnull(df['col_49']) or pd.isnull(df['col_39']) or pd.isnull(df['col_40']) or pd.isnull(df['col_43']):
                if pd.isnull(df['col_46']):
                    return 'Inconsistent  (3.1.1.a is blank)'
                elif pd.isnull(df['col_47']):
                    return 'Inconsistent (3.1.1.b is blank)'
                elif pd.isnull(df['col_49']):
                    return 'Inconsistent (3.1.3 is blank)'
                elif pd.isnull(float(df['col_39']) + float(df['col_40']) + float(df['col_43'])):
                    return 'Probable Reporting Error'
            elif float(df['col_46']) + float(df['col_47']) + float(df['col_49']) < float(df['col_39']) + float(df['col_40']) + float(df['col_43']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 8.1.1.c<=8.1.1.a
        def res29(df):

            if pd.isnull(df['col_131']) and pd.isnull(df['col_129']):
                return 'Blank'
            elif pd.isnull(df['col_131']) or pd.isnull(df['col_129']):
                if pd.isnull(df['col_131']):
                    return 'Probable Reporting Error (8.1.1.c is blank)'
                elif pd.isnull(df['col_129']):
                    return 'Inconsistent (8.1.1.a is blank)'
            elif float(df['col_131']) > float(df['col_129']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 9.2.1 + 9.2.2>= 9.1.1+ 9.1.2+ 9.1.3+ 9.1.4+ 9.1.5+ 9.1.6+ 9.1.7+ 9.1.8
        def res30(df):

            if pd.isnull(df['col_140']) and pd.isnull(df['col_141']) and pd.isnull(df['col_132']) and pd.isnull(df['col_133']) and pd.isnull(df['col_134']) and pd.isnull(df['col_135']) and pd.isnull(df['col_136']) and pd.isnull(df['col_137']) and pd.isnull(df['col_138']) and pd.isnull(df['col_139']):
                return 'Blank'
            elif pd.isnull(df['col_140']) or pd.isnull(df['col_141']) or pd.isnull(df['col_132']) or pd.isnull(df['col_133']) or pd.isnull(df['col_134']) or pd.isnull(df['col_135']) or pd.isnull(df['col_136']) or pd.isnull(df['col_137']) or pd.isnull(df['col_138']) or pd.isnull(df['col_139']):
                if pd.isnull(float(df['col_140']) + float(df['col_141'])):
                    return 'Inconsistent'
                # elif pd.isnull(df['col_141']):
                #     return 'Inconsistent (9.2.2 is blank)'
                elif pd.isnull(float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138']) + float(df['col_139'])):
                    return 'Probable Reporting Error'
            elif float(df['col_140']) + float(df['col_141']) < float(df['col_132']) + float(df['col_133']) + float(df['col_134']) + float(df['col_135']) + float(df['col_136']) + float(df['col_137']) + float(df['col_138']) + float(df['col_139']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # 8.1.1.b<=8.1.1.a
        def res31(df):

            if pd.isnull(df['col_130']) and pd.isnull(df['col_129']):
                return 'Blank'
            elif pd.isnull(df['col_130']) or pd.isnull(df['col_129']):
                if pd.isnull(df['col_130']):
                    return 'Probable Reporting Error (8.1.1.b is blank)'
                elif pd.isnull(df['col_129']):
                    return 'Inconsistent (8.1.1.a is blank)'
            elif float(df['col_130']) > float(df['col_129']):
                return 'Inconsistent'
            else:
                return 'consistent'
            return df

        # Renaming column names
        # =====================
        df['4.3 <= 2.1.1.a + 2.1.1.b + 2.2'] = df.apply(res1, axis=1)
        df['1.1 <= 1.1.1'] = df.apply(res2, axis=1)
        df['1.3.1.a <= 1.3.1'] = df.apply(res3, axis=1)
        df['1.2.7 <= 1.1'] = df.apply(res4, axis=1)
        df['1.5.1.a <= 1.1'] = df.apply(res5, axis=1)
        df['1.5.1.b <= 1.5.1.a'] = df.apply(res6, axis=1)
        df['2.1.2 <= 2.1.1.a + 2.1.1.b'] = df.apply(res7, axis=1)
        df['2.1.3 <= 2.1.1.a + 2.1.1.b'] = df.apply(res8, axis=1)
        df['2.2.2 <= 2.2'] = df.apply(res9, axis=1)
        df['4.4 <= 2.1.1.a + 2.1.1.b + 2.2'] = df.apply(res10, axis=1)
        df['6.1.1 <= 3.1.1.a + 3.1.1.b'] = df.apply(res11, axis=1)
        df['6.1.9 <= 3.1.1.a + 3.1.1.b'] = df.apply(res12, axis=1)
        df['6.1.13 <= 3.1.1.a + 3.1.1.b'] = df.apply(res13, axis=1)
        df['2.2.1 <= 2.2'] = df.apply(res14, axis=1)
        df['3.1.2 <= 3.1.1.a + 3.1.1.b'] = df.apply(res15, axis=1)
        df['3.3.1 <= 3.1.1.a + 3.1.1.b'] = df.apply(res16, axis=1)
        df['3.3.2 <= 3.3.1'] = df.apply(res17, axis=1)
        df['3.3.3 <= 3.1.1.a + 3.1.1.b'] = df.apply(res18, axis=1)
        df['4.1 <= 2.1.1.a + 2.1.1.b'] = df.apply(res19, axis=1)
        df['5.2 <= 2.1.1.a + 2.1.1.b + 2.2'] = df.apply(res20, axis=1)
        df['6.2.4.a + 6.2.4.b <= 6.2.1 + 6.2.2'] = df.apply(res21, axis=1)
        df['6.6.1<=6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4'] = df.apply(res22, axis=1)
        df['6.6.2<=6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4'] = df.apply(res23, axis=1)
        df['6.6.3<=6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4'] = df.apply(res24, axis=1)
        df['6.7.3<=6.7.2'] = df.apply(res25, axis=1)
        df['10.1.2<=10.1.1'] = df.apply(res26, axis=1)
        df['10.2.1.b<=10.2.1.a'] = df.apply(res27, axis=1)
        df['3.1.1.a+3.1.1.b+3.1.3 >= 2.1.1.a+2.1.1.b+2.2'] = df.apply(res28, axis=1)
        df['8.1.1.c<=8.1.1.a'] = df.apply(res29, axis=1)
        df['9.2.1 + 9.2.2>= 9.1.1+ 9.1.2+ 9.1.3+ 9.1.4+ 9.1.5+ 9.1.6+ 9.1.7+ 9.1.8'] = df.apply(res30, axis=1)
        df['8.1.1.b<=8.1.1.a'] = df.apply(res31, axis=1)

        # Merging all the renamed columns
        # ===============================
        df = pd.concat([df['4.3 <= 2.1.1.a + 2.1.1.b + 2.2'],
                        df['1.1 <= 1.1.1'],
                        df['1.3.1.a <= 1.3.1'],
                        df['1.2.7 <= 1.1'],
                        df['1.5.1.a <= 1.1'],
                        df['1.5.1.b <= 1.5.1.a'],
                        df['2.1.2 <= 2.1.1.a + 2.1.1.b'],
                            df['2.1.3 <= 2.1.1.a + 2.1.1.b'],
                            df['2.2.2 <= 2.2'],
                            df['4.4 <= 2.1.1.a + 2.1.1.b + 2.2'],
                            df['6.1.1 <= 3.1.1.a + 3.1.1.b'],
                            df['6.1.9 <= 3.1.1.a + 3.1.1.b'],
                            df['6.1.13 <= 3.1.1.a + 3.1.1.b'],
                            df['2.2.1 <= 2.2'],
                            df['3.1.2 <= 3.1.1.a + 3.1.1.b'],
                            df['3.3.1 <= 3.1.1.a + 3.1.1.b'],
                                df['3.3.2 <= 3.3.1'],
                                df['3.3.3 <= 3.1.1.a + 3.1.1.b'],
                                df['4.1 <= 2.1.1.a + 2.1.1.b'],
                                df['5.2 <= 2.1.1.a + 2.1.1.b + 2.2'],
                                df['6.2.4.a + 6.2.4.b <= 6.2.1 + 6.2.2'],
                                df['6.6.1<=6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4'],
                                df['6.6.2<=6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4'],
                                df['6.6.3<=6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4'],
                                df['6.7.3<=6.7.2'],
                                df['10.1.2<=10.1.1'],
                                    df['10.2.1.b<=10.2.1.a'],
                                    df['3.1.1.a+3.1.1.b+3.1.3 >= 2.1.1.a+2.1.1.b+2.2'],
                                    df['8.1.1.c<=8.1.1.a'],
                                    df['9.2.1 + 9.2.2>= 9.1.1+ 9.1.2+ 9.1.3+ 9.1.4+ 9.1.5+ 9.1.6+ 9.1.7+ 9.1.8'],
                                    df['8.1.1.b<=8.1.1.a']], axis=1)



        # Mergining current result of modified checks with original dataframe and displaying it on screen
        frames = [df_, df]
        print(frames)
        df = pd.concat(frames, axis=1, sort=False)
        #df = df.dropna(axis=0, subset=['col_2'])
        self.tableView.setModel(PandasModel(df))

        msg = QMessageBox()
        msg.setWindowTitle("Validation Completion Message")
        msg.setText("Health Sub Centers Validation Complete")
        msg.setIcon(QMessageBox.Information)
        msg.exec()

        return df

    '''
    # Filter to decide which filter button user clicked
    # =================================================
    '''
    def eventFilter(self, target, event):
        if target == self.pushButton_4 and event.type() == QtCore.QEvent.MouseButtonPress:
            print("Checked State")
            self.pushButton_4.clicked.connect(self.onSelectState)
            return True

        elif target == self.pushButton_5 and event.type() == QtCore.QEvent.MouseButtonPress:
            print("Checked District")
            self.pushButton_5.clicked.connect(self.onSelectDistrict)
            return True

        elif target == self.pushButton_6 and event.type() == QtCore.QEvent.MouseButtonPress:
            print("Checked Facility Name")
            self.pushButton_6.clicked.connect(self.onSelectFacilityName)
            return True

        elif target == self.pushButton_7 and event.type() == QtCore.QEvent.MouseButtonPress:
            print("Checked Month")
            self.pushButton_7.clicked.connect(self.onSelectMonth)
            return True

        elif target == self.pushButton_8 and event.type() == QtCore.QEvent.MouseButtonPress:
            print("Checked Year")
            self.pushButton_8.clicked.connect(self.onSelectYear)
            return True

        return False


    ################################################################################
    # Select State

    # Filter State Functionality
    def onSelectState(self, index):
        global list_set
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(Dialog)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 200 }')
        font = self.menu.font()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 3
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # list storing state data
        item = df['col_3'].to_list()

        ''' Adding search regex filter '''
        # model = QtGui.QStandardItemModel(len(item), 1)
        # model.setHorizontalHeaderLabels(['States'])
        # for row, state in enumerate(item):
        #     itm = QtGui.QStandardItem(state)
        #     model.setItem(row, 0, itm)

        # filter_proxy_model = QtCore.QSortFilterProxyModel()
        # filter_proxy_model.setSourceModel(model)
        # filter_proxy_model.setFilterCaseSensitivity(Qt.CaseInsensitive)
        # filter_proxy_model.setFilterKeyColumn(3)

        # search_field = QtWidgets.QLineEdit()
        # search_field.setStyleSheet('font-size: 15px; height: 20px; width: 240px')
        # search_field.textChanged.connect(self.onTextChanged)

        # searchAction = QWidgetAction(self.menu)
        # searchAction.setDefaultWidget(search_field)
        # self.menu.addAction(searchAction)

        ''''''

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all", self.menu)

        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelect)

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(df)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                         QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuClose)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)
        self.pushButton_4.setMenu(self.menu)

    def onTextChanged(self):
        print('working//////////////////')

    # method to check -> uncheck and vice versa
    def slotSelect(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuClose(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        self.filterdata()
        self.menu.close()

    # Filter data columnwise
    def filterdata(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_3'].shape[0])])

        # for i in range(df.shape[0]):
        j = 0
        for j in range(df.shape[0]):
            item = df['col_3'].to_list()

            print(self.keywords[self.col])
            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select State')
        df = df[final_lst]
        self.tableView.setModel(PandasModel(df))
        return df

    ################################################################################
    # Select District

        # Select District Functionality
    def onSelectDistrict(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(Dialog)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 200 }')
        font = self.menu.font()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 5
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all", self.menu)

        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectDistrict)

        # list storing state data
        list_set = df['col_5'].to_list()

        item = list_set

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                         QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseDistrict)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)
        self.pushButton_5.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectDistrict(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseDistrict(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataDistrict()
        self.menu.close()

    # Filter data columnwise
    def filterdataDistrict(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_5'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_5'].shape[0]):
            item = df['col_5'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select Month')
        df = df[final_lst]
        print(df)
        self.tableView.setModel(PandasModel(df))
        return df

    ################################################################################
    # Select Facility Name

    # Select FacilityName Functionality

    def onSelectFacilityName(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(Dialog)
        self.menu.setStyleSheet('QMenu { menu-scrollable: 20; width: 400 }')
        font = self.menu.font()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 14
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all", self.menu)

        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectFacilityName)

        # list storing Facility Name data
        list_set = df['col_14'].to_list()

        item = list_set

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                         QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseFacilityName)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)

        ############# Always set Pushbutton ####################
        self.pushButton_6.setMenu(self.menu)

        # self.scrollbar = QtWidgets.QScrollArea(widgetResizable=False)
        # self.scrollbar.setWidget(self.menu)

        # self.tabwidget = QtWidgets.QTabWidget()
        # self.tabwidget.addTab(self.scrollbar, "Tab1")
        # self.layout = QtWidgets.QVBoxLayout(Dialog)
        # self.layout.addWidget(self.tabwidget)
        # self.tabwidget.setTabBarAutoHide(True)
        # self.tabwidget.tabCloseRequested.connect(self.close_tab)
        

    # # close tab
    # def close_tab(self,index):
    #     self.tabwidget.removeTab(index)


    # method to check -> uncheck and vice versa
    def slotSelectFacilityName(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseFacilityName(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataFacilityName()
        self.menu.close()
        
        # self.scrollbar.close()
        # self.layout.deleteLater()


    # Filter data columnwise
    def filterdataFacilityName(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_14'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_14'].shape[0]):
            item = df['col_14'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select District')

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        print(df)
        self.tableView.setModel(PandasModel(df))
        return df

    ################################################################################
    # Select Month

    # Select Month Filter

    def onSelectMonth(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(Dialog)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 200 }')
        font = self.menu.font()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 1
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all", self.menu)

        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectMonth)

        # list storing Facility Name data
        list_set = df['col_1'].to_list()

        item = list_set

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                         QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseMonth)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)

        ############# Always set Pushbutton ####################
        self.pushButton_7.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectMonth(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseMonth(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataMonth()
        self.menu.close()

    # Filter data columnwise
    def filterdataMonth(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_1'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_1'].shape[0]):
            item = df['col_1'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select Month')

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        print(df)
        self.tableView.setModel(PandasModel(df))
        return df

    ################################################################################
    # Select Year

    # Select Year Filter
    def onSelectYear(self, index):
        self.keywords = dict([(i, []) for i in range(df.shape[0])])
        print(self.keywords)
        self.menu = QtWidgets.QMenu(Dialog)
        self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 200 }')
        font = self.menu.font()
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.menu.setFont(font)

        index = 2
        self.col = index

        data_unique = []

        self.checkBoxs = []

        # Selectall added into Dropdown
        checkBox = QtWidgets.QCheckBox("Select all", self.menu)

        # All the checkboxes are enabled to check
        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(checkBox)
        self.menu.addAction(checkableAction)
        checkBox.setChecked(True)
        checkBox.stateChanged.connect(self.slotSelectYear)

        # list storing Facility Name data
        list_set = df['col_2'].to_list()

        item = list_set

        # looping to fill checkboxes, initially all checkboxes will be checked
        for i in range(len(item)):
            if item[i] not in data_unique:
                data_unique.append(item[i])
                checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                checkBox.setChecked(True)
                checkableAction = QtWidgets.QWidgetAction(self.menu)
                checkableAction.setDefaultWidget(checkBox)
                self.menu.addAction(checkableAction)
                self.checkBoxs.append(checkBox)

        # Ok, cancel button
        btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                         QtCore.Qt.Vertical, self.menu)

        # ok selected
        btn.accepted.connect(self.menuCloseYear)
        # rejected , nothing selected
        btn.rejected.connect(self.menu.close)

        checkableAction = QtWidgets.QWidgetAction(self.menu)
        checkableAction.setDefaultWidget(btn)
        self.menu.addAction(checkableAction)

        ############# Always set Pushbutton ####################
        self.pushButton_8.setMenu(self.menu)

    # method to check -> uncheck and vice versa
    def slotSelectYear(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

    # after ok selected
    def menuCloseYear(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataYear()
        self.menu.close()

    # Filter data columnwise
    def filterdataYear(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_2'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_2'].shape[0]):
            item = df['col_2'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select Year')

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        print(df)
        self.tableView.setModel(PandasModel(df))
        return df


    # To count summary of the Modified Checks
    # =======================================
    def summaryReport(self):
        global final_result_summ1, final_result_summ2, col_sum, list_SctterPlot_Inc, list_SctterPlot_PRErr, list_SctterPlot_FName

        # For Health Sub Centre
        if FType == 'Health Sub Centre':
            df_SummReport = df.iloc[:, 200:]     ## Taking columns after 200th

            val_Description = [
                                'Number of mothers provided full course of 180 IFA tablets after delivery <= Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) + Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+ Number of Institutional Deliveries conducted',
                                'Out of the total ANC registered, number registered within 1st trimester (within 12 weeks)<=Total number of pregnant women registered for ANC', 
                                'Out of the new cases of PW with hypertension detected, cases managed at institution <=New cases of PW with hypertension detected',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC',
                                'Number of PW tested using POC test for Syphilis<=Total number of pregnant women registered for ANC',
                                'Out of the above, Number of PW found sero positive for Syphilis<=Number of PW tested using POC test for Syphilis',
                                'Number of PW given Tablet Misoprostol during home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                'Number of newborns received 7 Home Based Newborn Care (HBNC) visits in case of Home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                'Number of newborns received 6 HBNC  visits after Institutional Delivery<=Number of Institutional Deliveries conducted',
                                'Number of mothers provided 360 Calcium tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted',
                                    'Child immunisation - Vitamin K1 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                    'Child immunisation - OPV 0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                    'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                    'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted',
                                    'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                    'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                    'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                    'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                    'Women receiving 1st post partum checkup within 48 hours of home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                    'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunized - Female<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                    'Number of cases of AEFI - Abscess<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                        'Number of cases of AEFI - Death<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                        'Number of cases of AEFI - Others<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                        'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held',
                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg <=Number of Hb tests conducted',
                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                        'Live Birth - Male+Live Birth - Female+Still Birth>=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted',
                                        'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                        'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= 9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8',
                                        'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria' ]

        # For Primary Health Centre
        elif FType == 'Primary Health Centre':
            df_SummReport = df.iloc[:, 305:]     ## Taking columns after 305th
            val_Description = [
                                'Child immunisation - Vitamin K (Birth Dose) <= Live Birth - Male+Live Birth - Female',
                                'Out of the total ANC registered, number registered within 1st trimester (within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC', 
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC', 
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC', 
                                'Number of PW given Tablet Misoprostol during home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                'Number of newborns received 7 Home Based Newborn Care (HBNC) visits in case of Home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                    'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC', 
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Live Birth - Male+Live Birth - Female+Still Birth>=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception <=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth <=Live Birth - Male+Live Birth - Female',
                                        'Women receiving 1st post partum checkup within 48 hours of home delivery<=Number of Home Deliveries attended by Skill Birth Attendant (SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of mothers provided 360 Calcium tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - OPV 0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunized - Female<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                            'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= 14.1.1+14.1.2+14.1.3+14.1.4+14.1.5+14.1.6+14.1.7+14.1.8',
                                                'Number of Left Against Medical Advice (LAMA) cases<=Male Admissions +Female Admissions',
                                                'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                'number positive for HIV (Number confirmed positive at ICTCs)<=out of the above, Number screened positive',
                                                'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                'Number of cases of AEFI - Abscess<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                    'Number of cases of AEFI - Death<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                    'Number of cases of AEFI - Others<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                    'Out of the new cases of PW with hypertension detected, cases managed at institution <=New cases of PW with hypertension detected',
                                                    'Immunisation sessions held <=Immunisation sessions planned',
                                                    'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held', 
                                                    'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria', 
                                                    'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria',
                                                    'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                    'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                    'Inpatient - Malaria <=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Inpatient Deaths - Male<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Inpatient Deaths - Female<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                        'Male HIV - Number Positive<=Male HIV - Number Tested',
                                                        'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis',
                                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female']


        # For Sub District Hospital
        elif FType == 'Community Health Centre':
            df_SummReport = df.iloc[:, 318:]     ## Taking columns after 305th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                    'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                    'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                        'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                        'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                            'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                            'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                            'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                            'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                            'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology',
                                                'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                'Emergency - Burn<= Patients registered at Emergency Department',
                                                'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                    'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                    'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                    'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                    'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                    'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                    'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                    'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                    'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                    'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                    'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Immunisation sessions held <=Immunisation sessions planned ',
                                                        'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                        'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                        'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                        'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                        'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                        'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                        'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn – Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                        'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                        'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                        'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis'
                                                        ]

        # For Primary Health Centre
        elif FType == 'Sub District Hospital':
            df_SummReport = df.iloc[:, 321:]     ## Taking columns after 321th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                    'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                        'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                        'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                            'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                            'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                'Tests Positive for JE<=Tests Conducted for JE',
                                                'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology)',
                                                    'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                    'Emergency - Burn<= Patients registered at Emergency Department',
                                                    'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                        'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                        'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                        'Emergency - CVA ( Cerebovascular Disease)<= Patients registered at Emergency Department',
                                                        'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                        'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Immunisation sessions held <=Immunisation sessions planned ',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                            'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                            'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                            'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                            'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn – Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                            'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis']
            
        # For District Hospital
        elif FType == 'District Hospital':
            df_SummReport = df.iloc[:, 326:]     ## Taking columns after 326th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                'Tests Positive for JE<=Tests Conducted for JE',
                                                'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology',
                                                'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                    'Emergency - Burn<= Patients registered at Emergency Department',
                                                    'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                    'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                    'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                    'Emergency - CVA ( Cerebovascular Disease)<= Patients registered at Emergency Department',
                                                        'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                        'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Immunisation sessions held <=Immunisation sessions planned ',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                            'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                            'Total number of blood units issued during the month>=Number of blood units issued (Excluding C-Section)',
                                                            'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                            'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                            'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn – Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                            'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                            'Major Surgeries excluding Obstetrics, Gynaecology and Opthalmology etc.<=Operation major (General and spinal anaesthesia)',
                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis']



        '''
        ## First Summary Report
        ## ---------------------
        '''

        # count_Consistent = []
        count_Inconsistent = []
        # count_Blank = []
        count_ProbableRErr = []
        Columns = list(df_SummReport.columns.values.tolist())

        for col_name in Columns:

            # c1 = df_SummReport[col_name].str.count("consistent").sum()
            # count_Consistent.append(c1)

            c2 = df_SummReport[col_name].str.count("Inconsistent").sum()
            count_Inconsistent.append(c2)

            # c3 = df_SummReport[col_name].str.count("Blank").sum()
            # count_Blank.append(c3)

            c4 = df_SummReport[col_name].str.count("Probable Reporting Error").sum()
            count_ProbableRErr.append(c4)

        print(len(val_Description), len(count_Inconsistent), len(count_ProbableRErr))


        final_result_summ1 = pd.DataFrame({"Conditions": df_SummReport.columns, 
                                            "Description": val_Description,
                                            # "Consistent": count_Consistent,
                                                "Inconsistent": count_Inconsistent,
                                                    "Probable Reporting Error": count_ProbableRErr})

        
        def select_col_SumSheet(X):
            global cnt21, cnt22

            # COLORS
            c = ['background-color:  #C00000',              #>=50% RED
                    '']

            # Counting total numnber of facility names and calculating percentage and 
            # Coloring percentage >=25% to RED
            count_rows = df.shape[0]

            mask_21 = X['Inconsistent']/count_rows >= 0.25
            cnt21 = mask_21.values.sum()    

            mask_22 = X['Probable Reporting Error']/count_rows >= 0.25
            cnt22 = mask_22.values.sum()   

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[1], X.index, columns=X.columns)

            df1.loc[mask_21, 'Inconsistent'] = c[0]
            df1.loc[mask_22, 'Probable Reporting Error'] = c[0]
            return df1

        final_result_summ1 = final_result_summ1.style.apply(select_col_SumSheet, axis=None)


        '''
        ## Second Summary Report
        ## --------------------
        '''
        summ2_countInconsistent = []
        summ2_countProbableRErr = []
        All_Blank = []

        # Iterating over indices of each row and calculating number of Blanks for each Facility Name 
        for index in range(len(df_SummReport)):
            '''   For no. of Inconsistent   '''
            inconsistent = df_SummReport.iloc[index, :].str.count("Inconsistent").sum()
            summ2_countInconsistent.append(inconsistent)
            
            '''   For no. of Probable Reporting Errors   '''
            probableRErr = df_SummReport.iloc[index, :].str.count('Probable Reporting Error').sum()
            summ2_countProbableRErr.append(probableRErr)

            blank = df_SummReport.iloc[index, :].str.count("Blank").sum()
            if blank == 0:
                All_Blank.append('Yes')
            else:
                All_Blank.append('No')
                

        final_result_summ2 = pd.DataFrame({ "Facility Name": df['col_14'].tolist(),
                                                "Inconsistent": summ2_countInconsistent,
                                                    "Probable Reporting Error": summ2_countProbableRErr,
                                                        "All Blank": All_Blank,
                                                        })

                                                    
        # For Health Sub Centre
        if FType == 'Health Sub Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/31 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 31 * 100
        
        # For Primary Health Centre
        elif FType == 'Primary Health Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/78 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 78 * 100

        # For Community Health Centre
        elif FType == 'Community Health Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/83 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 83 * 100

        # For Sub District Hospital
        elif FType == 'Sub District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/85 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 85 * 100

        # For District Hospital
        elif FType == 'District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/85 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 85 * 100
            

        len0 = len(final_result_summ2['PercentageInc'])
        len1 = len(final_result_summ2['PercentagePRErr'])

        col_percentageInc_rank = []
        col_percentageErr_rank = []
        # col_percentageInc_Bucket = []
        # col_percentageErr_Bucket = []

        # For All Inconsistents in all five ranges respectively
        # Blank list for Inconsistents
        range1_Inc, range2_Inc, range3_Inc, range4_Inc, range5_Inc = [], [], [], [], []
        # Blank list for Probable Reporting Error
        range1_PRE, range2_PRE, range3_PRE, range4_PRE, range5_PRE = [], [], [], [], []

        for k in range(0,len0):
            primcol_Inc = final_result_summ2['PercentageInc'][k]
            if primcol_Inc < 5 :
                col_percentageInc_rank.append('1')
                #col_percentageInc_Bucket.append('< 5')
                range1_Inc.append(primcol_Inc)
            elif primcol_Inc >= 5 and primcol_Inc < 10:
                col_percentageInc_rank.append('2')
                #col_percentageInc_Bucket.append('5 & < 10')
                range2_Inc.append(primcol_Inc)
            elif primcol_Inc >= 10 and primcol_Inc < 25:
                col_percentageInc_rank.append('3')
                #col_percentageInc_Bucket.append('10 & < 25')
                range3_Inc.append(primcol_Inc)
            elif primcol_Inc >= 25 and primcol_Inc < 50:
                col_percentageInc_rank.append('4')
                #col_percentageInc_Bucket.append('25 & < 5')
                range4_Inc.append(primcol_Inc)
            elif primcol_Inc >= 50:
                col_percentageInc_rank.append('5')
                #col_percentageInc_Bucket.append('>= 5')
                range5_Inc.append(primcol_Inc)
            else:
                col_percentageInc_rank.append('NaN')
                #col_percentageInc_Bucket.append('NaN')

        for i in range(0,len1):
            primcol_Prc = final_result_summ2['PercentagePRErr'][i]
            if primcol_Prc < 5 :
                col_percentageErr_rank.append('1')
                #col_percentageErr_Bucket.append('< 5')
                range1_PRE.append(primcol_Prc)
            elif primcol_Prc >= 5 and primcol_Prc < 10:
                col_percentageErr_rank.append('2')
                #col_percentageErr_Bucket.append('5 & < 10')
                range2_PRE.append(primcol_Prc)
            elif primcol_Prc >= 10 and primcol_Prc < 25:
                col_percentageErr_rank.append('3')
                #col_percentageErr_Bucket.append('10 & < 25')
                range3_PRE.append(primcol_Prc)
            elif primcol_Prc >= 25 and primcol_Prc < 50:
                col_percentageErr_rank.append('4')
                #col_percentageErr_Bucket.append('25 & < 50')
                range4_PRE.append(primcol_Prc)
            elif primcol_Prc >= 50:
                col_percentageErr_rank.append('5')
                #col_percentageErr_Bucket.append('>= 50')
                range5_PRE.append(primcol_Prc)
            else:
                col_percentageErr_rank.append('NaN')
                #col_percentageErr_Bucket.append('NaN')

        ## Summing two buckets to generate final score
        # col_sum = []

        # for j in range(0,len0):
        #     num1 = int(col_percentageInc_rank[j])
        #     num2 = int(col_percentageErr_rank[j])
        #     sum = num1 + num2
        #     col_sum.append(sum)
        
        ## Adding a new column to final_result_summ2
        # final_result_summ2.loc[:,'Score'] = col_sum
        # final_result_summ2.loc[:,'BucketsInc'] = col_percentageInc_Bucket
        # final_result_summ2.loc[:,'BucketsPRE'] = col_percentageErr_Bucket

        list_SctterPlot_Inc = final_result_summ2['PercentageInc'].tolist()
        list_SctterPlot_FName = df['col_14'].tolist()
        list_SctterPlot_PRErr = final_result_summ2['PercentagePRErr'].tolist()

        # Deleting unnecessary columns
        del final_result_summ2['PercentageInc']
        del final_result_summ2['PercentagePRErr']

        def select_col(X):
            global cnt1, cnt2, cnt3, cnt4, cnt5, cnt6, cnt7, cnt8, cnt9, cnt10
            # COLORS
            c = ['background-color:  #C00000',              #>=50% RED
                    'background-color: #FFAF00',            #25-50% ORANGE
                        'background-color: #C0C000',        #10-25% YELLOW
                            'background-color: #00FF00',    #5-10% L GREEN
                                'background-color: #00AF5F',#<5% GREEN
                                    '']

            #compare columns
            # mask_5 = x['PercentilePRErr'] >= 0.50
            # mask_6 = (x['PercentilePRErr'] < 0.50) & (x['PercentilePRErr'] >= 0.25)
            # mask_7 = (x['PercentilePRErr'] < 0.25) & (x['PercentilePRErr'] >= 0.10)
            # mask_8 = (x['PercentilePRErr'] < 0.10) & (x['PercentilePRErr'] >= 0.05)
            # mask_9 = x['PercentilePRErr'] < 0.05


            # mask = x['PercentileInc'] >= 0.50
            # mask_1 = (x['PercentileInc'] < 0.50) & (x['PercentileInc'] >= 0.25)
            # mask_2 = (x['PercentileInc'] < 0.25) & (x['PercentileInc'] >= 0.10)
            # mask_3 = (x['PercentileInc'] < 0.10) & (x['PercentileInc'] >= 0.05)
            # mask_4 = x['PercentileInc'] < 0.05   
             
            mask_AllBlank = (X['All Blank'] == 'Yes')   

            mask_10 = (X['Inconsistent'] >= 5)
            cnt1 = mask_10.values.sum()
            mask_11 = (X['Inconsistent'] == 4)
            cnt2 = mask_11.values.sum()
            mask_12 = (X['Inconsistent'] == 3)
            cnt3 = mask_12.values.sum()
            mask_13 = (X['Inconsistent'] == 2)
            cnt4 = mask_13.values.sum()
            mask_14 = (X['Inconsistent'] <= 1)
            cnt5 = mask_14.values.sum()

            mask_5 = (X['Probable Reporting Error'] >= 5)
            cnt6 = mask_5.values.sum()
            mask_6 = (X['Probable Reporting Error'] == 4)
            cnt7 = mask_6.values.sum()
            mask_7 = (X['Probable Reporting Error'] == 3)
            cnt8 = mask_7.values.sum()
            mask_8 = (X['Probable Reporting Error'] == 2)
            cnt9 = mask_8.values.sum()
            mask_9 = (X['Probable Reporting Error'] <= 1)
            cnt10 = mask_9.values.sum()

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[5], X.index, columns=X.columns)

            #modify values of df1 column by boolean mask
            df1.loc[mask_5, 'Probable Reporting Error'] = c[0]
            df1.loc[mask_6, 'Probable Reporting Error'] = c[1]
            df1.loc[mask_7, 'Probable Reporting Error'] = c[2]
            df1.loc[mask_8, 'Probable Reporting Error'] = c[3]
            df1.loc[mask_9, 'Probable Reporting Error'] = c[4]

            df1.loc[mask_AllBlank, 'All Blank'] = c[0]
            df1.loc[mask_10, 'Inconsistent'] = c[0]
            df1.loc[mask_11, 'Inconsistent'] = c[1]
            df1.loc[mask_12, 'Inconsistent'] = c[2]
            df1.loc[mask_13, 'Inconsistent'] = c[3]
            df1.loc[mask_14, 'Inconsistent'] = c[4]

            return df1

        final_result_summ2 = final_result_summ2.style.apply(select_col, axis=None)


        # ## Third Summary Report
        # ## --------------------
        # summ3_countConsistent = []

        # # Iterating over indices of each row and calculating number of Consistent 
        # for index in range(len(df_SummReport)):
        #     '''   For no. of consistent   '''
        #     consistent = df_SummReport.iloc[index, :].str.count("consistent").sum()
        #     summ3_countConsistent.append(consistent)

        
        # final_result_summ3 = pd.DataFrame({ "Facility Name": df['col_14'].tolist(),
        #                                             "Consistent": summ3_countConsistent})

        # # For Health Sub Centre
        # if FType == 'Health Sub Centre':
        #     final_result_summ3['PercentageCon'] = final_result_summ3['Consistent']/31 * 100
            
        # # For Primary Health Centre
        # elif FType == 'Primary Health Centre':
        #     final_result_summ3['PercentageCon'] = final_result_summ3['Consistent']/77 * 100

        # len3 = len(final_result_summ3['PercentageCon'])

        # col_percentageCon_Bucket = []

        # for k in range(0,len3):
        #     primcol_Con = final_result_summ3['PercentageCon'][k]
        #     if primcol_Con < 5 :
        #         col_percentageCon_Bucket.append('< 5')
        #     elif primcol_Con >= 5 and primcol_Con < 10:
        #         col_percentageCon_Bucket.append('5 & < 10')
        #     elif primcol_Con >= 10 and primcol_Con < 25:
        #         col_percentageCon_Bucket.append('10 & < 25')
        #     elif primcol_Con >= 25 and primcol_Con < 50:
        #         col_percentageCon_Bucket.append('25 & < 50')
        #     elif primcol_Con >= 50:
        #         col_percentageCon_Bucket.append('>= 50')
        #     else:
        #         col_percentageCon_Bucket.append('NaN')

        # final_result_summ3.loc[:,'BucketsCon'] = col_percentageCon_Bucket

        # def select_col_Con(X):
        #     global cnt11, cnt12, cnt13, cnt14, cnt15
        #     # COLORS
        #     c = ['background-color:  #C00000',              #>=50% RED
        #             'background-color: #FFAF00',            #25-50% ORANGE
        #                 'background-color: #C0C000',        #10-25% YELLOW
        #                     'background-color: #00FF00',    #5-10% L GREEN
        #                         'background-color: #00AF5F',#<5% GREEN
        #                             '']

        #     #compare columns
        #     # mask_5 = x['PercentilePRErr'] >= 0.50
        #     # mask_6 = (x['PercentilePRErr'] < 0.50) & (x['PercentilePRErr'] >= 0.25)
        #     # mask_7 = (x['PercentilePRErr'] < 0.25) & (x['PercentilePRErr'] >= 0.10)
        #     # mask_8 = (x['PercentilePRErr'] < 0.10) & (x['PercentilePRErr'] >= 0.05)
        #     # mask_9 = x['PercentilePRErr'] < 0.05


        #     mask_15 = X['PercentageCon'] >= 50
        #     cnt11 = mask_15.values.sum()
        #     mask_16 = (X['PercentageCon'] < 50) & (X['PercentageCon'] >= 25)
        #     cnt12 = mask_16.values.sum()
        #     mask_17 = (X['PercentageCon'] < 25) & (X['PercentageCon'] >= 10)
        #     cnt13 = mask_17.values.sum()
        #     mask_18 = (X['PercentageCon'] < 10) & (X['PercentageCon'] >= 5)
        #     cnt14 = mask_18.values.sum()
        #     mask_19 = X['PercentageCon'] < 5  
        #     cnt15 = mask_19.values.sum()        

        #     #DataFrame with same index and columns names as original filled empty strings
        #     df1 =  pd.DataFrame(c[5], X.index, columns=X.columns)

        #     df1.loc[mask_15, 'Consistent'] = c[4]
        #     df1.loc[mask_16, 'Consistent'] = c[3]
        #     df1.loc[mask_17, 'Consistent'] = c[2]
        #     df1.loc[mask_18, 'Consistent'] = c[1]
        #     df1.loc[mask_19, 'Consistent'] = c[0]

        #     return df1

        # final_result_summ3 = final_result_summ3.style.apply(select_col_Con, axis=None)

        return final_result_summ1, final_result_summ2


    def export(self):
        #table_result1, table_result2, table_result3 = self.summaryReport()
        table_result1, table_result2 = self.summaryReport()

        # Rename orignal headers
        df.rename(res_dict , axis=1, inplace=True)

        # Save file dialog
        filename = QFileDialog.getSaveFileName(Dialog, "Save to Excel", "Summary_Sheet",
                                               "Excel Spreadsheet (*.xlsx);;"
                                               "All Files (*)")[0]


        with pd.ExcelWriter(filename) as writer:  # doctest: +SKIP
            df.to_excel(writer, sheet_name='Validated_Data')
            table_result1.to_excel(writer, sheet_name='Validation Summary Sheet', engine='openpyxl')
            table_result2.to_excel(writer, sheet_name='Facility Guidance Sheet', engine='openpyxl')
            #table_result3.to_excel(writer, sheet_name='Facility Guidance Sheet_2', engine='openpyxl')

        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.chart import BarChart, Reference, Series



        # PALETTES
        workbook = load_workbook(filename)
        sheet = workbook['Facility Guidance Sheet']
        sheet_1 = workbook['Validation Summary Sheet']

        # Activating sheets
        workbook.active = sheet
        workbook.active = sheet_1

        # sheet["G10"] = ">= 50%"
        # sheet["G11"] = "25 - 50%"
        # sheet["G12"] = "10 - 25%"
        # sheet["G13"] = "5 - 10%"
        # sheet["G14"] = "< 5%"

        sheet["H18"] = ">= 50%"
        sheet["H19"] = "25 - 50%"
        sheet["H20"] = "10 - 25%"
        sheet["H21"] = "5 - 10%"
        sheet["H22"] = "< 5%"
        sheet["H23"] = "Total Facilities"

        sheet["I17"] = "Inconsistent"
        sheet["I18"] = cnt1
        sheet["I19"] = cnt2
        sheet["I20"] = cnt3
        sheet["I21"] = cnt4
        sheet["I22"] = cnt5
        sheet["I23"] = cnt1 + cnt2 + cnt3 + cnt4 + cnt5

        sheet["J17"] = "Probable Reporting Error"
        sheet["J18"] = cnt6
        sheet["J19"] = cnt7
        sheet["J20"] = cnt8
        sheet["J21"] = cnt9
        sheet["J22"] = cnt10
        sheet["J23"] = cnt6 + cnt7 + cnt8 + cnt9 + cnt10

        sheet['G18'].fill = PatternFill(fgColor="C00000", fill_type = "solid")
        sheet['G19'].fill = PatternFill(fgColor="FFAF00", fill_type = "solid")
        sheet['G20'].fill = PatternFill(fgColor="C0C000", fill_type = "solid")
        sheet['G21'].fill = PatternFill(fgColor="00FF00", fill_type = "solid")
        sheet['G22'].fill = PatternFill(fgColor="00AF5F", fill_type = "solid")

        #sheet_1
        # sheet_1['I8'] = ">=25%"
        sheet_1['H11'].fill = PatternFill(fgColor="C00000", fill_type = "solid")

        sheet_1['I10'] = "Condition"
        sheet_1['I11'] = ">= 25%"

        sheet_1['J10'] = "Inconsistent"
        sheet_1['J11'] = cnt21

        sheet_1['K10'] = "Probable Reporting Error"
        sheet_1['K11'] = cnt22

        #PLOTTING GRAPH 1 (Sheet_1/validation check)
        # from matplotlib.pyplot import figure
        
        # Ranges = ['>= 25%']
        # Num_Inc = [cnt21]
        # Num_PRE = [cnt22]
        
        # # plotting a bar graph
        # Z_axis = np.arange(len(Ranges))

        # plt.bar(Z_axis - 0.2, Num_Inc, 0.4)
        # plt.bar(Z_axis + 0.2, Num_PRE, 0.4)
        # # figure(figsize=(8, 8), dpi=50)
        # plt.xticks(Z_axis, Ranges)
        # plt.xlabel("Condition   for checks")
        # plt.ylabel("Number")
        # plt.legend()
        # plt.title('Error Summary Facility Wise')
        # plt.savefig("myplot1.png", dpi = 80)

        # img = openpyxl.drawing.image.Image('myplot1.png')
        # img.anchor='N3'
        

        # sheet_1.add_image(img)
        # workbook.save(filename=filename)

        #PLOTTING GRAPH 2 (Sheet/Facility wise)
        from matplotlib.pyplot import figure
        
        Ranges = ['< 5%', '5 - 10%' , '10 - 25%', '25 - 50%', '>= 50%']
        Numbers_Inc = [cnt5, cnt4, cnt3, cnt2, cnt1]
        Numbers_PRE = [cnt10, cnt9, cnt8, cnt7, cnt6]
        
        # plotting a bar graph
        X_axis = np.arange(len(Ranges))

        plt.bar(X_axis - 0.2, Numbers_Inc, 0.4, label = 'Number of Inconsistents')
        plt.bar(X_axis + 0.2, Numbers_PRE, 0.4, label = 'Number of Probable Reporting Errors')
        # figure(figsize=(8, 8), dpi=50)
        plt.xticks(X_axis, Ranges)
        plt.xlabel("Condition")
        plt.ylabel("Number")
        plt.legend()
        plt.title('Error Summary Facility Wise')
        plt.savefig("myplot2.png", dpi = 80)

        img = openpyxl.drawing.image.Image('myplot2.png')
        img.anchor='M3'
        

        sheet.add_image(img)
        workbook.save(filename=filename)

        # Graph 3 Scatter plot for sheet_3
        
        ele = pd.DataFrame({ 
            'Facility':list_SctterPlot_FName,
            'INCONSISTENT': list_SctterPlot_Inc
        })
        ele.plot(x='Facility', y='INCONSISTENT', style='o')

        plt.xticks([])

        plt.title('Scatter plot for Inconsistent(Facility wise)')
        plt.savefig("myplot_3.png", dpi = 80)

        img = openpyxl.drawing.image.Image('myplot_3.png')
        img.anchor='M28'
        

        sheet.add_image(img)

        workbook.save(filename=filename)

        # Graph 4 Scatter plot for sheet_4
        
        ele = pd.DataFrame({ 
            'Facility':list_SctterPlot_FName,
            'Probable Reporting Error': list_SctterPlot_PRErr
        })
        ele.plot(x='Facility', y='Probable Reporting Error', style='o')

        plt.xticks([])

        plt.title('Scatter plot for Probable Reporting Error(Facility wise)')
        plt.savefig("myplot_4.png", dpi = 80)

        img = openpyxl.drawing.image.Image('myplot_4.png')
        img.anchor='V28'
        

        sheet.add_image(img)

        workbook.save(filename=filename)

        msg = QMessageBox()
        msg.setWindowTitle("Message")
        msg.setText(("{}\n \n\n exported to the file location you have chosen".format(filename)))
        msg.exec()

    def reset(self):
        self.comboBox.clear()
        self.pushButton.setEnabled(True)
        products = {'_': [' ']}
        blank_df = pd.DataFrame(products, columns=['_'])
        self.tableView.setModel(PandasModel(blank_df))


################################### Main Function ############################
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    sys.exit(app.exec_())